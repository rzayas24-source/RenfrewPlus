#!/usr/bin/env python3

import argparse
import json
import os
import re
from datetime import datetime
from pathlib import Path
import sys

from db import get_conn
from pdf2image import convert_from_path
from PIL import Image, ImageDraw, ImageFont
import xlrd
from openpyxl import load_workbook

EMAIL_FOLDER = r"C:\Renfrew\Workflow\4.Emails"
SNAPSHOT_FOLDER = r"C:\Renfrew\Workflow\snapshots"
POPPLER_CANDIDATES = [
    r"C:\Tools\poppler\Library\bin",
    r"C:\poppler\Library\bin",
]
BATCH_PREFIX_RE = re.compile(r"^(?P<batch_id>\d{2}\.\d{2}\.\d{2})-")

os.makedirs(EMAIL_FOLDER, exist_ok=True)
os.makedirs(SNAPSHOT_FOLDER, exist_ok=True)

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")


def _poppler_path():
    for candidate in POPPLER_CANDIDATES:
        if os.path.exists(candidate):
            return candidate
    return None


def _safe_font(size: int):
    candidates = [
        r"C:\Windows\Fonts\arial.ttf",
        r"C:\Windows\Fonts\segoeui.ttf",
    ]
    for candidate in candidates:
        if os.path.exists(candidate):
            try:
                return ImageFont.truetype(candidate, size=size)
            except Exception:
                pass
    return ImageFont.load_default()


def _wrap_text(draw, text, font, max_width):
    lines = []
    for raw_line in str(text).splitlines() or [""]:
        words = raw_line.split()
        if not words:
            lines.append("")
            continue

        current = words[0]
        for word in words[1:]:
            trial = f"{current} {word}"
            if draw.textbbox((0, 0), trial, font=font)[2] <= max_width:
                current = trial
            else:
                lines.append(current)
                current = word
        lines.append(current)
    return lines


def _draw_text_snapshot(lines, out_path, title, subtitle=None):
    width = 1400
    padding = 48
    title_font = _safe_font(34)
    body_font = _safe_font(24)
    small_font = _safe_font(20)

    temp = Image.new("RGB", (width, 1200), "white")
    draw = ImageDraw.Draw(temp)

    rendered_lines = []
    if title:
        rendered_lines.extend(_wrap_text(draw, title, title_font, width - padding * 2))
        rendered_lines.append("")
    if subtitle:
        rendered_lines.extend(_wrap_text(draw, subtitle, small_font, width - padding * 2))
        rendered_lines.append("")
    for line in lines:
        rendered_lines.extend(_wrap_text(draw, line, body_font, width - padding * 2))

    line_height = 34
    if rendered_lines:
        line_height = max(line_height, draw.textbbox((0, 0), "Ag", font=body_font)[3] + 10)
    height = max(420, padding * 2 + len(rendered_lines) * line_height + 20)

    image = Image.new("RGB", (width, height), "#ffffff")
    draw = ImageDraw.Draw(image)
    draw.rounded_rectangle((24, 24, width - 24, height - 24), radius=28, outline="#d7dde6", width=3, fill="#fbfcfe")

    y = padding
    if title:
        draw.text((padding, y), title, fill="#1f2933", font=title_font)
        y += draw.textbbox((0, 0), "Ag", font=title_font)[3] + 18
    if subtitle:
        draw.text((padding, y), subtitle, fill="#4b5563", font=small_font)
        y += draw.textbbox((0, 0), "Ag", font=small_font)[3] + 24

    for line in rendered_lines:
        draw.text((padding, y), line, fill="#1f2933", font=body_font)
        y += line_height
        if y > height - padding:
            break

    image.save(out_path, "PNG")
    return out_path


def _image_snapshot(image_path, out_path):
    with Image.open(image_path) as source:
        source = source.convert("RGB")
        source.thumbnail((1600, 1600))
        canvas = Image.new("RGB", (max(1600, source.width + 120), max(1200, source.height + 120)), "#ffffff")
        offset_x = (canvas.width - source.width) // 2
        offset_y = (canvas.height - source.height) // 2
        canvas.paste(source, (offset_x, offset_y))
        canvas.save(out_path, "PNG")
    return out_path


def _pdf_snapshot(pdf_path, out_path):
    poppler_path = _poppler_path()
    pages = convert_from_path(
        pdf_path,
        dpi=150,
        first_page=1,
        last_page=1,
        poppler_path=poppler_path,
    )
    pages[0].save(out_path, "PNG")
    return out_path


def _txt_snapshot(file_path, out_path, filename):
    with open(file_path, "r", encoding="utf-8", errors="replace") as handle:
        content = handle.read().strip()

    lines = content.splitlines()[:60]
    if not lines:
        lines = ["(empty text file)"]
    return _draw_text_snapshot(
        lines,
        out_path,
        title=filename,
        subtitle="Text attachment preview",
    )


def _excel_snapshot(file_path, out_path, filename):
    lower_name = filename.lower()
    if lower_name.endswith(".xls"):
        workbook = xlrd.open_workbook(file_path)
        sheet = workbook.sheet_by_index(0)
        lines = [f"Sheet: {sheet.name}", ""]
        for row_index in range(min(sheet.nrows, 25)):
            row = sheet.row_values(row_index)[:8]
            values = ["" if value in (None, "") else str(value) for value in row]
            values = [value for value in values if value != ""]
            lines.append(" | ".join(values) if values else "(blank row)")
        if sheet.nrows > 25:
            lines.append("...")
        return _draw_text_snapshot(
            lines,
            out_path,
            title=filename,
            subtitle="Spreadsheet preview",
        )

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        lines = [f"Sheet: {sheet.title}", ""]
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_index > 25:
                lines.append("...")
                break
            values = ["" if value is None else str(value) for value in row[:8]]
            values = [value for value in values if value != ""]
            lines.append(" | ".join(values) if values else "(blank row)")
        return _draw_text_snapshot(
            lines,
            out_path,
            title=filename,
            subtitle="Spreadsheet preview",
        )
    finally:
        workbook.close()


def _generic_snapshot(file_path, out_path, filename):
    size = os.path.getsize(file_path)
    modified = datetime.fromtimestamp(os.path.getmtime(file_path)).strftime("%Y-%m-%d %H:%M:%S")
    lines = [
        f"Filename: {filename}",
        f"Path: {file_path}",
        f"Size: {size} bytes",
        f"Modified: {modified}",
    ]
    return _draw_text_snapshot(lines, out_path, title="Attachment preview", subtitle="Generic file snapshot")


def _batch_info_from_filename(filename):
    match = BATCH_PREFIX_RE.match(filename)
    if not match:
        return None, None

    batch_id = match.group("batch_id")
    try:
        batch_date = datetime.strptime(batch_id, "%m.%d.%y").strftime("%Y-%m-%d")
    except ValueError:
        batch_date = batch_id
    return batch_id, batch_date


def ensure_db_row(cur, filename, batch_id=None, batch_date=None):
    cur.execute("SELECT id FROM imported_files WHERE filename = ?", (filename,))
    row = cur.fetchone()

    if row:
        if batch_id or batch_date:
            updates = []
            params = []
            if batch_id:
                updates.append("batch_id = ?")
                params.append(batch_id)
            if batch_date:
                updates.append("batch_date = ?")
                params.append(batch_date)
            if updates:
                params.append(row[0])
                cur.execute(
                    f"UPDATE imported_files SET {', '.join(updates)} WHERE id = ?",
                    params,
                )
        return row[0]

    cur.execute(
        """
        INSERT INTO imported_files (filename, review_status, source_type, batch_id, batch_date)
        VALUES (?, 'Pending', 'email', ?, ?)
        """,
        (filename, batch_id, batch_date),
    )
    return cur.lastrowid


def _snapshot_for_file(file_path, snapshot_path, filename):
    lower_name = filename.lower()
    if lower_name.endswith(".pdf"):
        return _pdf_snapshot(file_path, snapshot_path)
    if lower_name.endswith((".png", ".jpg", ".jpeg", ".bmp", ".gif", ".tif", ".tiff")):
        return _image_snapshot(file_path, snapshot_path)
    if lower_name.endswith(".txt"):
        return _txt_snapshot(file_path, snapshot_path, filename)
    if lower_name.endswith((".xls", ".xlsx")):
        return _excel_snapshot(file_path, snapshot_path, filename)
    return _generic_snapshot(file_path, snapshot_path, filename)


def process_folder_files():
    conn = get_conn()
    cur = conn.cursor()

    files = [
        f for f in os.listdir(EMAIL_FOLDER)
        if os.path.isfile(os.path.join(EMAIL_FOLDER, f))
    ]

    if not files:
        print("No email downloads found in the folder.")
        conn.close()
        return {
            "processed_count": 0,
            "generated_count": 0,
            "skipped_count": 0,
            "files": [],
        }

    total = len(files)
    print(f"\nFound {total} file(s) in folder.\n")
    generated_files = []
    skipped_count = 0

    for index, filename in enumerate(sorted(files), start=1):
        print(f"[{index}/{total}] Processing {filename}")

        file_path = os.path.join(EMAIL_FOLDER, filename)
        batch_id, batch_date = _batch_info_from_filename(filename)
        file_id = ensure_db_row(cur, filename, batch_id=batch_id, batch_date=batch_date)
        snapshot_path = os.path.join(SNAPSHOT_FOLDER, f"{file_id}.png")

        if os.path.exists(snapshot_path):
            print(f"   Snapshot already exists: {snapshot_path}")
            skipped_count += 1
            generated_files.append(snapshot_path)
            continue

        try:
            print("   Generating snapshot...")
            _snapshot_for_file(file_path, snapshot_path, filename)

            cur.execute(
                """
                UPDATE imported_files
                SET snapshot_path = ?, review_status = 'Pending'
                WHERE id = ?
                """,
                (snapshot_path, file_id),
            )

            print(f"   Snapshot saved: {snapshot_path}")
            generated_files.append(snapshot_path)

        except Exception as exc:
            print(f"Error generating snapshot for {filename}: {exc}")

    conn.commit()
    conn.close()
    print("\nAll folder-based snapshot processing complete.")
    return {
        "processed_count": total,
        "generated_count": len(generated_files),
        "skipped_count": skipped_count,
        "files": generated_files,
    }


def _write_json(path, payload):
    Path(path).write_text(json.dumps(payload, indent=2), encoding="utf-8")


def main(argv=None):
    parser = argparse.ArgumentParser(description="Snapshot generator helper")
    subparsers = parser.add_subparsers(dest="command")

    run_parser = subparsers.add_parser("run")
    run_parser.add_argument("--output", required=True)

    args = parser.parse_args(argv)

    if args.command == "run":
        result = process_folder_files()
        _write_json(args.output, result)
        return 0

    process_folder_files()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
