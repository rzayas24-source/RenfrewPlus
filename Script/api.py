import os
import shutil
import sqlite3
import re
import json
import html as html_lib
import tempfile
import zipfile
import subprocess
import sys
import uuid
import hashlib
import hmac
import secrets
from functools import lru_cache
from io import BytesIO, StringIO
from datetime import datetime
from collections import defaultdict
from pathlib import Path
from urllib.parse import quote
from xml.etree import ElementTree as ET
from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from openpyxl import load_workbook

from config_manager import CONFIG_PATH, load_config, resolve_path, save_config
from system_calendar_core import (
    add_days,
    advance_current_work_day,
    build_from,
    delete_days,
    get_current_bank_day,
    get_current_work_day,
    init_db,
    normalize_mmddyyyy,
    set_current_work_day,
    setup,
)
from source_table_schema import (
    ensure_edi_manifest_tables,
    ensure_eft_tables,
    ensure_eftload_schema,
    ensure_source_table_columns,
    refresh_source_table_mirrors,
)
from system_source_match_core import build_match_dashboard, build_match_history, commit_all_strong_matches, commit_match, ensure_match_indexes, get_match_detail, normalize_checknum
from system_banking_core import build_banking_spreadsheet
import pandas as pd

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
BACKEND_CONFIG = load_config()
WORKFLOW_ROOT = ""
DB_PATH = ""
ZIP_835_TRN_FOLDER = ""
ZIP_835_ERA_FOLDER = ""
ZIP_835_HTML_FOLDER = ""
ZIP_835_ERA_PROCESSING_FOLDER = ""
ZIP_835_HTML_PROCESSING_FOLDER = ""
EDI_PENDING_ROOT = ""
ZIP_835_TRN_ARCHIVE_FOLDER = ""
FLYWIRE_STORAGE_ROOT = ""
FLYWIRE_UPLOAD_FOLDER = ""
EMAILS_FOLDER = ""
SNAPSHOTS_FOLDER = ""
EMAIL_DOWNLOADER_SCRIPT = os.path.join(BASE_DIR, "site_emaildownloader.py")
SNAPSHOT_GENERATOR_SCRIPT = os.path.join(BASE_DIR, "site_snapshotgenerator.py")
PYTHON_EXECUTABLE = sys.executable


def refresh_runtime_config(config: dict | None = None):
    global BACKEND_CONFIG
    global WORKFLOW_ROOT
    global DB_PATH
    global ZIP_835_TRN_FOLDER
    global ZIP_835_ERA_FOLDER
    global ZIP_835_HTML_FOLDER
    global ZIP_835_ERA_PROCESSING_FOLDER
    global ZIP_835_HTML_PROCESSING_FOLDER
    global EDI_PENDING_ROOT
    global ZIP_835_TRN_ARCHIVE_FOLDER
    global FLYWIRE_STORAGE_ROOT
    global FLYWIRE_UPLOAD_FOLDER
    global EMAILS_FOLDER
    global SNAPSHOTS_FOLDER

    BACKEND_CONFIG = config or load_config(force=True)
    config_dir = Path(CONFIG_PATH).resolve().parent
    workflow_root = resolve_path(BACKEND_CONFIG, "workflow_root", Path(BASE_DIR).parent, relative_to=config_dir)
    WORKFLOW_ROOT = workflow_root
    DB_PATH = resolve_path(BACKEND_CONFIG, "db_path", Path(WORKFLOW_ROOT) / "database.db", relative_to=Path(WORKFLOW_ROOT))
    ZIP_835_TRN_FOLDER = resolve_path(BACKEND_CONFIG, "trn_folder", Path(WORKFLOW_ROOT) / "1.TRN", relative_to=Path(WORKFLOW_ROOT))
    ZIP_835_ERA_FOLDER = resolve_path(BACKEND_CONFIG, "era_folder", Path(WORKFLOW_ROOT) / "2.ERA", relative_to=Path(WORKFLOW_ROOT))
    ZIP_835_HTML_FOLDER = resolve_path(BACKEND_CONFIG, "html_folder", Path(WORKFLOW_ROOT) / "3.HTML", relative_to=Path(WORKFLOW_ROOT))
    ZIP_835_ERA_PROCESSING_FOLDER = os.path.join(ZIP_835_ERA_FOLDER, "Processing")
    ZIP_835_HTML_PROCESSING_FOLDER = os.path.join(ZIP_835_HTML_FOLDER, "Processing")
    EDI_PENDING_ROOT = os.path.join(WORKFLOW_ROOT, "EDI_Pending")
    ZIP_835_TRN_ARCHIVE_FOLDER = os.path.join(ZIP_835_TRN_FOLDER, "Loaded")
    FLYWIRE_STORAGE_ROOT = resolve_path(
        BACKEND_CONFIG,
        "flywire_storage_root",
        Path(WORKFLOW_ROOT) / "Import_Flywire",
        relative_to=Path(WORKFLOW_ROOT),
    )
    FLYWIRE_UPLOAD_FOLDER = os.path.join(FLYWIRE_STORAGE_ROOT, "Uploads")
    EMAILS_FOLDER = resolve_path(BACKEND_CONFIG, "emails_folder", Path(WORKFLOW_ROOT) / "4.Emails", relative_to=Path(WORKFLOW_ROOT))
    SNAPSHOTS_FOLDER = resolve_path(
        BACKEND_CONFIG,
        "snapshots_folder",
        Path(WORKFLOW_ROOT) / "snapshots",
        relative_to=Path(WORKFLOW_ROOT),
    )

    for path in [
        WORKFLOW_ROOT,
        DB_PATH,
        ZIP_835_TRN_FOLDER,
        ZIP_835_ERA_FOLDER,
        ZIP_835_HTML_FOLDER,
        ZIP_835_ERA_PROCESSING_FOLDER,
        ZIP_835_HTML_PROCESSING_FOLDER,
        EDI_PENDING_ROOT,
        FLYWIRE_STORAGE_ROOT,
        EMAILS_FOLDER,
        SNAPSHOTS_FOLDER,
    ]:
        if path and not os.path.splitext(path)[1]:
            os.makedirs(path, exist_ok=True)


refresh_runtime_config(BACKEND_CONFIG)
print(f"[startup] pid={os.getpid()} db_path={DB_PATH} workflow_root={WORKFLOW_ROOT}", flush=True)

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def _ensure_source_table_columns_on_startup():
    conn = get_conn()
    try:
        ensure_imported_files_table(conn)
        backfill_imported_file_batches(conn)
        ensure_keyproof_table(conn)
        ensure_itemization_table(conn)
        ensure_flywire_tables(conn)
        ensure_source_table_columns(conn)
        ensure_edi_manifest_tables(conn)
        ensure_eft_tables(conn)
        ensure_match_indexes(conn)
        ensure_balsheet_notes_table(conn)
        ensure_imaging_tables(conn)
        ensure_misc_table(conn)
        ensure_tasks_table(conn)
        ensure_auth_tables(conn)
        ensure_menu_table(conn)
        normalize_tasks_table_categories(conn)
        refresh_source_table_mirrors(conn)
    finally:
        conn.close()

def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA busy_timeout = 30000")
    return conn


@app.get("/config")
def get_config():
    return load_config()


@app.put("/config")
def put_config(payload: dict):
    if not isinstance(payload, dict):
        raise HTTPException(status_code=400, detail="Config payload must be an object.")

    saved = save_config(payload)
    refresh_runtime_config(saved)
    return saved


@app.get("/health")
def get_health():
    return {
        "status": "ok",
        "pid": os.getpid(),
        "db_path": DB_PATH,
        "workflow_root": WORKFLOW_ROOT,
    }


def _quote_identifier(name: str) -> str:
    return '"' + name.replace('"', '""') + '"'


_BATCH_PREFIX_RE = re.compile(r"^(?P<batch_id>\d{2}\.\d{2}\.\d{2})-")


def _batch_info_from_filename(filename: str):
    match = _BATCH_PREFIX_RE.match(filename or "")
    if not match:
        return None, None

    batch_id = match.group("batch_id")
    try:
        batch_date = datetime.strptime(batch_id, "%m.%d.%y").strftime("%Y-%m-%d")
    except ValueError:
        batch_date = batch_id
    return batch_id, batch_date


IMPORTED_FILES_TABLE_COLUMNS = [
    ("id", "INTEGER PRIMARY KEY"),
    ("filename", "TEXT"),
    ("moved_from", "TEXT"),
    ("moved_to", "TEXT"),
    ("archived_to", "TEXT"),
    ("processed_at", "TEXT"),
    ("site", "TEXT"),
    ("detail", "TEXT"),
    ("amount", "REAL"),
    ("snapshot_path", "TEXT"),
    ("review_status", "TEXT DEFAULT 'Pending'"),
    ("reviewer", "TEXT"),
    ("review_notes", "TEXT"),
    ("batch_id", "TEXT"),
    ("batch_date", "TEXT"),
    ("email_id", "TEXT"),
    ("original_filename", "TEXT"),
    ("download_notes", "TEXT"),
    ("source_type", "TEXT"),
]


def ensure_imported_files_table(conn=None):
    close_conn = False
    if conn is None:
        conn = get_conn()
        close_conn = True

    cur = conn.cursor()
    column_defs = ", ".join(
        f'{_quote_identifier(name)} {definition}' for name, definition in IMPORTED_FILES_TABLE_COLUMNS
    )
    cur.execute(f'CREATE TABLE IF NOT EXISTS {_quote_identifier("imported_files")} ({column_defs})')

    existing_columns = {
        row[1].lower()
        for row in cur.execute(f'PRAGMA table_info({_quote_identifier("imported_files")})').fetchall()
    }
    for column_name, column_type in IMPORTED_FILES_TABLE_COLUMNS:
        if column_name.lower() in existing_columns:
            continue
        cur.execute(
            f'ALTER TABLE {_quote_identifier("imported_files")} ADD COLUMN {_quote_identifier(column_name)} {column_type}'
        )

    conn.commit()

    if close_conn:
        conn.close()


def backfill_imported_file_batches(conn=None):
    close_conn = False
    if conn is None:
        conn = get_conn()
        close_conn = True

    cur = conn.cursor()
    rows = cur.execute(
        """
        SELECT id, filename, batch_id, batch_date
        FROM imported_files
        WHERE filename IS NOT NULL AND filename != ''
        """
    ).fetchall()

    changed = False
    for row_id, filename, current_batch_id, current_batch_date in rows:
        batch_id, batch_date = _batch_info_from_filename(filename)
        if not batch_id and not batch_date:
            continue

        updates = []
        params = []
        if batch_id and current_batch_id != batch_id:
            updates.append("batch_id = ?")
            params.append(batch_id)
        if batch_date and current_batch_date != batch_date:
            updates.append("batch_date = ?")
            params.append(batch_date)

        if not updates:
            continue

        params.append(row_id)
        cur.execute(
            f'UPDATE {_quote_identifier("imported_files")} SET {", ".join(updates)} WHERE id = ?',
            params,
        )
        changed = True

    if changed:
        conn.commit()

    if close_conn:
        conn.close()


FLYWIRE_DOCUMENTS_TABLE_COLUMNS = [
    ("id", "INTEGER PRIMARY KEY"),
    ("attachment_id", "INTEGER NOT NULL"),
    ("attachment_filename", "TEXT"),
    ("batch_id", "TEXT"),
    ("batch_date", "TEXT"),
    ("source_filename", "TEXT"),
    ("stored_filename", "TEXT"),
    ("stored_path", "TEXT"),
    ("sheet_name", "TEXT"),
    ("row_count", "INTEGER"),
    ("total_amount", "REAL"),
    ("summary_json", "TEXT"),
    ("created_at", "TEXT"),
    ("updated_at", "TEXT"),
]

FLYWIRE_ROWS_TABLE_COLUMNS = [
    ("id", "INTEGER PRIMARY KEY"),
    ("document_id", "INTEGER NOT NULL"),
    ("position", "INTEGER"),
    ("location", "TEXT"),
    ("department", "TEXT"),
    ("payment_method", "TEXT"),
    ("payment_type", "TEXT"),
    ("time_text", "TEXT"),
    ("amount", "REAL"),
    ("flywire_id", "TEXT"),
    ("account_number", "TEXT"),
    ("patient_name", "TEXT"),
    ("billing_name", "TEXT"),
    ("application", "TEXT"),
    ("raw_json", "TEXT"),
]


def ensure_flywire_tables(conn=None):
    close_conn = False
    if conn is None:
        conn = get_conn()
        close_conn = True

    os.makedirs(FLYWIRE_UPLOAD_FOLDER, exist_ok=True)

    cur = conn.cursor()

    document_column_defs = ", ".join(
        f'{_quote_identifier(name)} {definition}' for name, definition in FLYWIRE_DOCUMENTS_TABLE_COLUMNS
    )
    row_column_defs = ", ".join(f'{_quote_identifier(name)} {definition}' for name, definition in FLYWIRE_ROWS_TABLE_COLUMNS)

    cur.execute(f'CREATE TABLE IF NOT EXISTS {_quote_identifier("Import_FlywireDocuments")} ({document_column_defs})')
    cur.execute(f'CREATE TABLE IF NOT EXISTS {_quote_identifier("Import_FlywireRows")} ({row_column_defs})')
    cur.execute(
        f'CREATE UNIQUE INDEX IF NOT EXISTS idx_import_flywire_documents_attachment_id '
        f'ON {_quote_identifier("Import_FlywireDocuments")} ({_quote_identifier("attachment_id")})'
    )
    cur.execute(
        f'CREATE INDEX IF NOT EXISTS idx_import_flywire_rows_document_id '
        f'ON {_quote_identifier("Import_FlywireRows")} ({_quote_identifier("document_id")}, {_quote_identifier("position")})'
    )

    existing_document_columns = {
        row[1].lower()
        for row in cur.execute(f'PRAGMA table_info({_quote_identifier("Import_FlywireDocuments")})').fetchall()
    }
    existing_row_columns = {
        row[1].lower()
        for row in cur.execute(f'PRAGMA table_info({_quote_identifier("Import_FlywireRows")})').fetchall()
    }

    for column_name, column_type in FLYWIRE_DOCUMENTS_TABLE_COLUMNS:
        if column_name.lower() in existing_document_columns:
            continue
        cur.execute(
            f'ALTER TABLE {_quote_identifier("Import_FlywireDocuments")} ADD COLUMN {_quote_identifier(column_name)} {column_type}'
        )

    for column_name, column_type in FLYWIRE_ROWS_TABLE_COLUMNS:
        if column_name.lower() in existing_row_columns:
            continue
        cur.execute(
            f'ALTER TABLE {_quote_identifier("Import_FlywireRows")} ADD COLUMN {_quote_identifier(column_name)} {column_type}'
        )

    conn.commit()

    if close_conn:
        conn.close()


SAVED_ATTACHMENT_TABLE_COLUMNS = [
    ("attachment_id", "INTEGER PRIMARY KEY REFERENCES imported_files(id) ON DELETE CASCADE"),
    ("payload_json", "TEXT NOT NULL"),
    ("created_at", "TEXT NOT NULL"),
    ("updated_at", "TEXT NOT NULL"),
]


def _ensure_saved_attachment_table(conn, table_name: str):
    cur = conn.cursor()
    required_columns = [name.lower() for name, _ in SAVED_ATTACHMENT_TABLE_COLUMNS]
    existing_tables = {
        row[0].lower()
        for row in cur.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()
    }

    if table_name.lower() in existing_tables:
        existing_columns = [
            row[1].lower()
            for row in cur.execute(f'PRAGMA table_info({_quote_identifier(table_name)})').fetchall()
        ]
        has_saved_schema = all(column in existing_columns for column in required_columns)
        if not has_saved_schema:
            legacy_name = f"{table_name}_legacy"
            suffix = 1
            while legacy_name.lower() in existing_tables:
                suffix += 1
                legacy_name = f"{table_name}_legacy_{suffix}"
            cur.execute(
                f'ALTER TABLE {_quote_identifier(table_name)} RENAME TO {_quote_identifier(legacy_name)}'
            )
            existing_tables.add(legacy_name.lower())

    column_defs = ", ".join(f'{_quote_identifier(name)} {definition}' for name, definition in SAVED_ATTACHMENT_TABLE_COLUMNS)
    cur.execute(f'CREATE TABLE IF NOT EXISTS {_quote_identifier(table_name)} ({column_defs})')


def ensure_keyproof_table(conn=None):
    close_conn = False
    if conn is None:
        conn = get_conn()
        close_conn = True

    _ensure_saved_attachment_table(conn, "keyproof")
    conn.commit()

    if close_conn:
        conn.close()


def ensure_itemization_table(conn=None):
    close_conn = False
    if conn is None:
        conn = get_conn()
        close_conn = True

    _ensure_saved_attachment_table(conn, "itemization")
    conn.commit()

    if close_conn:
        conn.close()


def _ensure_attachment_exists(conn, attachment_id: int):
    row = conn.execute(
        f'SELECT {_quote_identifier("id")} FROM {_quote_identifier("imported_files")} WHERE {_quote_identifier("id")} = ?',
        (attachment_id,),
    ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Attachment not found")


def _load_saved_payload(conn, table_name: str, attachment_id: int):
    row = conn.execute(
        f'''
        SELECT *
        FROM {_quote_identifier(table_name)}
        WHERE {_quote_identifier("attachment_id")} = ?
        ''',
        (attachment_id,),
    ).fetchone()

    if not row:
        return {
            "attachment_id": attachment_id,
            "payload": None,
            "created_at": None,
            "updated_at": None,
        }

    try:
        payload = json.loads(row["payload_json"])
    except Exception:
        payload = None

    return {
        "attachment_id": attachment_id,
        "payload": payload,
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


def _save_saved_payload(conn, table_name: str, attachment_id: int, payload: dict):
    now = datetime.now().isoformat(timespec="seconds")
    payload_json = json.dumps(payload, ensure_ascii=False)
    existing = conn.execute(
        f'''
        SELECT {_quote_identifier("created_at")}
        FROM {_quote_identifier(table_name)}
        WHERE {_quote_identifier("attachment_id")} = ?
        ''',
        (attachment_id,),
    ).fetchone()
    created_at = existing[0] if existing else now

    conn.execute(
        f'''
        INSERT INTO {_quote_identifier(table_name)} (
            {_quote_identifier("attachment_id")},
            {_quote_identifier("payload_json")},
            {_quote_identifier("created_at")},
            {_quote_identifier("updated_at")}
        ) VALUES (?, ?, ?, ?)
        ON CONFLICT({_quote_identifier("attachment_id")}) DO UPDATE SET
            {_quote_identifier("payload_json")} = excluded.{_quote_identifier("payload_json")},
            {_quote_identifier("updated_at")} = excluded.{_quote_identifier("updated_at")}
        ''',
        (attachment_id, payload_json, created_at, now),
    )
    conn.commit()
    return _load_saved_payload(conn, table_name, attachment_id)


def _delete_saved_payload(conn, table_name: str, attachment_id: int):
    conn.execute(
        f'''
        DELETE FROM {_quote_identifier(table_name)}
        WHERE {_quote_identifier("attachment_id")} = ?
        ''',
        (attachment_id,),
    )
    conn.commit()
    return {"ok": True, "attachment_id": attachment_id}


def _keyproof_total_from_payload(payload):
    if not isinstance(payload, dict):
        return 0.0

    form = payload.get("form")
    if not isinstance(form, dict):
        return 0.0

    total = 0.0
    for field in ("cash", "check", "creditCard", "foreignCheck", "wireTransfer", "misc"):
        total += _parse_amount(form.get(field))
    return round(total, 2)


def _normalize_flywire_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    text = str(value).strip()
    return text


def _normalize_flywire_amount(value):
    if value in (None, ""):
        return None

    if isinstance(value, (int, float)) and not pd.isna(value):
        return float(value)

    try:
        parsed = float(str(value).replace("$", "").replace(",", "").strip())
    except Exception:
        return None

    return parsed


def _normalize_flywire_json_value(value):
    if value is None:
        return None
    if isinstance(value, float) and pd.isna(value):
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")

    try:
        return value.item()
    except Exception:
        return str(value)


def _normalize_flywire_column_name(value, index):
    text = str(value or "").strip()
    if not text or text.startswith("Unnamed:"):
        return f"column_{index}"
    return text


def _load_flywire_workbook(file_bytes: bytes, filename: str):
    extension = os.path.splitext(filename or "")[1].lower()
    if extension in {".xlsx", ".xlsm"}:
        engine = "openpyxl"
    elif extension == ".xls":
        engine = "xlrd"
    else:
        raise HTTPException(status_code=400, detail="Please upload a Fly Wire .xls or .xlsx workbook")

    try:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception:
        workbook = None

    if workbook is not None:
        worksheet = workbook[workbook.sheetnames[0]]
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(status_code=400, detail="Fly Wire workbook is empty")

        headers = [_normalize_flywire_column_name(value, index + 1) for index, value in enumerate(rows[0])]
        data_rows = []
        for raw_row in rows[1:]:
            row_map = {
                headers[index]: _normalize_flywire_json_value(value)
                for index, value in enumerate(raw_row[: len(headers)])
            }
            data_rows.append(row_map)

        return worksheet.title, data_rows

    try:
        dataframe = pd.read_excel(BytesIO(file_bytes), sheet_name=0, engine=engine)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Unable to read Fly Wire workbook: {exc}") from exc

    if dataframe.empty:
        raise HTTPException(status_code=400, detail="Fly Wire workbook is empty")

    dataframe.columns = [_normalize_flywire_column_name(value, index + 1) for index, value in enumerate(dataframe.columns)]
    data_rows = dataframe.where(pd.notna(dataframe), None).to_dict(orient="records")
    sheet_name = str(getattr(dataframe, "attrs", {}).get("sheet_name") or "Sheet1")
    return sheet_name, data_rows


def _flywire_row_is_data(row_map: dict):
    text_values = [str(value).strip() for value in row_map.values() if value not in (None, "")]
    if not text_values:
        return False

    upper_values = [value.upper() for value in text_values]
    if any(value == "TOTAL:" for value in upper_values):
        return False
    if any(value.startswith("AR COLLECTION:") for value in upper_values):
        return False

    meaningful_keys = (
        "location",
        "department",
        "payment method",
        "type",
        "time",
        "amount",
        "id",
        "account #",
        "patient name",
        "billing name",
        "application",
    )
    lower_map = {str(key).strip().lower(): value for key, value in row_map.items()}
    if any(_normalize_flywire_text(lower_map.get(key)) for key in meaningful_keys):
        return True

    amount = _normalize_flywire_amount(lower_map.get("amount"))
    return amount not in (None, 0)


def _row_dict_for_flywire(row_map: dict):
    lower_map = {str(key).strip().lower(): value for key, value in row_map.items()}
    amount = _normalize_flywire_amount(lower_map.get("amount"))
    normalized_row = {
        "location": _normalize_flywire_text(lower_map.get("location")),
        "department": _normalize_flywire_text(lower_map.get("department")),
        "payment_method": _normalize_flywire_text(lower_map.get("payment method")),
        "payment_type": _normalize_flywire_text(lower_map.get("type")),
        "time_text": _normalize_flywire_text(lower_map.get("time")),
        "amount": amount,
        "flywire_id": _normalize_flywire_text(lower_map.get("id")),
        "account_number": _normalize_flywire_text(lower_map.get("account #") or lower_map.get("account number")),
        "patient_name": _normalize_flywire_text(lower_map.get("patient name")),
        "billing_name": _normalize_flywire_text(lower_map.get("billing name")),
        "application": _normalize_flywire_text(lower_map.get("application")),
        "raw_json": json.dumps(row_map, ensure_ascii=False),
    }
    return normalized_row


def _build_flywire_summary(document_row, rows):
    amount_total = round(sum(float(row["amount"] or 0) for row in rows), 2)
    times = [row["time_text"] for row in rows if row["time_text"]]
    locations = []
    payment_methods = []
    for row in rows:
        if row["location"] and row["location"] not in locations:
            locations.append(row["location"])
        if row["payment_method"] and row["payment_method"] not in payment_methods:
            payment_methods.append(row["payment_method"])

    summary = {
        "attachment_id": document_row["attachment_id"],
        "attachment_filename": document_row["attachment_filename"],
        "batch_id": document_row["batch_id"],
        "batch_date": document_row["batch_date"],
        "source_filename": document_row["source_filename"],
        "sheet_name": document_row["sheet_name"],
        "row_count": document_row["row_count"],
        "total_amount": amount_total,
        "first_time": times[0] if times else None,
        "last_time": times[-1] if times else None,
        "unique_locations": len(locations),
        "payment_methods": payment_methods[:5],
    }
    return summary


def _remove_existing_flywire_document(conn, attachment_id: int):
    cur = conn.cursor()
    existing_documents = cur.execute(
        f'''
        SELECT id, stored_path
        FROM {_quote_identifier("Import_FlywireDocuments")}
        WHERE {_quote_identifier("attachment_id")} = ?
        ''',
        (attachment_id,),
    ).fetchall()

    for document_id, stored_path in existing_documents:
        cur.execute(
            f'DELETE FROM {_quote_identifier("Import_FlywireRows")} WHERE {_quote_identifier("document_id")} = ?',
            (document_id,),
        )
        cur.execute(
            f'DELETE FROM {_quote_identifier("Import_FlywireDocuments")} WHERE {_quote_identifier("id")} = ?',
            (document_id,),
        )
        if stored_path and os.path.exists(stored_path):
            try:
                os.remove(stored_path)
            except OSError:
                pass


def _flywire_document_payload(document_row, rows):
    summary = json.loads(document_row["summary_json"] or "{}")
    return {
        "document": {
            "id": document_row["id"],
            "attachment_id": document_row["attachment_id"],
            "attachment_filename": document_row["attachment_filename"],
            "batch_id": document_row["batch_id"],
            "batch_date": document_row["batch_date"],
            "source_filename": document_row["source_filename"],
            "stored_filename": document_row["stored_filename"],
            "stored_path": document_row["stored_path"],
            "sheet_name": document_row["sheet_name"],
            "row_count": document_row["row_count"],
            "total_amount": document_row["total_amount"],
            "created_at": document_row["created_at"],
            "updated_at": document_row["updated_at"],
        },
        "summary": summary,
        "rows": [
            {
                "id": row["id"],
                "document_id": row["document_id"],
                "position": row["position"],
                "location": row["location"],
                "department": row["department"],
                "payment_method": row["payment_method"],
                "payment_type": row["payment_type"],
                "time_text": row["time_text"],
                "amount": row["amount"],
                "flywire_id": row["flywire_id"],
                "account_number": row["account_number"],
                "patient_name": row["patient_name"],
                "billing_name": row["billing_name"],
                "application": row["application"],
                "raw_json": row["raw_json"],
            }
            for row in rows
        ],
    }


def _flywire_attachment_context(attachment_row):
    batch_id = attachment_row["batch_id"]
    batch_date = attachment_row["batch_date"]

    if not batch_id or not batch_date:
        fallback_batch_id, fallback_batch_date = _batch_info_from_filename(attachment_row["filename"] or "")
        if not batch_id:
          batch_id = fallback_batch_id
        if not batch_date:
            batch_date = fallback_batch_date

    return batch_id, batch_date


def _flywire_search_tokens(batch_id, batch_date):
    tokens = []

    def add_token(value):
        token = str(value or "").strip().lower()
        if token and token not in tokens:
            tokens.append(token)

    if batch_id:
        add_token(batch_id)
        add_token(str(batch_id).replace(".", ""))
        add_token(str(batch_id).replace(".", "-"))

    normalized_batch_date = _normalize_pending_day(batch_date)
    if normalized_batch_date:
        add_token(normalized_batch_date)
        try:
            parsed_date = datetime.strptime(normalized_batch_date, "%Y-%m-%d")
            add_token(parsed_date.strftime("%m%d%y"))
            add_token(parsed_date.strftime("%m.%d.%y"))
            add_token(parsed_date.strftime("%m-%d-%y"))
            add_token(parsed_date.strftime("%Y%m%d"))
        except ValueError:
            pass

    return tokens


def _find_flywire_candidate_path(batch_id, batch_date):
    if not os.path.isdir(EMAILS_FOLDER):
        return None

    search_tokens = _flywire_search_tokens(batch_id, batch_date)
    if not search_tokens:
        return None

    best_candidate = None
    best_score = 0

    for root, _dirs, files in os.walk(EMAILS_FOLDER):
        for file_name in files:
            lower_name = file_name.lower()
            if "flywire" not in lower_name:
                continue

            score = 0
            for token in search_tokens:
                if token and token in lower_name:
                    score += len(token)

            if score > best_score:
                best_score = score
                best_candidate = os.path.join(root, file_name)

    return best_candidate


def _import_flywire_document(conn, attachment_row, file_name, file_bytes):
    sheet_name, raw_rows = _load_flywire_workbook(file_bytes, file_name)
    parsed_rows = [_row_dict_for_flywire(row_map) for row_map in raw_rows if _flywire_row_is_data(row_map)]
    if not parsed_rows:
        raise HTTPException(status_code=400, detail="No Fly Wire payment rows were found in that workbook")

    attachment_id = attachment_row["id"]
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    extension = os.path.splitext(file_name)[1].lower()
    stored_filename = f"Import_Flywire_{attachment_id}_{timestamp}{extension}"
    stored_path = os.path.join(FLYWIRE_UPLOAD_FOLDER, stored_filename)
    with open(stored_path, "wb") as output_file:
        output_file.write(file_bytes)

    _remove_existing_flywire_document(conn, attachment_id)

    summary = {
        "attachment_id": attachment_id,
        "attachment_filename": attachment_row["filename"],
        "batch_id": attachment_row["batch_id"],
        "batch_date": attachment_row["batch_date"],
        "source_filename": file_name,
        "sheet_name": sheet_name,
        "row_count": len(parsed_rows),
        "total_amount": round(sum(float(row["amount"] or 0) for row in parsed_rows), 2),
        "first_time": next((row["time_text"] for row in parsed_rows if row["time_text"]), None),
        "last_time": next((row["time_text"] for row in reversed(parsed_rows) if row["time_text"]), None),
        "unique_locations": len({row["location"] for row in parsed_rows if row["location"]}),
        "payment_methods": list(dict.fromkeys(row["payment_method"] for row in parsed_rows if row["payment_method"]))[:5],
    }

    now = datetime.now().isoformat(timespec="seconds")
    cur = conn.cursor()
    cur.execute(
        f'''
        INSERT INTO {_quote_identifier("Import_FlywireDocuments")} (
            {_quote_identifier("attachment_id")},
            {_quote_identifier("attachment_filename")},
            {_quote_identifier("batch_id")},
            {_quote_identifier("batch_date")},
            {_quote_identifier("source_filename")},
            {_quote_identifier("stored_filename")},
            {_quote_identifier("stored_path")},
            {_quote_identifier("sheet_name")},
            {_quote_identifier("row_count")},
            {_quote_identifier("total_amount")},
            {_quote_identifier("summary_json")},
            {_quote_identifier("created_at")},
            {_quote_identifier("updated_at")}
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''',
        (
            attachment_id,
            attachment_row["filename"],
            attachment_row["batch_id"],
            attachment_row["batch_date"],
            file_name,
            stored_filename,
            stored_path,
            sheet_name,
            len(parsed_rows),
            summary["total_amount"],
            json.dumps(summary, ensure_ascii=False),
            now,
            now,
        ),
    )
    document_id = cur.lastrowid

    cur.executemany(
        f'''
        INSERT INTO {_quote_identifier("Import_FlywireRows")} (
            {_quote_identifier("document_id")},
            {_quote_identifier("position")},
            {_quote_identifier("location")},
            {_quote_identifier("department")},
            {_quote_identifier("payment_method")},
            {_quote_identifier("payment_type")},
            {_quote_identifier("time_text")},
            {_quote_identifier("amount")},
            {_quote_identifier("flywire_id")},
            {_quote_identifier("account_number")},
            {_quote_identifier("patient_name")},
            {_quote_identifier("billing_name")},
            {_quote_identifier("application")},
            {_quote_identifier("raw_json")}
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''',
        [
            (
                document_id,
                index,
                row["location"],
                row["department"],
                row["payment_method"],
                row["payment_type"],
                row["time_text"],
                row["amount"],
                row["flywire_id"],
                row["account_number"],
                row["patient_name"],
                row["billing_name"],
                row["application"],
                row["raw_json"],
            )
            for index, row in enumerate(parsed_rows, start=1)
        ],
    )

    conn.commit()

    document = conn.execute(
        f'''
        SELECT *
        FROM {_quote_identifier("Import_FlywireDocuments")}
        WHERE {_quote_identifier("id")} = ?
        ''',
        (document_id,),
    ).fetchone()
    rows = conn.execute(
        f'''
        SELECT *
        FROM {_quote_identifier("Import_FlywireRows")}
        WHERE {_quote_identifier("document_id")} = ?
        ORDER BY {_quote_identifier("position")} ASC, {_quote_identifier("id")} ASC
        ''',
        (document_id,),
    ).fetchall()
    return _flywire_document_payload(document, rows)


def _normalize_pending_day(value):
    if value in (None, ""):
        return None

    text = str(value).strip()
    if not text:
        return None

    formats = [
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M:%S.%f",
        "%m/%d/%Y",
        "%m/%d/%Y %H:%M:%S",
        "%m.%d.%Y",
        "%m.%d.%y",
        "%m-%d-%Y",
        "%m-%d-%y",
    ]

    for fmt in formats:
        try:
            if fmt in ("%Y-%m-%d %H:%M:%S", "%m/%d/%Y %H:%M:%S"):
                candidate = text[:19]
            elif fmt in ("%Y-%m-%dT%H:%M:%S.%f",):
                candidate = text[:26]
            else:
                candidate = text[:10]
            return datetime.strptime(candidate, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue

    if "T" in text:
        try:
            return datetime.fromisoformat(text.replace("Z", "")).strftime("%Y-%m-%d")
        except ValueError:
            pass

    return text[:10] if len(text) >= 10 else text


def _row_pending_day(row, batch_date_index, batch_id_index, processed_at_index):
    for index in (batch_date_index, batch_id_index, processed_at_index):
        if index is None or index >= len(row):
            continue
        normalized = _normalize_pending_day(row[index])
        if normalized:
            return normalized
    return "Unknown"


def _run_pythonw_worker(*args):
    if not os.path.exists(PYTHON_EXECUTABLE):
        raise HTTPException(status_code=500, detail="Python executable was not found")

    with tempfile.NamedTemporaryFile(delete=False, suffix=".json") as tmp:
        output_path = tmp.name

    try:
        completed = subprocess.run(
            [PYTHON_EXECUTABLE, EMAIL_DOWNLOADER_SCRIPT, *args, "--output", output_path],
            cwd=os.path.dirname(__file__),
            capture_output=True,
            text=True,
            timeout=120,
        )
        if completed.returncode != 0:
            stderr = (completed.stderr or "").strip()
            stdout = (completed.stdout or "").strip()
            detail = stderr or stdout or "Email downloader worker failed"
            raise HTTPException(status_code=500, detail=detail)

        if not os.path.exists(output_path):
            raise HTTPException(status_code=500, detail="Email downloader worker did not return output")

        with open(output_path, "r", encoding="utf-8") as handle:
            return json.load(handle)
    except subprocess.TimeoutExpired as exc:
        raise HTTPException(status_code=500, detail="Email downloader worker timed out") from exc
    finally:
        try:
            os.remove(output_path)
        except OSError:
            pass


def _run_script_worker(script_path, *args, timeout=120):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".json") as tmp:
        output_path = tmp.name

    try:
        completed = subprocess.run(
            [sys.executable, script_path, "run", *args, "--output", output_path],
            cwd=os.path.dirname(__file__),
            capture_output=True,
            text=True,
            timeout=timeout,
        )
        if completed.returncode != 0:
            stderr = (completed.stderr or "").strip()
            stdout = (completed.stdout or "").strip()
            detail = stderr or stdout or "Snapshot generator worker failed"
            raise HTTPException(status_code=500, detail=detail)

        if not os.path.exists(output_path):
            raise HTTPException(status_code=500, detail="Snapshot generator worker did not return output")

        with open(output_path, "r", encoding="utf-8") as handle:
            return json.load(handle)
    except subprocess.TimeoutExpired as exc:
        raise HTTPException(status_code=500, detail="Snapshot generator worker timed out") from exc
    finally:
        try:
            os.remove(output_path)
        except OSError:
            pass


BALSHEET_TABLE_COLUMNS = [
    ("EntryID", "TEXT PRIMARY KEY"),
    ("PostingDate", "TEXT"),
    ("Type", "TEXT"),
    ("Amount", "REAL"),
    ("Payer", "TEXT"),
    ("Check Number", "TEXT"),
    ("EDI", "TEXT"),
    ("Poster", "TEXT"),
    ("EOB", "TEXT"),
    ("UnPosted", "REAL"),
    ("Misc", "REAL"),
    ("Misc-Type", "TEXT"),
    ("Notes", "TEXT"),
    ("Nick", "REAL"),
    ("Raul", "REAL"),
    ("Needs", "TEXT"),
    ("From", "TEXT"),
    ("To", "TEXT"),
]

BALSHEET_TABLE_COLUMN_NAMES = [name for name, _ in BALSHEET_TABLE_COLUMNS]


def _balsheet_row_value(row, field_name: str, index: int):
    if isinstance(row, sqlite3.Row):
        return row[field_name]

    if isinstance(row, dict):
        return row.get(field_name)

    try:
        return row[index]
    except Exception:
        return None


def ensure_balsheet_table(conn=None):
    close_conn = False
    if conn is None:
        conn = get_conn()
        close_conn = True

    cur = conn.cursor()
    column_defs = ", ".join(f'{_quote_identifier(name)} {definition}' for name, definition in BALSHEET_TABLE_COLUMNS)
    cur.execute(f'CREATE TABLE IF NOT EXISTS {_quote_identifier("Balsheet")} ({column_defs})')
    conn.commit()

    if close_conn:
        conn.close()


def ensure_balsheet_notes_table(conn=None):
    close_conn = False
    if conn is None:
        conn = get_conn()
        close_conn = True

    cur = conn.cursor()
    cur.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {_quote_identifier("Balsheet_notes")} (
            {_quote_identifier("post_date")} TEXT,
            {_quote_identifier("notes")} TEXT,
            {_quote_identifier("message")} TEXT
        )
        """
    )
    existing_columns = [row[1] for row in cur.execute(f'PRAGMA table_info({_quote_identifier("Balsheet_notes")})').fetchall()]
    if "pk" in {column.lower() for column in existing_columns}:
        cur.execute(
            f'ALTER TABLE {_quote_identifier("Balsheet_notes")} RENAME TO {_quote_identifier("Balsheet_notes_legacy")}'
        )
        cur.execute(
            f"""
            CREATE TABLE {_quote_identifier("Balsheet_notes")} (
                {_quote_identifier("post_date")} TEXT,
                {_quote_identifier("notes")} TEXT,
                {_quote_identifier("message")} TEXT
            )
            """
        )
        cur.execute(
            f'INSERT INTO {_quote_identifier("Balsheet_notes")} ({_quote_identifier("post_date")}, {_quote_identifier("notes")}, {_quote_identifier("message")}) '
            f'SELECT {_quote_identifier("post_date")}, {_quote_identifier("notes")}, COALESCE({_quote_identifier("message")}, \'\') FROM {_quote_identifier("Balsheet_notes_legacy")}'
        )
        cur.execute(f'DROP TABLE {_quote_identifier("Balsheet_notes_legacy")}')
    elif "message" not in {column.lower() for column in existing_columns}:
        cur.execute(
            f'ALTER TABLE {_quote_identifier("Balsheet_notes")} ADD COLUMN {_quote_identifier("message")} TEXT'
        )
    conn.commit()

    if close_conn:
        conn.close()


MISC_TABLE_COLUMNS = [
    ("misc_id", "TEXT PRIMARY KEY"),
    ("posting_date", "TEXT NOT NULL"),
    ("amount", "REAL NOT NULL DEFAULT 0"),
    ("misc_type", "TEXT NOT NULL DEFAULT ''"),
    ("details", "TEXT NOT NULL DEFAULT ''"),
    ("created_at", "TEXT NOT NULL DEFAULT ''"),
]


def ensure_misc_table(conn=None):
    close_conn = False
    if conn is None:
        conn = get_conn()
        close_conn = True

    cur = conn.cursor()
    column_defs = ", ".join(f'{_quote_identifier(name)} {definition}' for name, definition in MISC_TABLE_COLUMNS)
    cur.execute(f'CREATE TABLE IF NOT EXISTS {_quote_identifier("Misc")} ({column_defs})')

    existing_columns = [row[1] for row in cur.execute(f'PRAGMA table_info({_quote_identifier("Misc")})').fetchall()]
    existing_columns_lower = {column.lower() for column in existing_columns}
    for column_name, definition in MISC_TABLE_COLUMNS:
        if column_name.lower() not in existing_columns_lower:
            cur.execute(
                f'ALTER TABLE {_quote_identifier("Misc")} ADD COLUMN {_quote_identifier(column_name)} {definition}'
            )

    conn.commit()

    if close_conn:
        conn.close()


IMAGING_DOCUMENT_INDEX_COLUMNS = [
    ("file_path", "TEXT PRIMARY KEY"),
    ("file_name", "TEXT NOT NULL"),
    ("file_ext", "TEXT NOT NULL"),
    ("normalized_name", "TEXT NOT NULL"),
    ("check_numbers_json", "TEXT NOT NULL"),
    ("is_archived", "INTEGER NOT NULL DEFAULT 0"),
    ("source_folder", "TEXT NOT NULL DEFAULT ''"),
    ("size_bytes", "INTEGER NOT NULL DEFAULT 0"),
    ("page_count", "INTEGER NOT NULL DEFAULT 0"),
    ("text_content", "TEXT NOT NULL DEFAULT ''"),
    ("outline_json", "TEXT NOT NULL DEFAULT '[]'"),
    ("scanned_at", "TEXT NOT NULL"),
    ("updated_at", "TEXT NOT NULL"),
]

IMAGING_BALSHEET_LINK_COLUMNS = [
    ("link_id", "TEXT PRIMARY KEY"),
    ("entry_id", "TEXT NOT NULL"),
    ("posting_date", "TEXT NOT NULL"),
    ("lockbox_image_date", "TEXT NOT NULL DEFAULT ''"),
    ("amount", "REAL NOT NULL DEFAULT 0"),
    ("payer", "TEXT NOT NULL DEFAULT ''"),
    ("check_number", "TEXT NOT NULL DEFAULT ''"),
    ("file_path", "TEXT NOT NULL"),
    ("file_name", "TEXT NOT NULL DEFAULT ''"),
    ("match_method", "TEXT NOT NULL DEFAULT 'manual'"),
    ("confidence", "REAL NOT NULL DEFAULT 0"),
    ("bookmark_page", "INTEGER NOT NULL DEFAULT 0"),
    ("bookmark_title", "TEXT NOT NULL DEFAULT ''"),
    ("source_query", "TEXT NOT NULL DEFAULT ''"),
    ("confirmed", "INTEGER NOT NULL DEFAULT 1"),
    ("created_at", "TEXT NOT NULL"),
    ("updated_at", "TEXT NOT NULL"),
]

IMAGING_PDF_TOOL_NAMES = {
    "pdfinfo": "pdfinfo.exe",
    "pdftotext": "pdftotext.exe",
    "pdftohtml": "pdftohtml.exe",
}


def _poppler_tool_path(tool_name: str) -> str:
    executable = IMAGING_PDF_TOOL_NAMES.get(tool_name, tool_name)
    candidate = Path(WORKFLOW_ROOT) / "poppler" / "Library" / "bin" / executable
    return str(candidate) if candidate.exists() else executable


def _run_text_command(command: list[str], *, cwd: str | None = None) -> str:
    try:
        result = subprocess.run(
            command,
            cwd=cwd,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            check=False,
        )
    except Exception:
        return ""

    if result.returncode != 0:
        return result.stdout or result.stderr or ""

    return result.stdout


def _normalize_search_text(value: str | None) -> str:
    return " ".join(str(value or "").upper().split())


def _date_prefix_from_mmddyyyy(value: str | None) -> str:
    normalized = normalize_mmddyyyy(value)
    if not normalized:
        return ""

    try:
        parsed = datetime.strptime(normalized, "%m/%d/%Y")
    except ValueError:
        return ""

    return parsed.strftime("%m.%d.%y")


def _extract_pdf_outline(xml_text: str) -> list[dict]:
    if not xml_text.strip():
        return []

    try:
        root = ET.fromstring(xml_text)
    except Exception:
        return []

    outline_root = root.find("outline")
    if outline_root is None:
        return []

    items: list[dict] = []

    def walk(node, depth: int = 0):
        for child in list(node):
            if child.tag != "item":
                continue

            title = html_lib.unescape("".join(child.itertext()).strip())
            page_value = str(child.attrib.get("page") or "").strip()
            try:
                page_number = int(page_value)
            except ValueError:
                page_number = 0

            if title:
                items.append({"title": title, "page": page_number, "depth": depth})

            nested_outline = child.find("outline")
            if nested_outline is not None:
                walk(nested_outline, depth + 1)

    walk(outline_root, 0)
    return items


def _extract_html_text(file_path: Path) -> str:
    try:
        raw = file_path.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return ""

    stripped = re.sub(r"<script[\s\S]*?</script>", " ", raw, flags=re.IGNORECASE)
    stripped = re.sub(r"<style[\s\S]*?</style>", " ", stripped, flags=re.IGNORECASE)
    stripped = re.sub(r"<[^>]+>", " ", stripped)
    stripped = html_lib.unescape(stripped)
    return " ".join(stripped.split())


@lru_cache(maxsize=256)
def _extract_pdf_search_bundle(file_path: str, modified_time: float) -> tuple[str, int, str]:
    path = Path(file_path)
    if not path.exists():
        return "", 0, "[]"

    pdfinfo_output = _run_text_command([_poppler_tool_path("pdfinfo"), str(path)])
    page_count = 0
    for line in pdfinfo_output.splitlines():
        if line.startswith("Pages:"):
            try:
                page_count = int(line.split(":", 1)[1].strip())
            except ValueError:
                page_count = 0
            break

    text_content = _run_text_command([_poppler_tool_path("pdftotext"), "-layout", str(path), "-"])
    outline_xml = _run_text_command([_poppler_tool_path("pdftohtml"), "-xml", "-stdout", str(path)])
    outline_json = json.dumps(_extract_pdf_outline(outline_xml))
    return text_content.strip(), page_count, outline_json


def _build_document_search_payload(file_row: sqlite3.Row) -> dict:
    file_path = str(file_row["file_path"] or "")
    file_ext = str(file_row["file_ext"] or "").lower()
    modified_time = 0.0
    try:
        modified_time = Path(file_path).stat().st_mtime
    except Exception:
        modified_time = 0.0

    if file_ext == "pdf":
        text_content, page_count, outline_json = _extract_pdf_search_bundle(file_path, modified_time)
    else:
        text_content = _extract_html_text(Path(file_path))
        page_count = 1 if text_content else 0
        outline_json = "[]"

    return {
        "filePath": file_path,
        "fileName": str(file_row["file_name"] or ""),
        "fileExt": file_ext,
        "normalizedName": str(file_row["normalized_name"] or ""),
        "checkNumbers": json.loads(file_row["check_numbers_json"] or "[]"),
        "isArchived": bool(file_row["is_archived"]),
        "sourceFolder": str(file_row["source_folder"] or ""),
        "sizeBytes": int(file_row["size_bytes"] or 0),
        "pageCount": page_count,
        "textContent": text_content,
        "outline": json.loads(outline_json or "[]"),
        "scannedAt": str(file_row["scanned_at"] or ""),
        "updatedAt": str(file_row["updated_at"] or ""),
    }


def _build_search_snippet(text: str, query: str, width: int = 80) -> str:
    normalized_text = " ".join(str(text or "").split())
    normalized_query = _normalize_search_text(query)
    if not normalized_text:
        return ""

    if not normalized_query:
        return normalized_text[:width]

    haystack = normalized_text.upper()
    needle_index = haystack.find(normalized_query)
    if needle_index < 0:
        return normalized_text[:width]

    start = max(0, needle_index - width // 2)
    end = min(len(normalized_text), needle_index + len(normalized_query) + width // 2)
    snippet = normalized_text[start:end].strip()
    if start > 0:
        snippet = f"... {snippet}"
    if end < len(normalized_text):
        snippet = f"{snippet} ..."
    return snippet


def _search_lockbox_documents(posting_date: str, query: str) -> list[dict]:
    root = _imaging_root_folder()
    if not root.exists():
        return []

    date_prefix = _date_prefix_from_mmddyyyy(posting_date)
    normalized_query = _normalize_search_text(query)
    candidates: list[dict] = []

    def _lockbox_filename_schema_score(file_name: str, normalized_date_prefix: str) -> tuple[float, str]:
        normalized_file_name = _normalize_search_text(file_name)
        score = 0.0
        tags: list[str] = []

        if normalized_date_prefix and normalized_date_prefix in normalized_file_name:
            score += 0.25
            tags.append("date")

        if "WF" in normalized_file_name:
            score += 0.20
            tags.append("wf")

        if "LOCKBOX" in normalized_file_name:
            score += 0.20
            tags.append("lockbox")

        schema_count_match = re.search(r"(?:^|[^A-Z0-9])[-_ ](\d{1,3})(?:[^A-Z0-9]|$)", str(file_name).upper())
        if schema_count_match:
            count_value = int(schema_count_match.group(1))
            if count_value > 0:
                score += 0.05
                tags.append("count")

        if score >= 0.45:
            tags.append("schema")

        return min(score, 0.6), "+".join(tags)

    for path in sorted(root.rglob("*.pdf")):
        if not path.is_file():
            continue

        file_name_upper = path.name.upper()
        if "WF" not in file_name_upper and "LOCKBOX" not in file_name_upper:
            continue
        if date_prefix and date_prefix.upper() not in file_name_upper:
            continue

        try:
            file_stat = path.stat()
        except Exception:
            continue

        text_content, page_count, outline_json = _extract_pdf_search_bundle(str(path), file_stat.st_mtime)
        page_texts = [page.strip() for page in text_content.split("\f") if page.strip()]
        outline = json.loads(outline_json or "[]")
        normalized_file_name = _normalize_search_text(path.name)
        normalized_text = _normalize_search_text(text_content)
        score = 0.0
        matched_page = 0
        matched_bookmark = ""
        matched_snippet = ""
        match_method = "date-prefix"

        schema_score, schema_tag = _lockbox_filename_schema_score(path.name, date_prefix.upper() if date_prefix else "")
        if schema_score:
            score += schema_score
            match_method = "filename-schema"
            if schema_tag:
                match_method = f"{match_method}-{schema_tag}"
        elif date_prefix:
            score += 0.2

        if normalized_query:
            if normalized_query in normalized_file_name:
                score += 0.2
                match_method = "filename"
                if schema_score:
                    match_method = f"{match_method}-schema"

            if normalized_query in normalized_text:
                score += 0.25
                match_method = "text"

            for page_number, page_text in enumerate(page_texts, start=1):
                if normalized_query not in _normalize_search_text(page_text):
                    continue
                matched_page = page_number
                matched_snippet = _build_search_snippet(page_text, query)
                score = max(score, 0.7)
                match_method = "page-text"
                if schema_score:
                    match_method = f"{match_method}-schema"
                break

            for outline_item in outline:
                title = str(outline_item.get("title") or "")
                if normalized_query not in _normalize_search_text(title):
                    continue
                matched_bookmark = title
                matched_page = int(outline_item.get("page") or matched_page or 0)
                score = max(score, 0.9)
                match_method = "bookmark"
                if schema_score:
                    match_method = f"{match_method}-schema"
                break
        else:
            if outline:
                first_outline = outline[0]
                matched_page = int(first_outline.get("page") or 0)
                matched_bookmark = str(first_outline.get("title") or "")
                score += 0.1

        if not normalized_query and not score:
            score = 0.2

        if normalized_query and score <= 0:
            continue

        if not matched_snippet and page_texts:
            matched_snippet = _build_search_snippet(page_texts[0], query)

        candidates.append(
            {
                "filePath": str(path),
                "fileName": path.name,
                "pageCount": page_count,
                "confidence": round(min(score, 1.0), 2),
                "matchMethod": match_method,
                "bookmarkPage": matched_page,
                "bookmarkTitle": matched_bookmark,
                "snippet": matched_snippet,
                "openUrl": f"/imaging/files/open?path={quote(str(path), safe='')}" + (f"#page={matched_page}" if matched_page > 0 else ""),
                "sourceFolder": str(path.parent.relative_to(root)) if path.parent.is_relative_to(root) else str(path.parent),
            }
        )

    candidates.sort(key=lambda item: (item["confidence"], item["pageCount"], item["fileName"]), reverse=True)
    return candidates[:10]


def _normalize_imaging_check_number(value):
    return re.sub(r"[^A-Z0-9]", "", str(value or "").upper().strip())


def _extract_imaging_filename_checks(path: Path) -> list[str]:
    tokens = re.split(r"[^A-Z0-9]+", path.stem.upper())
    checks: list[str] = []
    for token in tokens:
        normalized = _normalize_imaging_check_number(token)
        if not normalized:
            continue
        if normalized.isdigit():
            if len(normalized) < 4:
                continue
        elif len(normalized) < 5:
            continue
        if not any(char.isdigit() for char in normalized):
            continue
        if normalized not in checks:
            checks.append(normalized)
    return checks


def _extract_imaging_filename_tokens(file_name: str) -> list[str]:
    tokens = re.split(r"[^A-Z0-9]+", str(file_name or "").upper())
    normalized_tokens: list[str] = []
    for token in tokens:
        normalized = _normalize_imaging_check_number(token)
        if not normalized:
            continue
        if normalized.isdigit():
            if len(normalized) < 4:
                continue
        elif len(normalized) < 5:
            continue
        if not any(char.isdigit() for char in normalized):
            continue
        if normalized not in normalized_tokens:
            normalized_tokens.append(normalized)
    return normalized_tokens


def _imaging_filename_has_exact_check(file_name: str, check_number: str) -> bool:
    normalized_check = _normalize_imaging_check_number(check_number)
    if not normalized_check:
        return False

    normalized_file_name = str(file_name or "").upper()
    return re.search(rf"(?<![A-Z0-9]){re.escape(normalized_check)}(?![A-Z0-9])", normalized_file_name) is not None


def _imaging_filename_partial_check_score(file_name: str, check_number: str) -> float:
    normalized_check = _normalize_imaging_check_number(check_number)
    if not normalized_check:
        return 0.0

    normalized_file_name = _normalize_imaging_check_number(file_name)
    if normalized_check not in normalized_file_name:
        return 0.0

    tokens = _extract_imaging_filename_tokens(file_name)
    best_score = 0.0
    for token in tokens:
        if normalized_check not in token:
            continue

        # If the check number is embedded inside a longer alphanumeric token,
        # keep it as a weaker signal than an exact token boundary match.
        if token == normalized_check:
            return 1.0

        ratio = len(normalized_check) / max(len(token), 1)
        best_score = max(best_score, 0.55 + (0.35 * ratio))

    return min(best_score, 0.94)


def _normalize_imaging_amount_token(value) -> str:
    try:
        amount = float(str(value or 0).replace(",", "").replace("$", "").strip())
    except Exception:
        return ""

    if amount.is_integer():
        return f"{int(amount):d}"

    return f"{amount:.2f}".rstrip("0").rstrip(".")


def _imaging_amount_tokens(value) -> list[str]:
    normalized = _normalize_imaging_amount_token(value)
    if not normalized:
        return []

    try:
        amount = float(normalized)
    except Exception:
        amount = 0.0

    tokens = {
        normalized,
        f"{amount:,.2f}",
        f"{amount:.2f}",
        f"${amount:,.2f}",
    }
    if amount.is_integer():
        tokens.add(f"{int(amount):,d}")
        tokens.add(f"{int(amount):d}.00")
    return [token for token in tokens if token]


def _build_lockbox_recommendation_snippet(page_texts: list[str], page_number: int, query: str) -> str:
    if page_number > 0 and page_number <= len(page_texts):
        snippet = _build_search_snippet(page_texts[page_number - 1], query)
        if snippet:
            return snippet
    if page_texts:
        return _build_search_snippet(page_texts[0], query)
    return ""


def _build_lockbox_row_recommendations(posting_date: str, row: dict, document_rows: list[sqlite3.Row]) -> list[dict]:
    date_prefix = _date_prefix_from_mmddyyyy(posting_date)
    normalized_date_prefix = date_prefix.upper() if date_prefix else ""
    entry_check = _normalize_imaging_check_number(row.get("checkNumber") or row.get("check_number") or "")
    payer = str(row.get("payer") or "").strip()
    normalized_payer = _normalize_search_text(payer)
    amount_tokens = _imaging_amount_tokens(row.get("amount") or 0)
    amount_display = f"{float(row.get('amount') or 0):,.2f}"
    recommendations: list[dict] = []

    for document_row in document_rows:
        score, match_method = _imaging_match_score(entry_check, posting_date, document_row)
        file_path = str(document_row["file_path"] or "")
        file_name = str(document_row["file_name"] or "")
        normalized_file_name = _normalize_search_text(file_name)
        normalized_name = str(document_row["normalized_name"] or "").upper()
        text_content = str(document_row["text_content"] or "")
        normalized_text = _normalize_search_text(text_content)
        outline = []
        try:
            outline = json.loads(document_row["outline_json"] or "[]")
        except Exception:
            outline = []
        page_texts = [page.strip() for page in text_content.split("\f") if page.strip()]
        matched_page = 0
        matched_title = ""
        matched_check = ""
        matched_amount = ""
        notes: list[str] = []

        if normalized_date_prefix and normalized_date_prefix not in normalized_file_name:
            continue

        if entry_check:
            if _imaging_filename_has_exact_check(file_name, entry_check):
                matched_check = entry_check
                score = max(score, 0.82)
                notes.append("check")
            else:
                for page_number, page_text in enumerate(page_texts, start=1):
                    normalized_page = _normalize_search_text(page_text)
                    if entry_check not in normalized_page:
                        continue
                    matched_page = page_number
                    matched_check = entry_check
                    score = max(score, 0.86)
                    notes.append("check-page")
                    break

        for token in amount_tokens:
            if token and (token in normalized_name or token in normalized_text):
                matched_amount = token
                score = max(score, 0.82 if token in normalized_name else 0.75)
                notes.append("amount")
                break

            for page_number, page_text in enumerate(page_texts, start=1):
                normalized_page = _normalize_search_text(page_text)
                if token not in normalized_page:
                    continue
                matched_page = page_number if not matched_page else matched_page
                matched_amount = token
                score = max(score, 0.88)
                notes.append("amount-page")
                break
            if matched_amount:
                break

        if normalized_payer and normalized_payer in normalized_text:
            score = min(1.0, score + 0.08)
            notes.append("payer")

        if outline:
            for outline_item in outline:
                title = str(outline_item.get("title") or "")
                normalized_title = _normalize_search_text(title)
                if entry_check and entry_check not in normalized_title:
                    continue
                if amount_tokens and not any(token in normalized_title for token in amount_tokens):
                    continue
                matched_title = title
                matched_page = int(outline_item.get("page") or matched_page or 0)
                score = max(score, 0.9)
                notes.append("bookmark")
                break

        if matched_page == 0 and outline:
            first_outline = outline[0]
            matched_page = int(first_outline.get("page") or 0)
            if not matched_title:
                matched_title = str(first_outline.get("title") or "")

        if matched_page == 0 and page_texts:
            for page_number, page_text in enumerate(page_texts, start=1):
                normalized_page = _normalize_search_text(page_text)
                if (entry_check and entry_check in normalized_page) or any(token in normalized_page for token in amount_tokens):
                    matched_page = page_number
                    break

        if not matched_check and entry_check:
            matched_check = entry_check
        if not matched_amount and amount_tokens:
            matched_amount = amount_tokens[0]
        if score <= 0:
            continue

        matched_text = " ".join(notes) if notes else "search"
        recommendations.append(
            {
                "filePath": file_path,
                "fileName": file_name,
                "confidence": round(min(score, 1.0), 2),
                "matchMethod": f"{match_method}-{matched_text}" if matched_text else match_method,
                "bookmarkPage": matched_page,
                "bookmarkTitle": matched_title,
                "snippet": _build_lockbox_recommendation_snippet(page_texts, matched_page, f"{entry_check} {amount_display} {payer}".strip()),
                "sourceFolder": str(document_row["source_folder"] or ""),
                "foundCheckNumber": matched_check,
                "foundAmount": amount_display,
            }
        )

    recommendations.sort(key=lambda item: (item["confidence"], item["bookmarkPage"], item["fileName"]), reverse=True)
    return recommendations[:3]


def _imaging_posting_date_tokens(posting_date: str) -> set[str]:
    normalized = normalize_mmddyyyy(posting_date)
    if not normalized:
        return set()

    try:
        parsed = datetime.strptime(normalized, "%m/%d/%Y")
    except ValueError:
        return { _normalize_imaging_check_number(normalized) }

    month = parsed.strftime("%m")
    day = parsed.strftime("%d")
    year = parsed.strftime("%Y")
    year_short = parsed.strftime("%y")
    return {
        f"{year}{month}{day}",
        f"{month}{day}{year}",
        f"{month}{day}{year_short}",
    }


def _imaging_root_folder() -> Path:
    return Path(ZIP_835_HTML_FOLDER).resolve()


def _is_within_root(candidate_path: str, root_path: Path) -> bool:
    try:
        Path(candidate_path).resolve().relative_to(root_path.resolve())
        return True
    except Exception:
        return False


def ensure_imaging_tables(conn=None):
    close_conn = False
    if conn is None:
        conn = get_conn()
        close_conn = True

    cur = conn.cursor()

    index_column_defs = ", ".join(f'{_quote_identifier(name)} {definition}' for name, definition in IMAGING_DOCUMENT_INDEX_COLUMNS)
    cur.execute(f'CREATE TABLE IF NOT EXISTS {_quote_identifier("Imaging_DocumentFileIndex")} ({index_column_defs})')

    index_columns = {row[1].lower() for row in cur.execute(f'PRAGMA table_info({_quote_identifier("Imaging_DocumentFileIndex")})').fetchall()}
    for column_name, column_type in IMAGING_DOCUMENT_INDEX_COLUMNS:
        if column_name.lower() not in index_columns:
            cur.execute(
                f'ALTER TABLE {_quote_identifier("Imaging_DocumentFileIndex")} ADD COLUMN {_quote_identifier(column_name)} {column_type}'
            )

    link_column_defs = ", ".join(f'{_quote_identifier(name)} {definition}' for name, definition in IMAGING_BALSHEET_LINK_COLUMNS)
    cur.execute(
        f'CREATE TABLE IF NOT EXISTS {_quote_identifier("Imaging_BalsheetDocumentLinks")} '
        f'({link_column_defs}, UNIQUE({_quote_identifier("entry_id")}, {_quote_identifier("file_path")}))'
    )

    link_columns = {row[1].lower() for row in cur.execute(f'PRAGMA table_info({_quote_identifier("Imaging_BalsheetDocumentLinks")})').fetchall()}
    for column_name, column_type in IMAGING_BALSHEET_LINK_COLUMNS:
        if column_name.lower() not in link_columns:
            cur.execute(
                f'ALTER TABLE {_quote_identifier("Imaging_BalsheetDocumentLinks")} ADD COLUMN {_quote_identifier(column_name)} {column_type}'
            )

    cur.execute(
        f'CREATE INDEX IF NOT EXISTS {_quote_identifier("idx_imaging_links_posting_date")} '
        f'ON {_quote_identifier("Imaging_BalsheetDocumentLinks")} ({_quote_identifier("posting_date")})'
    )
    cur.execute(
        f'CREATE INDEX IF NOT EXISTS {_quote_identifier("idx_imaging_links_check_number")} '
        f'ON {_quote_identifier("Imaging_BalsheetDocumentLinks")} ({_quote_identifier("check_number")})'
    )

    conn.commit()

    if close_conn:
        conn.close()


def rebuild_imaging_document_index(conn=None):
    close_conn = False
    if conn is None:
        conn = get_conn()
        close_conn = True

    ensure_imaging_tables(conn)

    cur = conn.cursor()
    root = _imaging_root_folder()
    scanned_at = datetime.now().isoformat(timespec="seconds")
    rows: list[tuple] = []

    cur.execute(f'DELETE FROM {_quote_identifier("Imaging_DocumentFileIndex")}')

    if root.exists():
        for path in root.rglob("*"):
            if not path.is_file():
                continue

            suffix = path.suffix.lower()
            if suffix not in {".pdf", ".html", ".htm"}:
                continue

            try:
                resolved_path = str(path.resolve())
                stat = path.stat()
            except Exception:
                continue

            check_numbers = _extract_imaging_filename_checks(path)
            normalized_name = _normalize_imaging_check_number(path.stem)
            is_archived = 1 if any(part.lower() == "archived" for part in path.parts) else 0
            try:
                source_folder = str(path.parent.resolve().relative_to(root))
            except Exception:
                source_folder = str(path.parent)

            if suffix == ".pdf":
                text_content, page_count, outline_json = _extract_pdf_search_bundle(resolved_path, stat.st_mtime)
            else:
                text_content = _extract_html_text(path)
                page_count = 1 if text_content else 0
                outline_json = "[]"

            rows.append(
                (
                    resolved_path,
                    path.name,
                    suffix.lstrip("."),
                    normalized_name,
                    json.dumps(check_numbers),
                    is_archived,
                    source_folder,
                    int(stat.st_size),
                    int(page_count),
                    text_content,
                    outline_json,
                    scanned_at,
                    scanned_at,
                )
            )

    if rows:
        cur.executemany(
            f"""
            INSERT INTO {_quote_identifier("Imaging_DocumentFileIndex")} (
                {_quote_identifier("file_path")},
                {_quote_identifier("file_name")},
                {_quote_identifier("file_ext")},
                {_quote_identifier("normalized_name")},
                {_quote_identifier("check_numbers_json")},
                {_quote_identifier("is_archived")},
                {_quote_identifier("source_folder")},
                {_quote_identifier("size_bytes")},
                {_quote_identifier("page_count")},
                {_quote_identifier("text_content")},
                {_quote_identifier("outline_json")},
                {_quote_identifier("scanned_at")},
                {_quote_identifier("updated_at")}
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            rows,
        )

    conn.commit()
    result = len(rows)

    if close_conn:
        conn.close()

    return result


def _load_imaging_document_index(conn):
    ensure_imaging_tables(conn)
    row_count = conn.execute(f'SELECT COUNT(*) FROM {_quote_identifier("Imaging_DocumentFileIndex")}').fetchone()[0]
    if not row_count:
        rebuild_imaging_document_index(conn)

    conn.row_factory = sqlite3.Row
    return conn.execute(
        f'''
        SELECT *
        FROM {_quote_identifier("Imaging_DocumentFileIndex")}
        ORDER BY {_quote_identifier("is_archived")} ASC, {_quote_identifier("file_name")} ASC
        '''
    ).fetchall()


def _imaging_match_score(entry_check: str, posting_date: str, document_row) -> tuple[float, str]:
    if not entry_check:
        return 0.0, "none"

    try:
        file_checks = json.loads(document_row["check_numbers_json"] or "[]")
    except Exception:
        file_checks = []

    normalized_name = str(document_row["normalized_name"] or "").upper()
    file_name = str(document_row["file_name"] or "").upper()
    filename_tokens = _extract_imaging_filename_tokens(file_name)
    posting_date_tokens = _imaging_posting_date_tokens(posting_date)
    has_posting_date = any(token and token in normalized_name for token in posting_date_tokens)
    exact_filename_check = _imaging_filename_has_exact_check(file_name, entry_check)
    partial_filename_check_score = _imaging_filename_partial_check_score(file_name, entry_check)

    if exact_filename_check:
        score = 1.0
        match_method = "filename-token"

        if filename_tokens and filename_tokens[-1] == entry_check:
            match_method = "filename-schema"

        if has_posting_date:
            match_method = "filename-schema-date" if match_method != "filename-token" else "filename-token-date"

        return score, match_method

    if entry_check in file_checks and partial_filename_check_score > 0:
        score = max(partial_filename_check_score, 0.85 if document_row["file_ext"].lower() == "pdf" else 0.77)
        match_method = "filename-token-partial"

        if has_posting_date:
            score = min(1.0, score + 0.03)
            match_method = "filename-token-partial-date"

        return score, match_method

    if entry_check in file_checks:
        score = 0.98 if document_row["file_ext"].lower() == "pdf" else 0.92
        match_method = "filename-token"

        if filename_tokens and filename_tokens[-1] == entry_check:
            score = min(1.0, score + 0.01)
            match_method = "filename-schema"

        if has_posting_date:
            score = min(1.0, score + 0.02)
            match_method = "filename-schema-date" if match_method != "filename-token" else "filename-token-date"

        return score, match_method

    if entry_check in normalized_name:
        score = 0.9 if document_row["file_ext"].lower() == "pdf" else 0.82
        match_method = "filename"

        if filename_tokens and entry_check in filename_tokens:
            score = max(score, 0.92 if document_row["file_ext"].lower() == "pdf" else 0.86)
            match_method = "filename-token"

        if has_posting_date:
            score = min(1.0, score + 0.05)
            match_method = "filename-schema-date"

        return score, match_method

    if has_posting_date and entry_check in _normalize_imaging_check_number(file_name):
        return (0.86 if document_row["file_ext"].lower() == "pdf" else 0.78), "filename-schema"

    return 0.0, "none"


def _build_imaging_association_payload(conn, posting_date: str):
    balsheet_rows = _load_balsheet_review_day_rows(conn, posting_date)
    index_rows = _load_imaging_document_index(conn)

    link_rows = conn.execute(
        f'''
        SELECT *
        FROM {_quote_identifier("Imaging_BalsheetDocumentLinks")}
        WHERE {_quote_identifier("posting_date")} = ?
        ORDER BY {_quote_identifier("updated_at")} DESC
        ''',
        (posting_date,),
    ).fetchall()

    confirmed_links_by_entry: dict[str, list[sqlite3.Row]] = {}
    for link_row in link_rows:
        confirmed_links_by_entry.setdefault(str(link_row["entry_id"]), []).append(link_row)

    rows = []
    for balsheet_row in balsheet_rows:
        entry_id = str(balsheet_row.get("entry_id") or "")
        entry_check = _normalize_imaging_check_number(balsheet_row.get("check_number"))
        matches = []
        for document_row in index_rows:
            score, match_method = _imaging_match_score(entry_check, posting_date, document_row)
            if score <= 0:
                continue

            matches.append(
                {
                    "filePath": document_row["file_path"],
                    "fileName": document_row["file_name"],
                    "fileExt": document_row["file_ext"],
                    "isArchived": bool(document_row["is_archived"]),
                    "sourceFolder": document_row["source_folder"],
                    "confidence": round(score, 2),
                    "matchMethod": match_method,
                }
            )

        matches.sort(key=lambda item: (item["confidence"], item["fileExt"].lower() == "pdf", not item["isArchived"], item["fileName"]), reverse=True)
        linked_rows = confirmed_links_by_entry.get(entry_id, [])

        rows.append(
            {
                "entryId": entry_id,
                "postingDate": balsheet_row.get("posting_date") or "",
                "type": str(balsheet_row.get("type") or ""),
                "amount": float(balsheet_row.get("amount") or 0),
                "payer": str(balsheet_row.get("payer") or ""),
                "checkNumber": str(balsheet_row.get("check_number") or ""),
                "linkedFiles": [
                    {
                        "linkId": link_row["link_id"],
                        "filePath": link_row["file_path"],
                        "fileName": link_row["file_name"],
                        "matchMethod": link_row["match_method"],
                        "confidence": float(link_row["confidence"] or 0),
                        "bookmarkPage": int(link_row["bookmark_page"] or 0),
                        "bookmarkTitle": str(link_row["bookmark_title"] or ""),
                        "sourceQuery": str(link_row["source_query"] or ""),
                        "confirmed": bool(link_row["confirmed"]),
                        "openUrl": (
                            f"/imaging/balsheet-links/{link_row['link_id']}/open"
                            + (f"#page={int(link_row['bookmark_page'] or 0)}" if int(link_row["bookmark_page"] or 0) > 0 else "")
                        ),
                    }
                    for link_row in linked_rows
                ],
                "matches": [
                {
                    **match,
                    "openUrl": f"/imaging/files/open?path={quote(match['filePath'], safe='')}",
                }
                    for match in matches[:5]
                ],
                "recommendations": [],
            }
        )

    return {
        "postingDate": posting_date,
        "rowCount": len(rows),
        "indexCount": len(index_rows),
        "rows": rows,
    }


def _build_imaging_lockbox_recommendations_payload(conn, posting_date: str):
    payload = _build_imaging_association_payload(conn, posting_date)
    index_rows = _load_imaging_document_index(conn)

    for row in payload["rows"]:
        if str(row.get("type") or "").strip().lower() != "lockbox":
            row["recommendations"] = []
            continue

        row["recommendations"] = _build_lockbox_row_recommendations(posting_date, row, index_rows)

    return payload


def _safe_imaging_file_response(file_path: str):
    root = _imaging_root_folder()
    if not _is_within_root(file_path, root):
        raise HTTPException(status_code=404, detail="File not found")

    resolved = Path(file_path).resolve()
    if not resolved.exists():
        raise HTTPException(status_code=404, detail="File not found")

    return FileResponse(str(resolved))


TASK_TABLE_COLUMNS = [
    ("task_id", "TEXT PRIMARY KEY"),
    ("task_list", "TEXT NOT NULL"),
    ("title", "TEXT NOT NULL"),
    ("details", "TEXT NOT NULL DEFAULT ''"),
    ("category", "TEXT NOT NULL DEFAULT ''"),
    ("recurrence", "TEXT NOT NULL DEFAULT 'none'"),
    ("action_type", "TEXT NOT NULL DEFAULT 'none'"),
    ("action_label", "TEXT NOT NULL DEFAULT ''"),
    ("action_value", "TEXT NOT NULL DEFAULT ''"),
    ("done", "INTEGER NOT NULL DEFAULT 0"),
    ("sort_order", "INTEGER NOT NULL DEFAULT 0"),
    ("next_due_at", "TEXT"),
    ("completed_at", "TEXT"),
    ("created_at", "TEXT NOT NULL"),
    ("updated_at", "TEXT NOT NULL"),
]

WORKLIST_TASK_CATEGORY = "worklist"
NORMAL_TASK_CATEGORY = "task"


def ensure_tasks_table(conn=None):
    close_conn = False
    if conn is None:
        conn = get_conn()
        close_conn = True

    cur = conn.cursor()
    column_defs = ", ".join(f'{_quote_identifier(name)} {definition}' for name, definition in TASK_TABLE_COLUMNS)
    cur.execute(f'CREATE TABLE IF NOT EXISTS {_quote_identifier("tasks")} ({column_defs})')
    cur.execute(f'CREATE INDEX IF NOT EXISTS idx_tasks_list_order ON {_quote_identifier("tasks")} ({_quote_identifier("task_list")}, {_quote_identifier("sort_order")}, {_quote_identifier("title")})')
    conn.commit()

    if close_conn:
        conn.close()


def normalize_tasks_table_categories(conn=None):
    close_conn = False
    if conn is None:
        conn = get_conn()
        close_conn = True

    cur = conn.cursor()
    cur.execute(
        f"""
        UPDATE {_quote_identifier("tasks")}
        SET {_quote_identifier("category")} = CASE
            WHEN LOWER({_quote_identifier("task_list")}) = 'template' THEN ?
            ELSE ?
        END
        WHERE {_quote_identifier("task_list")} IN ('template', 'live')
        """,
        (WORKLIST_TASK_CATEGORY, NORMAL_TASK_CATEGORY),
    )
    conn.commit()

    if close_conn:
        conn.close()


AUTH_ROLE_TABLE_COLUMNS = [
    ("id", "INTEGER PRIMARY KEY"),
    ("name", "TEXT NOT NULL UNIQUE"),
    ("description", "TEXT NOT NULL DEFAULT ''"),
    ("permissions_json", "TEXT NOT NULL DEFAULT '[]'"),
    ("is_system", "INTEGER NOT NULL DEFAULT 0"),
    ("active", "INTEGER NOT NULL DEFAULT 1"),
    ("created_at", "TEXT NOT NULL"),
    ("updated_at", "TEXT NOT NULL"),
]

AUTH_USER_TABLE_COLUMNS = [
    ("id", "INTEGER PRIMARY KEY"),
    ("signin", "TEXT NOT NULL UNIQUE"),
    ("display_name", "TEXT NOT NULL DEFAULT ''"),
    ("password_hash", "TEXT NOT NULL"),
    ("role_id", "INTEGER NOT NULL REFERENCES roles(id) ON UPDATE CASCADE ON DELETE RESTRICT"),
    ("active", "INTEGER NOT NULL DEFAULT 1"),
    ("last_login_at", "TEXT"),
    ("created_at", "TEXT NOT NULL"),
    ("updated_at", "TEXT NOT NULL"),
]

MENU_TABLE_COLUMNS = [
    ("id", "INTEGER PRIMARY KEY"),
    ("menu_key", "TEXT NOT NULL"),
    ("item_id", "TEXT NOT NULL"),
    ("position", "INTEGER NOT NULL DEFAULT 0"),
    ("back", "INTEGER NOT NULL DEFAULT 0"),
    ("darken", "INTEGER NOT NULL DEFAULT 0"),
    ("enabled", "INTEGER NOT NULL DEFAULT 1"),
    ("created_at", "TEXT NOT NULL"),
    ("updated_at", "TEXT NOT NULL"),
]

DEFAULT_AUTH_ROLES = [
    {
        "name": "Admin",
        "description": "Full access across screens, settings, and permissions.",
        "permissions": ["*"],
    },
    {
        "name": "Manager",
        "description": "Can manage day-to-day operations and shared menus.",
        "permissions": [
            "menu.view",
            "menu.manage",
            "feature.view",
            "feature.manage",
        ],
    },
    {
        "name": "User",
        "description": "Standard access for everyday work.",
        "permissions": [
            "menu.view",
            "feature.view",
        ],
    },
]


def _utc_now_iso() -> str:
    return datetime.utcnow().isoformat(timespec="seconds")


def _json_permissions(value) -> str:
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return "[]"
        try:
            parsed = json.loads(text)
        except Exception:
            parsed = [item.strip() for item in text.split(",") if item.strip()]
        return json.dumps(parsed)

    if value is None:
        return "[]"

    if isinstance(value, (list, tuple, set)):
        return json.dumps([str(item).strip() for item in value if str(item).strip()])

    return json.dumps([str(value).strip()])


def _hash_password(password: str) -> str:
    password = str(password or "")
    if not password:
        raise HTTPException(status_code=400, detail="password is required")

    salt = secrets.token_hex(16)
    iterations = 120000
    digest = hashlib.pbkdf2_hmac(
        "sha256",
        password.encode("utf-8"),
        salt.encode("utf-8"),
        iterations,
    ).hex()
    return f"pbkdf2_sha256${iterations}${salt}${digest}"


def _verify_password(password: str, stored_hash: str) -> bool:
    password = str(password or "")
    stored_hash = str(stored_hash or "")
    if not password or not stored_hash:
        return False

    parts = stored_hash.split("$")
    if len(parts) != 4:
        return hmac.compare_digest(password, stored_hash)

    algorithm, iterations_text, salt, expected = parts
    if algorithm != "pbkdf2_sha256":
        return False

    try:
        iterations = int(iterations_text)
    except ValueError:
        return False

    actual = hashlib.pbkdf2_hmac(
        "sha256",
        password.encode("utf-8"),
        salt.encode("utf-8"),
        iterations,
    ).hex()
    return hmac.compare_digest(actual, expected)


def ensure_auth_tables(conn=None):
    close_conn = False
    if conn is None:
        conn = get_conn()
        close_conn = True

    cur = conn.cursor()

    role_column_defs = ", ".join(
        f'{_quote_identifier(name)} {definition}' for name, definition in AUTH_ROLE_TABLE_COLUMNS
    )
    user_column_defs = ", ".join(
        f'{_quote_identifier(name)} {definition}' for name, definition in AUTH_USER_TABLE_COLUMNS
    )

    cur.execute(f'CREATE TABLE IF NOT EXISTS {_quote_identifier("roles")} ({role_column_defs})')
    cur.execute(f'CREATE TABLE IF NOT EXISTS {_quote_identifier("users")} ({user_column_defs})')
    cur.execute(
        f'CREATE INDEX IF NOT EXISTS idx_users_role_id ON {_quote_identifier("users")} ({_quote_identifier("role_id")})'
    )

    existing_role_columns = {
        row[1].lower()
        for row in cur.execute(f'PRAGMA table_info({_quote_identifier("roles")})').fetchall()
    }
    for column_name, column_type in AUTH_ROLE_TABLE_COLUMNS:
        if column_name.lower() in existing_role_columns:
            continue
        cur.execute(
            f'ALTER TABLE {_quote_identifier("roles")} ADD COLUMN {_quote_identifier(column_name)} {column_type}'
        )

    existing_user_columns = {
        row[1].lower()
        for row in cur.execute(f'PRAGMA table_info({_quote_identifier("users")})').fetchall()
    }
    for column_name, column_type in AUTH_USER_TABLE_COLUMNS:
        if column_name.lower() in existing_user_columns:
            continue
        cur.execute(
            f'ALTER TABLE {_quote_identifier("users")} ADD COLUMN {_quote_identifier(column_name)} {column_type}'
        )

    for role_seed in DEFAULT_AUTH_ROLES:
        permissions_json = _json_permissions(role_seed["permissions"])
        existing_role = cur.execute(
            f'SELECT id FROM {_quote_identifier("roles")} WHERE LOWER({_quote_identifier("name")}) = LOWER(?)',
            (role_seed["name"],),
        ).fetchone()
        if existing_role:
            continue

        cur.execute(
            f'INSERT INTO {_quote_identifier("roles")} ({_quote_identifier("name")}, {_quote_identifier("description")}, {_quote_identifier("permissions_json")}, {_quote_identifier("is_system")}, {_quote_identifier("active")}, {_quote_identifier("created_at")}, {_quote_identifier("updated_at")}) VALUES (?, ?, ?, 1, 1, ?, ?)',
            (
                role_seed["name"],
                role_seed["description"],
                permissions_json,
                _utc_now_iso(),
                _utc_now_iso(),
            ),
        )

    conn.commit()

    if close_conn:
        conn.close()


def ensure_menu_table(conn=None):
    close_conn = False
    if conn is None:
        conn = get_conn()
        close_conn = True

    cur = conn.cursor()

    menu_column_defs = ", ".join(
        f'{_quote_identifier(name)} {definition}' for name, definition in MENU_TABLE_COLUMNS
    )

    cur.execute(f'CREATE TABLE IF NOT EXISTS {_quote_identifier("Menu")} ({menu_column_defs})')
    cur.execute(
        f'CREATE INDEX IF NOT EXISTS idx_menu_menu_key ON {_quote_identifier("Menu")} ({_quote_identifier("menu_key")}, {_quote_identifier("position")})'
    )
    cur.execute(
        f'CREATE UNIQUE INDEX IF NOT EXISTS idx_menu_menu_key_item_id ON {_quote_identifier("Menu")} ({_quote_identifier("menu_key")}, {_quote_identifier("item_id")})'
    )

    existing_menu_columns = {
        row[1].lower()
        for row in cur.execute(f'PRAGMA table_info({_quote_identifier("Menu")})').fetchall()
    }
    for column_name, column_type in MENU_TABLE_COLUMNS:
        if column_name.lower() in existing_menu_columns:
            continue
        cur.execute(
            f'ALTER TABLE {_quote_identifier("Menu")} ADD COLUMN {_quote_identifier(column_name)} {column_type}'
        )

    conn.commit()

    if close_conn:
        conn.close()


def _menu_row_to_payload(row):
    return {
        "id": row["id"],
        "menu_key": row["menu_key"],
        "item_id": row["item_id"],
        "position": row["position"],
        "back": bool(row["back"]),
        "darken": bool(row["darken"]),
        "enabled": bool(row["enabled"]),
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


def _normalize_menu_key(menu_key: str) -> str:
    value = str(menu_key or "").strip()
    if not value:
        raise HTTPException(status_code=400, detail="menu_key is required")
    return value


def _normalize_menu_selection(selection):
    if selection is None:
        return []

    if not isinstance(selection, list):
        raise HTTPException(status_code=400, detail="selection must be a list")

    normalized = []
    for entry in selection:
        if not isinstance(entry, dict):
            raise HTTPException(status_code=400, detail="selection entries must be objects")

        item_id = str(entry.get("id") or "").strip()
        if not item_id:
            raise HTTPException(status_code=400, detail="selection entry id is required")

        normalized.append(
            {
                "id": item_id,
                "back": bool(entry.get("back")),
                "darken": bool(entry.get("darken")),
                "enabled": True if entry.get("enabled") is None else bool(entry.get("enabled")),
            }
        )

    unique = []
    seen = set()
    for entry in normalized:
        if entry["id"] in seen:
            continue
        seen.add(entry["id"])
        unique.append(entry)
    return unique


def _replace_menu_rows(conn, menu_key: str, selection):
    normalized_key = _normalize_menu_key(menu_key)
    normalized_selection = _normalize_menu_selection(selection)
    timestamp = _utc_now_iso()

    cur = conn.cursor()
    cur.execute(f'DELETE FROM {_quote_identifier("Menu")} WHERE {_quote_identifier("menu_key")} = ?', (normalized_key,))

    for position, entry in enumerate(normalized_selection):
        cur.execute(
            f'INSERT INTO {_quote_identifier("Menu")} ({_quote_identifier("menu_key")}, {_quote_identifier("item_id")}, {_quote_identifier("position")}, {_quote_identifier("back")}, {_quote_identifier("darken")}, {_quote_identifier("enabled")}, {_quote_identifier("created_at")}, {_quote_identifier("updated_at")}) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
            (
                normalized_key,
                entry["id"],
                position,
                1 if entry["back"] else 0,
                1 if entry["darken"] else 0,
                1 if entry["enabled"] else 0,
                timestamp,
                timestamp,
            ),
        )


def _role_row_to_payload(row):
    permissions = []
    try:
        permissions = json.loads(row["permissions_json"] or "[]")
    except Exception:
        permissions = []

    return {
        "id": row["id"],
        "name": row["name"],
        "description": row["description"],
        "permissions": permissions,
        "is_system": bool(row["is_system"]),
        "active": bool(row["active"]),
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


def _user_row_to_payload(row):
    return {
        "id": row["id"],
        "signin": row["signin"],
        "display_name": row["display_name"],
        "role_id": row["role_id"],
        "role_name": row["role_name"],
        "active": bool(row["active"]),
        "last_login_at": row["last_login_at"],
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


@app.get("/menu")
def list_menu_entries():
    conn = get_conn()
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            f'''
            SELECT *
            FROM {_quote_identifier("Menu")}
            ORDER BY {_quote_identifier("menu_key")} ASC, {_quote_identifier("position")} ASC, {_quote_identifier("id")} ASC
            '''
        ).fetchall()
        return [_menu_row_to_payload(row) for row in rows]
    finally:
        conn.close()


@app.get("/menu/{menu_key:path}")
def get_menu_entries(menu_key: str):
    normalized_key = _normalize_menu_key(menu_key)
    conn = get_conn()
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            f'''
            SELECT *
            FROM {_quote_identifier("Menu")}
            WHERE {_quote_identifier("menu_key")} = ?
            ORDER BY {_quote_identifier("position")} ASC, {_quote_identifier("id")} ASC
            ''',
            (normalized_key,),
        ).fetchall()
        return [_menu_row_to_payload(row) for row in rows]
    finally:
        conn.close()


@app.put("/menu/{menu_key:path}")
def save_menu_entries(menu_key: str, payload: dict):
    conn = get_conn()
    try:
        _replace_menu_rows(conn, menu_key, payload.get("selection"))
        conn.commit()
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            f'''
            SELECT *
            FROM {_quote_identifier("Menu")}
            WHERE {_quote_identifier("menu_key")} = ?
            ORDER BY {_quote_identifier("position")} ASC, {_quote_identifier("id")} ASC
            ''',
            (_normalize_menu_key(menu_key),),
        ).fetchall()
        return [_menu_row_to_payload(row) for row in rows]
    finally:
        conn.close()


@app.delete("/menu/{menu_key:path}")
def delete_menu_entries(menu_key: str):
    normalized_key = _normalize_menu_key(menu_key)
    conn = get_conn()
    try:
        conn.execute(
            f'DELETE FROM {_quote_identifier("Menu")} WHERE {_quote_identifier("menu_key")} = ?',
            (normalized_key,),
        )
        conn.commit()
        return {"ok": True}
    finally:
        conn.close()


@app.delete("/menu")
def delete_all_menu_entries():
    conn = get_conn()
    try:
        conn.execute(f'DELETE FROM {_quote_identifier("Menu")}')
        conn.commit()
        return {"ok": True}
    finally:
        conn.close()


def _fetch_role_or_404(conn, role_id: int):
    role = conn.execute(
        f'SELECT * FROM {_quote_identifier("roles")} WHERE id = ?',
        (role_id,),
    ).fetchone()
    if not role:
        raise HTTPException(status_code=404, detail="Role not found")
    return role


def _create_user_record(conn, payload: dict):
    signin = str(payload.get("signin") or "").strip()
    password = str(payload.get("password") or "")
    role_id = payload.get("role_id")
    display_name = str(payload.get("display_name") or "").strip()

    if not signin:
        raise HTTPException(status_code=400, detail="signin is required")
    if role_id in (None, ""):
        raise HTTPException(status_code=400, detail="role_id is required")

    try:
        role_id = int(role_id)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="role_id must be a number")

    _fetch_role_or_404(conn, role_id)

    password_hash = _hash_password(password)
    now = _utc_now_iso()
    cur = conn.cursor()
    cur.execute(
        f'INSERT INTO {_quote_identifier("users")} ({_quote_identifier("signin")}, {_quote_identifier("display_name")}, {_quote_identifier("password_hash")}, {_quote_identifier("role_id")}, {_quote_identifier("active")}, {_quote_identifier("created_at")}, {_quote_identifier("updated_at")}) VALUES (?, ?, ?, ?, ?, ?, ?)',
        (signin, display_name, password_hash, role_id, 1 if payload.get("active", True) else 0, now, now),
    )
    return cur.lastrowid


def _get_user_by_signin(conn, signin: str):
    return conn.execute(
        f"""
        SELECT
            u.id,
            u.signin,
            u.display_name,
            u.password_hash,
            u.role_id,
            u.active,
            u.last_login_at,
            u.created_at,
            u.updated_at,
            r.name AS role_name,
            r.description AS role_description,
            r.permissions_json AS role_permissions_json
        FROM {_quote_identifier("users")} u
        LEFT JOIN {_quote_identifier("roles")} r ON r.id = u.role_id
        WHERE LOWER(u.signin) = LOWER(?)
        """,
        (signin,),
    ).fetchone()


def _balsheet_order_clause() -> str:
    entry_id = _quote_identifier("EntryID")
    posting_date = _quote_identifier("PostingDate")
    return (
        f"ORDER BY {posting_date} ASC, "
        f"CASE WHEN instr({entry_id}, '-') > 0 THEN CAST(substr({entry_id}, instr({entry_id}, '-') + 1) AS INTEGER) ELSE 0 END ASC, "
        f"{entry_id} ASC"
    )


def _normalize_balsheet_amount(value):
    try:
        return float(str(value).replace("$", "").replace(",", "").strip() or 0)
    except Exception:
        return 0.0


def _generate_balsheet_entry_id() -> str:
    return f"BS-{datetime.now().strftime('%m%d%Y-%H%M%S%f')}"


def _balsheet_row_to_payload(row):
    return {
        "entry_id": _balsheet_row_value(row, "EntryID", 0),
        "posting_date": normalize_mmddyyyy(_balsheet_row_value(row, "PostingDate", 1))
        or str(_balsheet_row_value(row, "PostingDate", 1) or ""),
        "type": str(_balsheet_row_value(row, "Type", 2) or ""),
        "amount": _balsheet_row_value(row, "Amount", 3),
        "payer": str(_balsheet_row_value(row, "Payer", 4) or ""),
        "check_number": str(_balsheet_row_value(row, "Check Number", 5) or ""),
        "edi": str(_balsheet_row_value(row, "EDI", 6) or ""),
        "poster": str(_balsheet_row_value(row, "Poster", 7) or ""),
        "eob": str(_balsheet_row_value(row, "EOB", 8) or ""),
        "unposted": _balsheet_row_value(row, "UnPosted", 9),
        "misc": _balsheet_row_value(row, "Misc", 10),
        "misc_type": str(_balsheet_row_value(row, "Misc-Type", 11) or ""),
        "notes": str(_balsheet_row_value(row, "Notes", 12) or ""),
        "nick": _balsheet_row_value(row, "Nick", 13),
        "raul": _balsheet_row_value(row, "Raul", 14),
        "needs": str(_balsheet_row_value(row, "Needs", 15) or ""),
        "from_date": str(_balsheet_row_value(row, "From", 16) or ""),
        "to_date": str(_balsheet_row_value(row, "To", 17) or ""),
    }


def _normalize_review_value(field_name: str, value):
    if field_name in {"amount", "unposted", "misc", "nick", "raul"}:
        return round(_normalize_balsheet_amount(value), 2)
    return str(value or "").strip()


def _normalize_review_keyproof_payload(payload):
    if not isinstance(payload, dict):
        return {}

    form = payload.get("form")
    return form if isinstance(form, dict) else {}


def _load_saved_itemization_rows(conn, attachment_id: int):
    row = conn.execute(
        f'''
        SELECT payload_json
        FROM {_quote_identifier("itemization")}
        WHERE {_quote_identifier("attachment_id")} = ?
        ''',
        (attachment_id,),
    ).fetchone()

    if not row or not row[0]:
        return []

    try:
        payload = json.loads(row[0])
    except Exception:
        return []

    items = payload.get("items", []) if isinstance(payload, dict) else []
    return items if isinstance(items, list) else []


def _load_balsheet_review_day_rows(conn, posting_date: str):
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        f'''
        SELECT *
        FROM {_quote_identifier("Balsheet")}
        WHERE {_quote_identifier("PostingDate")} = ?
        {_balsheet_order_clause()}
        ''',
        (posting_date,),
    ).fetchall()
    return [_balsheet_row_to_payload(row) for row in rows]


def _normalize_review_site_key(value):
    return " ".join(str(value or "").strip().lower().split())


def _live_balsheet_total_for_site(balsheet_rows, site_name):
    normalized_site = _normalize_review_site_key(site_name)
    total = 0.0

    for row in balsheet_rows:
        row_type = _normalize_review_site_key(row.get("type"))
        if row_type == normalized_site:
            total += _normalize_balsheet_amount(row.get("amount"))
        elif "spring lane" in normalized_site and row_type in {"eft", "lockbox"}:
            total += _normalize_balsheet_amount(row.get("amount"))

    return round(total, 2)


def _is_close_to_zero(value: float, tolerance: float = 0.005) -> bool:
    return abs(round(value, 2)) < tolerance


@app.get("/balsheet/keyproof-review")
def get_balsheet_keyproof_review(posting_date: str | None = None):
    init_db()
    conn = get_conn()
    conn.row_factory = sqlite3.Row
    ensure_balsheet_table(conn)
    ensure_keyproof_table(conn)
    ensure_itemization_table(conn)

    try:
        normalized_posting_date = normalize_mmddyyyy(posting_date) if posting_date else normalize_mmddyyyy(get_current_work_day() or "") or normalize_mmddyyyy(get_current_bank_day() or "") or datetime.today().strftime("%m/%d/%Y")
        if not normalized_posting_date:
            raise HTTPException(status_code=400, detail="posting_date is required")

        balsheet_rows = _load_balsheet_review_day_rows(conn, normalized_posting_date)

        candidates = conn.execute(
            f'''
            SELECT
                f.id,
                f.filename,
                f.site,
                f.batch_date,
                f.review_status,
                k.payload_json AS keyproof_payload_json
            FROM {_quote_identifier("imported_files")} f
            INNER JOIN {_quote_identifier("keyproof")} k ON k.attachment_id = f.id
            WHERE {_quote_identifier("batch_date")} IS NOT NULL
            ORDER BY f.batch_date ASC, f.site ASC, f.id ASC
            '''
        ).fetchall()

        review_rows = []
        for candidate in candidates:
            candidate_day = normalize_mmddyyyy(candidate["batch_date"]) or str(candidate["batch_date"] or "")
            if candidate_day != normalized_posting_date:
                continue

            try:
                keyproof_payload = json.loads(candidate["keyproof_payload_json"]) if candidate["keyproof_payload_json"] else None
            except Exception:
                keyproof_payload = None

            keyproof_form = _normalize_review_keyproof_payload(keyproof_payload)
            keyproof_total = _keyproof_total_from_payload(keyproof_payload)
            eft_expected = _normalize_balsheet_amount(keyproof_form.get("eft"))
            lockbox_expected = _normalize_balsheet_amount(keyproof_form.get("lockbox"))

            itemization_rows = _load_saved_itemization_rows(conn, candidate["id"])
            itemization_total = 0.0
            for item in itemization_rows:
                if isinstance(item, dict):
                    itemization_total += _normalize_balsheet_amount(item.get("amount"))

            itemization_difference = round(itemization_total - keyproof_total, 2)
            itemization_status = "no_itemization" if not itemization_rows else ("matched" if _is_close_to_zero(itemization_difference) else "partial")

            eft_balsheet_total = 0.0
            lockbox_balsheet_total = 0.0
            for row in balsheet_rows:
                row_type = str(row.get("type") or "").strip().lower()
                if row_type == "eft":
                    eft_balsheet_total += _normalize_balsheet_amount(row.get("amount"))
                elif row_type == "lockbox":
                    lockbox_balsheet_total += _normalize_balsheet_amount(row.get("amount"))

            balsheet_total = _live_balsheet_total_for_site(balsheet_rows, candidate["site"])
            balsheet_difference = round(keyproof_total - balsheet_total, 2)
            balsheet_needed = not (_is_close_to_zero(keyproof_total) and _is_close_to_zero(balsheet_total))
            if not balsheet_needed:
                balsheet_status = "not_applicable"
            elif _is_close_to_zero(balsheet_difference):
                balsheet_status = "matched"
            elif keyproof_total > 0 and balsheet_total > 0:
                balsheet_status = "partial"
            elif keyproof_total > 0 and _is_close_to_zero(balsheet_total):
                balsheet_status = "missing"
            else:
                balsheet_status = "partial"

            spring_lane_expected_total = round(eft_expected + lockbox_expected, 2)
            spring_lane_balsheet_total = round(eft_balsheet_total + lockbox_balsheet_total, 2)
            spring_lane_difference = round(spring_lane_balsheet_total - spring_lane_expected_total, 2)
            spring_lane_needed = not (_is_close_to_zero(spring_lane_expected_total) and _is_close_to_zero(spring_lane_balsheet_total))
            if not spring_lane_needed:
                spring_lane_status = "not_applicable"
            elif _is_close_to_zero(spring_lane_difference):
                spring_lane_status = "matched"
            elif spring_lane_balsheet_total > 0 and spring_lane_expected_total > 0:
                spring_lane_status = "partial"
            elif spring_lane_expected_total > 0 and _is_close_to_zero(spring_lane_balsheet_total):
                spring_lane_status = "missing"
            else:
                spring_lane_status = "partial"

            if balsheet_status == "not_applicable":
                status = "no_itemization"
            elif balsheet_status == "partial" or balsheet_status == "missing" or spring_lane_status in {"partial", "missing"}:
                status = "partial"
            else:
                status = "matched"

            site_row = {
                "attachmentId": candidate["id"],
                "filename": candidate["filename"],
                "site": candidate["site"],
                "sourceSite": candidate["site"],
                "rowKind": "site",
                "batchDate": normalize_mmddyyyy(candidate["batch_date"]) or str(candidate["batch_date"] or ""),
                "reviewStatus": candidate["review_status"],
                "keyproofTotal": round(keyproof_total, 2),
                "eftExpectedTotal": round(eft_expected, 2),
                "lockboxExpectedTotal": round(lockbox_expected, 2),
                "eftBalsheetTotal": round(eft_balsheet_total, 2),
                "lockboxBalsheetTotal": round(lockbox_balsheet_total, 2),
                "itemizationBalsheetTotal": round(itemization_total, 2),
                "itemizationDifference": itemization_difference,
                "itemizationStatus": itemization_status,
                "balsheetActualTotal": round(balsheet_total, 2),
                "balsheetDifference": balsheet_difference,
                "balsheetStatus": balsheet_status,
                "springLaneExpectedTotal": spring_lane_expected_total,
                "springLaneBalsheetTotal": spring_lane_balsheet_total,
                "springLaneDifference": spring_lane_difference,
                "springLaneStatus": spring_lane_status,
                "balsheetRowCount": len(balsheet_rows),
                "status": status,
            }
            review_rows.append(site_row)

        itemization_matched_count = sum(1 for row in review_rows if row["itemizationStatus"] == "matched")
        itemization_partial_count = sum(1 for row in review_rows if row["itemizationStatus"] == "partial")
        itemization_missing_count = sum(1 for row in review_rows if row["itemizationStatus"] == "no_itemization")
        balsheet_matched_count = sum(1 for row in review_rows if row["balsheetStatus"] == "matched")
        balsheet_partial_count = sum(1 for row in review_rows if row["balsheetStatus"] == "partial")
        balsheet_missing_count = sum(1 for row in review_rows if row["balsheetStatus"] == "missing")
        spring_lane_matched_count = sum(1 for row in review_rows if row["springLaneStatus"] == "matched")
        spring_lane_partial_count = sum(1 for row in review_rows if row["springLaneStatus"] == "partial")
        spring_lane_missing_count = sum(1 for row in review_rows if row["springLaneStatus"] == "missing")
        needs_review_count = sum(1 for row in review_rows if row["status"] != "matched")

        return {
            "postingDate": normalized_posting_date,
            "balsheetRowCount": len(balsheet_rows),
            "keyproofCount": len(review_rows),
            "needsReviewCount": needs_review_count,
            "itemizationMatchedCount": itemization_matched_count,
            "itemizationPartialCount": itemization_partial_count,
            "itemizationMissingCount": itemization_missing_count,
            "balsheetMatchedCount": balsheet_matched_count,
            "balsheetPartialCount": balsheet_partial_count,
            "balsheetMissingCount": balsheet_missing_count,
            "springLaneMatchedCount": spring_lane_matched_count,
            "springLanePartialCount": spring_lane_partial_count,
            "springLaneMissingCount": spring_lane_missing_count,
            "rows": review_rows,
        }
    finally:
        conn.close()


@app.get("/balsheet/keyproof-review-open")
def get_balsheet_keyproof_review_open():
    init_db()
    conn = get_conn()
    conn.row_factory = sqlite3.Row
    ensure_balsheet_table(conn)
    ensure_keyproof_table(conn)

    try:
        candidates = conn.execute(
            f'''
            SELECT
                f.id,
                f.filename,
                f.site,
                f.batch_date,
                f.review_status,
                k.payload_json AS keyproof_payload_json
            FROM {_quote_identifier("imported_files")} f
            INNER JOIN {_quote_identifier("keyproof")} k ON k.attachment_id = f.id
            WHERE {_quote_identifier("batch_date")} IS NOT NULL
            ORDER BY f.batch_date DESC, f.site ASC, f.id ASC
            '''
        ).fetchall()

        balsheet_rows_by_day: dict[str, list[dict]] = {}
        issue_rows: list[dict] = []
        open_balance_total = 0.0

        for candidate in candidates:
            candidate_day = normalize_mmddyyyy(candidate["batch_date"]) or str(candidate["batch_date"] or "")
            if not candidate_day:
                continue

            balsheet_rows = balsheet_rows_by_day.get(candidate_day)
            if balsheet_rows is None:
                balsheet_rows = _load_balsheet_review_day_rows(conn, candidate_day)
                balsheet_rows_by_day[candidate_day] = balsheet_rows

            try:
                keyproof_payload = json.loads(candidate["keyproof_payload_json"]) if candidate["keyproof_payload_json"] else None
            except Exception:
                keyproof_payload = None

            keyproof_form = _normalize_review_keyproof_payload(keyproof_payload)
            keyproof_total = _keyproof_total_from_payload(keyproof_payload)
            eft_expected = _normalize_balsheet_amount(keyproof_form.get("eft"))
            lockbox_expected = _normalize_balsheet_amount(keyproof_form.get("lockbox"))
            spring_lane_expected_total = round(eft_expected + lockbox_expected, 2)

            site_name = str(candidate["site"] or "")
            keyproof_subtotal = round(
                keyproof_total + spring_lane_expected_total if "spring lane" in site_name.lower() else keyproof_total,
                2,
            )
            balsheet_total = _live_balsheet_total_for_site(balsheet_rows, site_name)
            difference = round(keyproof_subtotal - balsheet_total, 2)
            if _is_close_to_zero(difference):
                continue

            issue_rows.append(
                {
                    "attachmentId": candidate["id"],
                    "filename": candidate["filename"],
                    "site": candidate["site"],
                    "batchDate": normalize_mmddyyyy(candidate["batch_date"]) or str(candidate["batch_date"] or ""),
                    "reviewStatus": candidate["review_status"],
                    "keyproofTotal": keyproof_subtotal,
                    "balsheetActualTotal": round(balsheet_total, 2),
                    "difference": difference,
                }
            )
            open_balance_total += abs(difference)

        return {
            "openCount": len(issue_rows),
            "postingDateCount": len({row["batchDate"] for row in issue_rows if row.get("batchDate")}),
            "openBalanceTotal": round(open_balance_total, 2),
            "rows": issue_rows,
        }
    finally:
        conn.close()


def _balsheet_note_row_to_payload(row):
    return {
        "rowid": row["rowid"],
        "post_date": normalize_mmddyyyy(row["post_date"]) or str(row["post_date"] or ""),
        "notes": str(row["notes"] or ""),
        "message": str(row["message"] or ""),
    }


def _normalize_balsheet_payload(entry: dict, entry_id: str | None = None):
    posting_date = normalize_mmddyyyy(entry.get("posting_date")) or ""
    if not posting_date:
        raise HTTPException(status_code=400, detail="posting_date is required")

    amount = _normalize_balsheet_amount(entry.get("amount"))
    unposted = _normalize_balsheet_amount(entry.get("unposted"))
    misc = _normalize_balsheet_amount(entry.get("misc"))
    poster = str(entry.get("poster") or "").strip()
    poster_key = poster.lower() or "nick"
    poster_amount = amount - unposted - misc
    nick = poster_amount if poster_key == "nick" else 0.0
    raul = poster_amount if poster_key == "raul" else 0.0

    return {
        "EntryID": entry_id or str(entry.get("entry_id") or "").strip() or _generate_balsheet_entry_id(),
        "PostingDate": posting_date,
        "Type": str(entry.get("type") or "").strip(),
        "Amount": amount,
        "Payer": str(entry.get("payer") or "").strip(),
        "Check Number": str(entry.get("check_number") or "").strip(),
        "EDI": str(entry.get("edi") or "").strip(),
        "Poster": poster,
        "EOB": str(entry.get("eob") or "").strip(),
        "UnPosted": unposted,
        "Misc": misc,
        "Misc-Type": str(entry.get("misc_type") or "").strip(),
        "Notes": str(entry.get("notes") or "").strip(),
        "Nick": nick,
        "Raul": raul,
        "Needs": str(entry.get("needs") or "").strip(),
        "From": str(entry.get("from_date") or "").strip(),
        "To": str(entry.get("to_date") or "").strip(),
    }


def _normalize_balsheet_note_payload(note: dict, rowid: int | None = None):
    post_date = normalize_mmddyyyy(note.get("post_date")) or ""
    if not post_date:
        raise HTTPException(status_code=400, detail="post_date is required")

    return {
        "rowid": rowid,
        "post_date": post_date,
        "notes": str(note.get("notes") or "").strip(),
        "message": str(note.get("message") or "").strip(),
    }


def _balsheet_insert_or_replace(conn, entry: dict):
    normalized = _normalize_balsheet_payload(entry)
    columns = [name for name, _ in BALSHEET_TABLE_COLUMNS]
    quoted_columns = ", ".join(_quote_identifier(name) for name in columns)
    placeholders = ", ".join(["?"] * len(columns))
    conn.execute(
        f'INSERT OR REPLACE INTO {_quote_identifier("Balsheet")} ({quoted_columns}) VALUES ({placeholders})',
        tuple(normalized[column] for column in columns),
    )
    return normalized["EntryID"]


def _list_user_tables(conn):
    cur = conn.cursor()
    cur.execute("""
        SELECT name
        FROM sqlite_master
        WHERE type = 'table'
          AND name NOT LIKE 'sqlite_%'
        ORDER BY name
    """)
    return [row[0] for row in cur.fetchall()]


def _table_exists(conn, table_name: str) -> bool:
    return table_name in _list_user_tables(conn)


def _parse_calendar_date(value):
    normalized = normalize_mmddyyyy(value)
    if not normalized:
        return None

    try:
        return datetime.strptime(normalized, "%m/%d/%Y")
    except ValueError:
        return None


def _parse_amount(value):
    try:
        if value in (None, ""):
            return 0.0
        return float(str(value).replace("$", "").replace(",", "").strip())
    except Exception:
        return 0.0


def _normalize_yyyy_mm_dd_to_mmddyyyy(value):
    normalized = normalize_mmddyyyy(value)
    if normalized:
        return normalized

    try:
        parsed = pd.to_datetime(value, errors="coerce")
    except Exception:
        return ""

    if pd.isna(parsed):
        return ""

    return parsed.strftime("%m/%d/%Y")


def _parse_eft_descriptive_text_1(value):
    text = str(value or "").strip()
    if not text:
        return "", ""

    match = re.search(r"^(.*?)TRN\*1\*(.*)$", text, flags=re.IGNORECASE)
    if not match:
        return "", ""

    payer = match.group(1).strip()
    remainder = match.group(2).strip()
    check_number = remainder.split("*", 1)[0].strip()
    return payer, check_number


def _replace_table_from_dataframe(conn, table_name: str, df: pd.DataFrame):
    cur = conn.cursor()
    cur.execute(f"DELETE FROM {_quote_identifier(table_name)}")
    if df.empty:
        return 0

    columns = list(df.columns)
    quoted_columns = ", ".join(_quote_identifier(column) for column in columns)
    placeholders = ", ".join(["?"] * len(columns))
    insert_sql = f"INSERT INTO {_quote_identifier(table_name)} ({quoted_columns}) VALUES ({placeholders})"
    clean_df = df.where(pd.notna(df), None)
    cur.executemany(insert_sql, clean_df.itertuples(index=False, name=None))
    return int(len(clean_df))


def _append_table_from_dataframe(conn, table_name: str, df: pd.DataFrame):
    if df.empty:
        return 0

    cur = conn.cursor()
    columns = list(df.columns)
    quoted_columns = ", ".join(_quote_identifier(column) for column in columns)
    placeholders = ", ".join(["?"] * len(columns))
    insert_sql = f"INSERT INTO {_quote_identifier(table_name)} ({quoted_columns}) VALUES ({placeholders})"
    clean_df = df.where(pd.notna(df), None)
    cur.executemany(insert_sql, clean_df.itertuples(index=False, name=None))
    return int(len(clean_df))


def _normalize_task_payload(task: dict, task_id: str | None = None, sort_order: int | None = None):
    title = str(task.get("title") or "").strip()
    if not title:
        raise HTTPException(status_code=400, detail="title is required")

    task_list = str(task.get("task_list") or "live").strip().lower() or "live"
    recurrence = str(task.get("recurrence") or "none").strip().lower()
    if recurrence not in {"none", "daily", "weekly", "monthly"}:
        recurrence = "none"

    action_type = str(task.get("action_type") or "none").strip().lower()
    if action_type not in {"none", "url", "copy", "copy_details"}:
        action_type = "none"

    now = datetime.now().isoformat(timespec="seconds")
    next_due_at = str(task.get("next_due_at") or "").strip() or None
    completed_at = str(task.get("completed_at") or "").strip() or None

    return {
        "task_id": str(task_id or task.get("task_id") or "").strip() or f"task-{datetime.now().strftime('%Y%m%d%H%M%S%f')}",
        "task_list": task_list,
        "title": title,
        "details": str(task.get("details") or "").strip(),
        "category": WORKLIST_TASK_CATEGORY if task_list == "template" else NORMAL_TASK_CATEGORY,
        "recurrence": recurrence,
        "action_type": action_type,
        "action_label": str(task.get("action_label") or "").strip(),
        "action_value": str(task.get("action_value") or "").strip(),
        "done": 1 if bool(task.get("done")) else 0,
        "sort_order": int(sort_order if sort_order is not None else task.get("sort_order") or 0),
        "next_due_at": next_due_at,
        "completed_at": completed_at,
        "created_at": str(task.get("created_at") or now).strip(),
        "updated_at": str(task.get("updated_at") or now).strip(),
    }


def _task_row_to_payload(row):
    return {
        "id": str(row["task_id"]),
        "task_list": str(row["task_list"] or ""),
        "title": str(row["title"] or ""),
        "details": str(row["details"] or ""),
        "category": str(row["category"] or ""),
        "recurrence": str(row["recurrence"] or "none"),
        "action_type": str(row["action_type"] or "none"),
        "action_label": str(row["action_label"] or ""),
        "action_value": str(row["action_value"] or ""),
        "done": bool(row["done"]),
        "sort_order": int(row["sort_order"] or 0),
        "next_due_at": str(row["next_due_at"] or "") or None,
        "completed_at": str(row["completed_at"] or "") or None,
        "created_at": str(row["created_at"] or ""),
        "updated_at": str(row["updated_at"] or ""),
    }


def _normalize_eft_key_columns(df: pd.DataFrame, key_columns: list[str]) -> pd.DataFrame:
    normalized = df.loc[:, key_columns].copy()
    for column in key_columns:
        if column == "Amount":
            normalized[column] = pd.to_numeric(normalized[column], errors="coerce").round(2)
        else:
            normalized[column] = normalized[column].fillna("").astype(str).str.strip()
    return normalized


def _all_rows_already_exist(conn, table_name: str, df: pd.DataFrame, key_columns: list[str]) -> bool:
    if df.empty:
        return False

    select_columns = ", ".join(_quote_identifier(column) for column in key_columns)
    existing_df = pd.read_sql_query(
        f"SELECT {select_columns} FROM {_quote_identifier(table_name)}",
        conn,
    )
    if existing_df.empty:
        return False

    candidate_df = _normalize_eft_key_columns(df, key_columns)
    existing_keys = _normalize_eft_key_columns(existing_df, key_columns).drop_duplicates()
    merged = candidate_df.merge(existing_keys, on=key_columns, how="left", indicator=True)
    return bool(len(merged) > 0 and (merged["_merge"] == "both").all())


def _all_rows_already_exist_as_text(conn, table_name: str, df: pd.DataFrame, key_columns: list[str]) -> bool:
    if df.empty:
        return False

    select_columns = ", ".join(_quote_identifier(column) for column in key_columns)
    existing_df = pd.read_sql_query(
        f"SELECT {select_columns} FROM {_quote_identifier(table_name)}",
        conn,
    )
    if existing_df.empty:
        return False

    candidate_df = df.loc[:, key_columns].copy()
    existing_df = existing_df.loc[:, key_columns].copy()

    for column in key_columns:
        candidate_df[column] = candidate_df[column].fillna("").astype(str).str.strip()
        existing_df[column] = existing_df[column].fillna("").astype(str).str.strip()

    merged = candidate_df.merge(existing_df.drop_duplicates(), on=key_columns, how="left", indicator=True)
    return bool(len(merged) > 0 and (merged["_merge"] == "both").all())


def _load_calendar_rows(conn):
    cur = conn.cursor()
    cur.execute("""
        SELECT bank_day, weekday, is_closed, closure_reason, paperwork_day
        FROM calendar
    """)

    rows = []
    for bank_day, weekday, is_closed, closure_reason, paperwork_day in cur.fetchall():
        rows.append({
            "bankDay": normalize_mmddyyyy(bank_day) or bank_day,
            "weekday": weekday,
            "isClosed": bool(is_closed),
            "closureReason": closure_reason or "",
            "paperworkDay": normalize_mmddyyyy(paperwork_day) or paperwork_day,
            "_bankSort": _parse_calendar_date(bank_day),
            "_paperSort": _parse_calendar_date(paperwork_day),
        })

    rows.sort(key=lambda row: row["_bankSort"] or datetime.max)
    return rows


def _load_task_rows(conn, task_list: str = "live"):
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        f"""
        SELECT *
        FROM {_quote_identifier("tasks")}
        WHERE {_quote_identifier("task_list")} = ?
        ORDER BY {_quote_identifier("sort_order")} ASC, {_quote_identifier("created_at")} ASC, {_quote_identifier("task_id")} ASC
        """,
        (task_list,),
    ).fetchall()
    return [_task_row_to_payload(row) for row in rows]


def _live_cashing_totals(conn):
    totals = defaultdict(lambda: {
        "lockboxTotal": 0.0,
        "lockboxCount": 0,
        "eftTotal": 0.0,
        "eftCount": 0,
    })

    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT [Transaction Total] AS amount,
                   [Deposit Date] AS deposit_date
            FROM Lockbox
        """)
        for amount, deposit_date in cur.fetchall():
            day = normalize_mmddyyyy(deposit_date)
            if not day:
                continue
            totals[day]["lockboxTotal"] += _parse_amount(amount)
            totals[day]["lockboxCount"] += 1
    except Exception:
        pass

    try:
        cur.execute("""
            SELECT Date AS as_of_date,
                   Amount AS amount
            FROM EFT
        """)
        for as_of_date, amount in cur.fetchall():
            day = normalize_mmddyyyy(as_of_date)
            if not day:
                continue
            totals[day]["eftTotal"] += _parse_amount(amount)
            totals[day]["eftCount"] += 1
    except Exception:
        pass

    return totals


def _calendar_status_payload():
    conn = get_conn()
    init_db()

    current_work_day = get_current_work_day()
    current_bank_day = get_current_bank_day()
    today = datetime.today().strftime("%m/%d/%Y")
    rows = _load_calendar_rows(conn)

    total_days = len(rows)
    open_days = sum(1 for row in rows if not row["isClosed"])
    closed_days = total_days - open_days

    today_bank_day = None
    for row in rows:
        if row["paperworkDay"] == today:
            today_bank_day = row["bankDay"]
            break

    current_sort = _parse_calendar_date(current_work_day) if current_work_day else None
    next_open_work_day = None

    for row in rows:
        if current_sort and row["_paperSort"] and row["_paperSort"] > current_sort and not row["isClosed"]:
            next_open_work_day = row["paperworkDay"]
            break

    last_bank_day = rows[-1]["bankDay"] if rows else None

    conn.close()
    return {
        "today": today,
        "currentWorkDay": current_work_day,
        "currentBankDay": current_bank_day,
        "todayBankDay": today_bank_day,
        "nextOpenWorkDay": next_open_work_day,
        "lastBankDay": last_bank_day,
        "totalDays": total_days,
        "openDays": open_days,
        "closedDays": closed_days,
    }


def _calendar_range_payload(start_str, end_str):
    start_norm = normalize_mmddyyyy(start_str)
    end_norm = normalize_mmddyyyy(end_str)

    if not start_norm or not end_norm:
        raise HTTPException(status_code=400, detail="Start and end dates must be valid dates")

    start_dt = _parse_calendar_date(start_norm)
    end_dt = _parse_calendar_date(end_norm)

    if not start_dt or not end_dt:
        raise HTTPException(status_code=400, detail="Start and end dates must be valid dates")

    if start_dt > end_dt:
        start_dt, end_dt = end_dt, start_dt
        start_norm, end_norm = end_norm, start_norm

    conn = get_conn()
    init_db()

    calendar_rows = _load_calendar_rows(conn)
    totals = _live_cashing_totals(conn)
    current_work_day = get_current_work_day()

    rows = []
    for row in calendar_rows:
      bank_dt = row["_bankSort"]
      if not bank_dt or bank_dt < start_dt or bank_dt > end_dt:
          continue

      source = totals.get(row["bankDay"], {})
      lockbox_total = float(source.get("lockboxTotal", 0.0))
      eft_total = float(source.get("eftTotal", 0.0))
      rows.append({
          "bankDay": row["bankDay"],
          "weekday": row["weekday"],
          "isClosed": row["isClosed"],
          "closureReason": row["closureReason"],
          "paperworkDay": row["paperworkDay"],
          "isCurrentWorkDay": row["paperworkDay"] == current_work_day,
          "lockboxTotal": round(lockbox_total, 2),
          "lockboxCount": int(source.get("lockboxCount", 0)),
          "eftTotal": round(eft_total, 2),
          "eftCount": int(source.get("eftCount", 0)),
          "combinedTotal": round(lockbox_total + eft_total, 2),
      })

    conn.close()
    return {
        "start": start_norm,
        "end": end_norm,
        "currentWorkDay": current_work_day,
        "rows": rows,
    }


# ------------------------------------------------------------
# GET FIRST PENDING IMPORTED FILE
# ------------------------------------------------------------
@app.get("/attachments/pending")
def get_first_pending(day: str | None = None):
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, filename, site, snapshot_path, review_status, batch_date, batch_id, processed_at
        FROM imported_files
        WHERE review_status = 'Pending'
        ORDER BY id ASC
    """)

    rows = cur.fetchall()
    conn.close()

    desired_day = _normalize_pending_day(day) if day else None
    row = None
    for candidate in rows:
        if desired_day and _row_pending_day(candidate, 5, 6, 7) != desired_day:
            continue
        row = candidate
        break

    if not row:
        return {"done": True}

    return {
        "id": row[0],
        "filename": row[1],
        "site": row[2],
        "snapshot": row[3],
        "status": row[4],
        "done": False
    }


@app.get("/attachments/{attachment_id}")
def get_attachment_by_id(attachment_id: int):
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, filename, site, snapshot_path, review_status, batch_date, batch_id, processed_at
        FROM imported_files
        WHERE id = ?
    """, (attachment_id,))

    row = cur.fetchone()
    conn.close()

    if not row:
        raise HTTPException(status_code=404, detail="Attachment not found")

    return {
        "id": row[0],
        "filename": row[1],
        "site": row[2],
        "snapshot": row[3],
        "status": row[4],
        "done": False,
    }


@app.get("/email-downloader/folders")
def get_email_downloader_folders():
    try:
        return _run_pythonw_worker("folders")
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to load folders: {exc}") from exc


@app.get("/email-downloader/dates")
def get_email_downloader_dates(folder_index: int):
    try:
        return _run_pythonw_worker("dates", "--folder-index", str(int(folder_index)))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to load dates: {exc}") from exc


@app.get("/email-downloader/last-uploaded")
def get_email_downloader_last_uploaded():
    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT date(MAX(processed_at))
            FROM imported_files
            WHERE source_type = 'email' AND processed_at IS NOT NULL AND processed_at != ''
            """
        )
        row = cur.fetchone()
        return row[0] if row and row[0] else None
    finally:
        conn.close()


@app.post("/email-downloader/run")
def run_email_downloader(payload: dict):
    try:
        folder_index = int(payload.get("folder_index"))
    except Exception as exc:
        raise HTTPException(status_code=400, detail="folder_index is required") from exc

    date_value = payload.get("date_value")
    move_messages_after = bool(payload.get("move_messages_after"))
    dest_folder_index = payload.get("dest_folder_index")

    try:
        args = ["run", "--folder-index", str(folder_index)]
        if date_value:
            args.extend(["--date-value", str(date_value)])
        if move_messages_after:
            args.append("--move-messages-after")
        if dest_folder_index not in (None, ""):
            args.extend(["--dest-folder-index", str(int(dest_folder_index))])
        result = _run_pythonw_worker(*args)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to run email downloader: {exc}") from exc

    return result


@app.post("/snapshot-generator/run")
def run_snapshot_generator():
    try:
        return _run_script_worker(SNAPSHOT_GENERATOR_SCRIPT)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to run snapshot generator: {exc}") from exc


@app.get("/pending/by-day")
def get_pending_by_day():
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, filename, batch_date, batch_id, processed_at
        FROM imported_files
        WHERE review_status = 'Pending'
        ORDER BY id ASC
    """)

    grouped = {}
    for row in cur.fetchall():
        day = _row_pending_day(row, 2, 3, 4)
        grouped.setdefault(day, []).append({
            "id": row[0],
            "filename": row[1],
        })

    conn.close()
    return grouped


@app.get("/approved")
def get_approved():
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, filename, site, detail, amount, processed_at
        FROM imported_files
        WHERE review_status = 'Approved'
        ORDER BY id DESC
    """)

    rows = cur.fetchall()
    conn.close()

    return [
        {
            "id": row[0],
            "filename": row[1],
            "site": row[2],
            "detail": row[3],
            "total": row[4] or 0,
            "date": row[5],
        }
        for row in rows
    ]


@app.get("/rejectlist")
def get_rejectlist():
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, filename, review_notes, processed_at
        FROM imported_files
        WHERE review_status = 'Rejected'
        ORDER BY id DESC
    """)

    rows = cur.fetchall()
    conn.close()

    return [
        {
            "id": row[0],
            "filename": row[1],
            "reason": row[2],
            "date": row[3],
        }
        for row in rows
    ]


@app.get("/site-review/history")
def get_site_review_history(view: str | None = None):
    normalized_view = (view or "complete").strip().lower()
    status_filter = {
        "approved": "Approved",
        "rejected": "Rejected",
    }.get(normalized_view)

    conn = get_conn()
    cur = conn.cursor()

    query = """
        SELECT
            f.id,
            f.filename,
            f.site,
            f.detail,
            f.review_notes,
            f.amount,
            f.review_status,
            f.processed_at,
            f.batch_id,
            f.batch_date,
            k.payload_json AS keyproof_payload_json
        FROM imported_files f
        LEFT JOIN keyproof k ON k.attachment_id = f.id
    """
    params = []
    if status_filter:
        query += " WHERE review_status = ?"
        params.append(status_filter)
    elif normalized_view == "complete":
        query += " WHERE batch_id IS NOT NULL AND TRIM(batch_id) != ''"

    query += " ORDER BY COALESCE(batch_date, processed_at, filename) DESC, id DESC"
    cur.execute(query, params)
    rows = cur.fetchall()
    conn.close()

    result = []
    for row in rows:
        amount = float(row[5] or 0)
        if amount == 0 and row[10]:
            try:
                payload = json.loads(row[10])
            except Exception:
                payload = None
            amount = _keyproof_total_from_payload(payload)

        result.append(
            {
                "id": row[0],
                "filename": row[1],
                "site": row[2],
                "detail": row[3],
                "reason": row[4],
                "total": amount,
                "status": row[6],
                "processedAt": row[7],
                "batchId": row[8],
                "batchDate": row[9],
            }
        )

    return result


@app.get("/calendar/status")
def get_calendar_status():
    return _calendar_status_payload()


@app.get("/calendar/work-day/lookup")
def get_calendar_work_day_lookup(work_day: str):
    normalized = normalize_mmddyyyy(work_day)
    if not normalized:
        raise HTTPException(status_code=400, detail="work_day is required")

    conn = get_conn()
    init_db()
    row = conn.execute(
        "SELECT bank_day FROM calendar WHERE paperwork_day = ?",
        (normalized,),
    ).fetchone()
    conn.close()

    return {
        "workDay": normalized,
        "bankDay": row[0] if row else None,
    }


@app.get("/calendar/range")
def get_calendar_range(start: str, end: str):
    return _calendar_range_payload(start, end)


@app.post("/calendar/setup")
def post_calendar_setup(payload: dict):
    start_date = normalize_mmddyyyy(payload.get("start_date"))
    if not start_date:
        raise HTTPException(status_code=400, detail="start_date is required")

    setup(start_date)
    return _calendar_status_payload()


@app.post("/calendar/add")
def post_calendar_add(payload: dict):
    try:
        days = int(payload.get("days"))
    except Exception:
        raise HTTPException(status_code=400, detail="days is required")

    if days <= 0:
        raise HTTPException(status_code=400, detail="days must be greater than zero")

    add_days(days)
    return _calendar_status_payload()


@app.post("/calendar/build-from")
def post_calendar_build_from(payload: dict):
    start_date = normalize_mmddyyyy(payload.get("start_date"))
    if not start_date:
        raise HTTPException(status_code=400, detail="start_date is required")

    try:
        days = int(payload.get("days"))
    except Exception:
        raise HTTPException(status_code=400, detail="days is required")

    if days <= 0:
        raise HTTPException(status_code=400, detail="days must be greater than zero")

    build_from(start_date, days)
    return _calendar_status_payload()


@app.delete("/calendar/days")
def delete_calendar_days(from_date: str, to_date: str):
    start_date = normalize_mmddyyyy(from_date)
    end_date = normalize_mmddyyyy(to_date)
    if not start_date or not end_date:
        raise HTTPException(status_code=400, detail="from_date and to_date are required")

    delete_days(start_date, end_date)
    return _calendar_status_payload()


@app.post("/calendar/work-day/set")
def post_calendar_set_work_day(payload: dict):
    work_day = normalize_mmddyyyy(payload.get("work_day"))
    if not work_day:
        raise HTTPException(status_code=400, detail="work_day is required")

    set_current_work_day(work_day)
    return _calendar_status_payload()


@app.post("/calendar/work-day/advance")
def post_calendar_advance_work_day():
    advance_current_work_day()
    return _calendar_status_payload()


@app.get("/tasks")
def get_tasks(task_list: str = "live"):
    init_db()
    conn = get_conn()
    ensure_tasks_table(conn)

    try:
        return _load_task_rows(conn, task_list=task_list)
    finally:
        conn.close()


@app.post("/tasks")
def post_task(task: dict):
    init_db()
    conn = get_conn()
    ensure_tasks_table(conn)

    try:
        normalized = _normalize_task_payload(task)
        columns = [name for name, _ in TASK_TABLE_COLUMNS]
        quoted_columns = ", ".join(_quote_identifier(name) for name in columns)
        placeholders = ", ".join(["?"] * len(columns))
        conn.execute(
            f'INSERT INTO {_quote_identifier("tasks")} ({quoted_columns}) VALUES ({placeholders})',
            tuple(normalized[column] for column in columns),
        )
        conn.commit()
        row = conn.execute(
            f'SELECT * FROM {_quote_identifier("tasks")} WHERE {_quote_identifier("task_id")} = ?',
            (normalized["task_id"],),
        ).fetchone()
        return _task_row_to_payload(row)
    finally:
        conn.close()


@app.post("/tasks/bulk-replace")
def replace_tasks(payload: dict):
    init_db()
    conn = get_conn()
    ensure_tasks_table(conn)

    task_list = str(payload.get("task_list") or "live").strip() or "live"
    tasks = payload.get("tasks", [])
    if not isinstance(tasks, list):
        raise HTTPException(status_code=400, detail="tasks must be a list")

    try:
        conn.execute(
            f'DELETE FROM {_quote_identifier("tasks")} WHERE {_quote_identifier("task_list")} = ?',
            (task_list,),
        )
        normalized_rows = []
        for index, task in enumerate(tasks):
            if not isinstance(task, dict):
                continue
            normalized = _normalize_task_payload(task, sort_order=index)
            normalized["task_list"] = task_list
            normalized_rows.append(normalized)

        if normalized_rows:
            columns = [name for name, _ in TASK_TABLE_COLUMNS]
            quoted_columns = ", ".join(_quote_identifier(name) for name in columns)
            placeholders = ", ".join(["?"] * len(columns))
            conn.executemany(
                f'INSERT INTO {_quote_identifier("tasks")} ({quoted_columns}) VALUES ({placeholders})',
                [tuple(row[column] for column in columns) for row in normalized_rows],
            )
        conn.commit()
        return {"status": "ok", "task_list": task_list, "rows": len(normalized_rows)}
    finally:
        conn.close()


@app.post("/tasks/import-template")
def import_template_to_live(payload: dict | None = None):
    init_db()
    conn = get_conn()
    ensure_tasks_table(conn)
    conn.row_factory = sqlite3.Row

    payload = payload or {}
    source_list = str(payload.get("source_list") or "template").strip() or "template"
    target_list = str(payload.get("target_list") or "live").strip() or "live"

    try:
        rows = conn.execute(
            f"""
            SELECT *
            FROM {_quote_identifier("tasks")}
            WHERE {_quote_identifier("task_list")} = ?
            ORDER BY {_quote_identifier("sort_order")} ASC, {_quote_identifier("created_at")} ASC, {_quote_identifier("task_id")} ASC
            """,
            (source_list,),
        ).fetchall()
        conn.execute(
            f'DELETE FROM {_quote_identifier("tasks")} WHERE {_quote_identifier("task_list")} = ?',
            (target_list,),
        )
        copied_rows = []
        now = datetime.now().isoformat(timespec="seconds")
        for index, row in enumerate(rows):
            copied = _task_row_to_payload(row)
            copied["id"] = f'task-{datetime.now().strftime("%Y%m%d%H%M%S%f")}-{index}'
            copied["task_list"] = target_list
            copied["sort_order"] = index
            copied["created_at"] = now
            copied["updated_at"] = now
            copied_rows.append(_normalize_task_payload(copied, task_id=copied["id"], sort_order=index))

        if copied_rows:
            columns = [name for name, _ in TASK_TABLE_COLUMNS]
            quoted_columns = ", ".join(_quote_identifier(name) for name in columns)
            placeholders = ", ".join(["?"] * len(columns))
            conn.executemany(
                f'INSERT INTO {_quote_identifier("tasks")} ({quoted_columns}) VALUES ({placeholders})',
                [tuple(row[column] for column in columns) for row in copied_rows],
            )
        conn.commit()
        return {"status": "ok", "source_list": source_list, "target_list": target_list, "rows": len(copied_rows)}
    finally:
        conn.close()


@app.put("/tasks/{task_id}")
def put_task(task_id: str, task: dict):
    init_db()
    conn = get_conn()
    ensure_tasks_table(conn)

    try:
        conn.row_factory = sqlite3.Row
        existing = conn.execute(
            f'SELECT * FROM {_quote_identifier("tasks")} WHERE {_quote_identifier("task_id")} = ?',
            (task_id,),
        ).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Task not found")

        normalized = _normalize_task_payload(task, task_id=task_id)
        conn.execute(
            f"""
            UPDATE {_quote_identifier("tasks")}
            SET
                {_quote_identifier("task_list")} = ?,
                {_quote_identifier("title")} = ?,
                {_quote_identifier("details")} = ?,
                {_quote_identifier("category")} = ?,
                {_quote_identifier("recurrence")} = ?,
                {_quote_identifier("action_type")} = ?,
                {_quote_identifier("action_label")} = ?,
                {_quote_identifier("action_value")} = ?,
                {_quote_identifier("done")} = ?,
                {_quote_identifier("sort_order")} = ?,
                {_quote_identifier("next_due_at")} = ?,
                {_quote_identifier("completed_at")} = ?,
                {_quote_identifier("created_at")} = ?,
                {_quote_identifier("updated_at")} = ?
            WHERE {_quote_identifier("task_id")} = ?
            """,
            (
                normalized["task_list"],
                normalized["title"],
                normalized["details"],
                normalized["category"],
                normalized["recurrence"],
                normalized["action_type"],
                normalized["action_label"],
                normalized["action_value"],
                normalized["done"],
                normalized["sort_order"],
                normalized["next_due_at"],
                normalized["completed_at"],
                normalized["created_at"],
                normalized["updated_at"],
                task_id,
            ),
        )
        conn.commit()
        row = conn.execute(
            f'SELECT * FROM {_quote_identifier("tasks")} WHERE {_quote_identifier("task_id")} = ?',
            (task_id,),
        ).fetchone()
        return _task_row_to_payload(row)
    finally:
        conn.close()


@app.delete("/tasks/{task_id}")
def delete_task(task_id: str):
    init_db()
    conn = get_conn()
    ensure_tasks_table(conn)

    try:
        cur = conn.cursor()
        cur.execute(
            f'DELETE FROM {_quote_identifier("tasks")} WHERE {_quote_identifier("task_id")} = ?',
            (task_id,),
        )
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="Task not found")
        conn.commit()
        return {"status": "ok", "task_id": task_id}
    finally:
        conn.close()


@app.get("/tasks/template")
def get_task_template():
    init_db()
    conn = get_conn()
    ensure_tasks_table(conn)

    try:
        return _load_task_rows(conn, task_list="template")
    finally:
        conn.close()


@app.get("/admin/tables")
def get_admin_tables():
    conn = get_conn()
    cur = conn.cursor()
    tables = []

    for table_name in _list_user_tables(conn):
        try:
            cur.execute(f"SELECT COUNT(*) FROM {_quote_identifier(table_name)}")
            row_count = int(cur.fetchone()[0] or 0)
        except Exception:
            row_count = 0

        try:
            cur.execute(f"PRAGMA table_info({_quote_identifier(table_name)})")
            columns = [
                {
                    "name": row[1],
                    "type": row[2],
                    "notNull": bool(row[3]),
                    "defaultValue": row[4],
                    "primaryKey": bool(row[5]),
                }
                for row in cur.fetchall()
            ]
        except Exception:
            columns = []

        tables.append({
            "name": table_name,
            "rowCount": row_count,
            "columnCount": len(columns),
            "columns": columns,
        })

    conn.close()
    return tables


@app.get("/admin/tables/{table_name}")
def get_admin_table_rows(table_name: str, limit: int = 250, offset: int = 0, sort_by: str = "rowid", sort_direction: str = "asc"):
    conn = get_conn()

    if not _table_exists(conn, table_name):
        conn.close()
        raise HTTPException(status_code=404, detail="Table not found")

    safe_limit = max(1, min(int(limit or 250), 250))
    safe_offset = max(0, int(offset or 0))
    cur = conn.cursor()

    cur.execute(f"PRAGMA table_info({_quote_identifier(table_name)})")
    columns = [
        {
            "name": row[1],
            "type": row[2],
            "notNull": bool(row[3]),
            "defaultValue": row[4],
            "primaryKey": bool(row[5]),
        }
        for row in cur.fetchall()
    ]

    column_lookup = {column["name"].lower(): column["name"] for column in columns}
    requested_sort = str(sort_by or "rowid").strip()
    requested_direction = str(sort_direction or "asc").strip().lower()
    if requested_direction not in ("asc", "desc"):
        raise HTTPException(status_code=400, detail="Invalid sort_direction")

    if requested_sort.lower() == "rowid":
        order_clause = f"rowid {requested_direction.upper()}"
    else:
        resolved_sort = column_lookup.get(requested_sort.lower())
        if not resolved_sort:
            raise HTTPException(status_code=400, detail="Invalid sort_by")
        order_clause = f"{_quote_identifier(resolved_sort)} {requested_direction.upper()}"

    cur.execute(
        f"SELECT rowid, * FROM {_quote_identifier(table_name)} ORDER BY {order_clause} LIMIT ? OFFSET ?",
        (safe_limit, safe_offset),
    )
    rows = cur.fetchall()
    headers = [description[0] for description in cur.description]

    payload_rows = [dict(zip(headers, row)) for row in rows]

    cur.execute(f"SELECT COUNT(*) FROM {_quote_identifier(table_name)}")
    total_rows = int(cur.fetchone()[0] or 0)

    conn.close()
    return {
        "name": table_name,
        "rowCount": total_rows,
        "columns": columns,
        "rows": payload_rows,
    }


# ------------------------------------------------------------
# SOURCE-DRIVEN 835 MATCH
# ------------------------------------------------------------
@app.get("/match/worklist")
def get_match_worklist(
    limit: int = 250,
    revision: str | None = None,
    page: int = 1,
    sort_by: str = "edi",
    sort_dir: str = "asc",
    show_matched: bool = True,
    show_unmatched: bool = True,
    latest_year_only: bool = False,
):
    return build_match_dashboard(
        limit=limit,
        revision=revision,
        page=page,
        sort_by=sort_by,
        sort_dir=sort_dir,
        show_matched=show_matched,
        show_unmatched=show_unmatched,
        latest_year_only=latest_year_only,
    )


@app.get("/match/matches")
def get_match_history_view(limit: int = 100):
    return build_match_history(limit=limit)


@app.get("/match/{edi_id}")
def get_match_detail_view(edi_id: int):
    detail = get_match_detail(edi_id)
    if detail is None:
        raise HTTPException(status_code=404, detail="EDI row not found")
    return detail


@app.post("/match/commit")
def post_match_commit(payload: dict):
    edi_id = payload.get("edi_id")
    if edi_id in (None, ""):
        raise HTTPException(status_code=400, detail="edi_id is required")

    try:
        result = commit_match(
            edi_id,
            eft_ids=payload.get("eft_ids") or [],
            lockbox_ids=payload.get("lockbox_ids") or [],
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    return result


@app.post("/match/commit-strong-hits")
def post_match_commit_strong_hits():
    return commit_all_strong_matches(match_source="MANUAL")


@app.post("/match/commit-exact-hits")
def post_match_commit_exact_hits():
    return commit_all_strong_matches(match_source="MANUAL")


# ------------------------------------------------------------
# 835 ZIP EXTRACTION
# ------------------------------------------------------------
def _route_835_zip_member(member_name: str) -> tuple[str, str] | tuple[None, None]:
    base_name = os.path.basename(member_name or "").strip()
    lower_name = base_name.lower()

    if lower_name.endswith(".trn"):
        return ZIP_835_TRN_FOLDER, base_name
    if lower_name.endswith((".html", ".htm")):
        return ZIP_835_HTML_FOLDER, base_name
    if lower_name.endswith(".era"):
        return ZIP_835_ERA_FOLDER, base_name

    return None, None


def _ensure_edi_pending_root():
    os.makedirs(EDI_PENDING_ROOT, exist_ok=True)
    return EDI_PENDING_ROOT


def _edi_manifest_dir(manifest_id: int):
    return os.path.join(_ensure_edi_pending_root(), f"manifest_{manifest_id}")


def _edi_manifest_member_dir(manifest_id: int, member_kind: str):
    return os.path.join(_edi_manifest_dir(manifest_id), member_kind.upper())


def _edi_manifest_processing_member_dir(manifest_id: int, member_kind: str):
    member_kind = str(member_kind or "").strip().upper()
    processing_root = ZIP_835_ERA_PROCESSING_FOLDER if member_kind == "ERA" else ZIP_835_HTML_PROCESSING_FOLDER
    return os.path.join(processing_root, f"manifest_{manifest_id}")


def _save_bytes_to_path(destination: str, content: bytes):
    os.makedirs(os.path.dirname(destination), exist_ok=True)
    with open(destination, "wb") as handle:
        handle.write(content)


def _sha256_bytes(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def _create_edi_batch_manifest(
    conn,
    *,
    upload_group_id: str,
    batch_id: str,
    zip_filename: str,
    zip_path: str,
    zip_hash: str,
    created_at: str,
):
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO EDI_BatchManifest (
            upload_group_id,
            batch_id,
            zip_filename,
            zip_path,
            zip_hash,
            status,
            trn_file_count,
            era_file_count,
            html_file_count,
            accepted_count,
            blocked_count,
            duplicate_count,
            notes,
            created_at,
            updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            upload_group_id,
            batch_id,
            zip_filename,
            zip_path,
            zip_hash,
            "UPLOADING",
            0,
            0,
            0,
            0,
            0,
            0,
            "",
            created_at,
            created_at,
        ),
    )
    manifest_id = int(cur.lastrowid)
    return manifest_id, batch_id


def _update_edi_batch_manifest(
    conn,
    manifest_id: int,
    *,
    status: str,
    trn_file_count: int,
    era_file_count: int,
    html_file_count: int,
    accepted_count: int,
    blocked_count: int,
    duplicate_count: int,
    notes: str,
    updated_at: str,
):
    conn.execute(
        """
        UPDATE EDI_BatchManifest
        SET status = ?,
            trn_file_count = ?,
            era_file_count = ?,
            html_file_count = ?,
            accepted_count = ?,
            blocked_count = ?,
            duplicate_count = ?,
            notes = ?,
            updated_at = ?
        WHERE id = ?
        """,
        (
            status,
            trn_file_count,
            era_file_count,
            html_file_count,
            accepted_count,
            blocked_count,
            duplicate_count,
            notes,
            updated_at,
            manifest_id,
        ),
    )


def _build_edi_manifest_item_rows(manifest_id: int, trn_file_rows, existing_edi_numbers, created_at: str):
    seen_numbers = set(str(number).strip() for number in existing_edi_numbers if str(number).strip())
    accepted_trn_file_rows = []
    manifest_items = []
    accepted_count = 0
    blocked_count = 0
    duplicate_count = 0
    next_row_index = 1

    for file_row in trn_file_rows:
        filename = str(file_row.get("filename") or "").strip()
        parsed_rows = list(file_row.get("rows") or [])
        if not filename:
            continue

        accepted_rows_for_file = []
        for file_row_index, (check_date, check_number, check_amount) in enumerate(parsed_rows, start=1):
            normalized_check = str(check_number or "").strip()
            is_duplicate = bool(normalized_check and normalized_check in seen_numbers)
            row_status = "BLOCKED_DUPLICATE" if is_duplicate else "ACCEPTED"
            blocked_reason = "Duplicate check number already exists in EDI" if is_duplicate else None

            manifest_items.append(
                {
                    "manifest_id": manifest_id,
                    "row_index": next_row_index,
                    "trn_filename": filename,
                    "trn_row_index": file_row_index,
                    "check_date": check_date,
                    "check_number": normalized_check,
                    "check_amount": check_amount,
                    "row_status": row_status,
                    "blocked_reason": blocked_reason,
                    "edi_id": None,
                    "ediload_id": None,
                    "edistage_id": None,
                    "accepted_at": created_at if not is_duplicate else None,
                    "created_at": created_at,
                    "updated_at": created_at,
                }
            )

            if is_duplicate:
                blocked_count += 1
                duplicate_count += 1
                next_row_index += 1
                continue

            seen_numbers.add(normalized_check)
            accepted_count += 1
            accepted_rows_for_file.append((check_date, check_number, check_amount))
            next_row_index += 1

        if accepted_rows_for_file:
            accepted_trn_file_rows.append({"filename": filename, "rows": accepted_rows_for_file})

    return accepted_trn_file_rows, pd.DataFrame(manifest_items), accepted_count, blocked_count, duplicate_count


def _normalize_sqlite_row(row, cursor=None):
    if row is None:
        return None
    if hasattr(row, "keys"):
        return dict(row)
    if cursor is not None and getattr(cursor, "description", None):
        return dict(zip([column[0] for column in cursor.description], row))
    return row


def _get_latest_edi_manifest(conn, statuses: tuple[str, ...] = ("PENDING_ACCEPTANCE", "PARTIAL", "TRANSFER_PENDING")):
    if not statuses:
        return None

    placeholders = ", ".join(["?"] * len(statuses))
    cur = conn.execute(
        f"""
        SELECT *
        FROM EDI_BatchManifest
        WHERE UPPER(TRIM(status)) IN ({placeholders})
        ORDER BY datetime(created_at) DESC, id DESC
        LIMIT 1
        """,
        tuple(statuses),
    )
    return _normalize_sqlite_row(cur.fetchone(), cur)


def _short_edi_batch_id(upload_group_id: str):
    token = str(upload_group_id or "").replace("-", "").strip()
    if not token:
        token = uuid.uuid4().hex
    return f"835-{token[:8].upper()}"


def _collect_edi_manifest_group(conn, manifest_row):
    upload_group_id = str(manifest_row.get("upload_group_id") or "").strip()
    if not upload_group_id:
        return [_normalize_sqlite_row(manifest_row)]

    cur = conn.execute(
        """
        SELECT *
        FROM EDI_BatchManifest
        WHERE upload_group_id = ?
        ORDER BY datetime(created_at) ASC, id ASC
        """,
        (upload_group_id,),
    )
    rows = [_normalize_sqlite_row(row, cur) for row in cur.fetchall()]
    return [row for row in rows if row]


def _manifest_member_signature(manifest_dir: str):
    stems = []
    if not manifest_dir or not os.path.isdir(manifest_dir):
        return ""

    for root_dir, _dirs, files in os.walk(manifest_dir):
        for filename in files:
            if filename.lower().endswith(".zip"):
                continue
            stem = Path(filename).stem.strip().lower()
            if stem:
                stems.append(stem)

    if not stems:
        return ""

    digest_input = "|".join(sorted(stems))
    return hashlib.sha256(digest_input.encode("utf-8", errors="ignore")).hexdigest()


def _collect_edi_manifest_family(conn, manifest_row):
    manifest_id = int(manifest_row["id"])
    manifest_dir = _edi_manifest_dir(manifest_id)
    signature = _manifest_member_signature(manifest_dir)
    if not signature:
        return [{"row": _normalize_sqlite_row(manifest_row), "dir": manifest_dir, "signature": ""}]

    family = []
    pending_root = Path(EDI_PENDING_ROOT)
    if not pending_root.is_dir():
        return [{"row": _normalize_sqlite_row(manifest_row), "dir": manifest_dir, "signature": signature}]

    for child in sorted(pending_root.iterdir(), key=lambda item: item.name):
        if not child.is_dir() or not child.name.startswith("manifest_"):
            continue
        try:
            sibling_id = int(child.name.split("_", 1)[1])
        except (IndexError, ValueError):
            continue

        sibling_signature = _manifest_member_signature(str(child))
        if sibling_signature != signature:
            continue

        sibling_row = conn.execute("SELECT * FROM EDI_BatchManifest WHERE id = ?", (sibling_id,)).fetchone()
        if sibling_row is None:
            continue
        family.append(
            {
                "row": _normalize_sqlite_row(sibling_row),
                "dir": str(child),
                "signature": sibling_signature,
            }
        )

    if not any(item["row"]["id"] == manifest_id for item in family):
        family.append({"row": _normalize_sqlite_row(manifest_row), "dir": manifest_dir, "signature": signature})

    family.sort(key=lambda item: int(item["row"]["id"]))
    return family


def _count_manifest_live_files(manifest_row, batch_id: str | None = None):
    manifest_id = int(manifest_row["id"])
    batch_id = str(batch_id or manifest_row["batch_id"] or f"835-{manifest_id}").strip()
    counts = {"trn": 0, "era": 0, "html": 0}

    for member_kind, target_root in (
        ("trn", ZIP_835_TRN_ARCHIVE_FOLDER),
        ("era", ZIP_835_ERA_FOLDER),
        ("html", ZIP_835_HTML_FOLDER),
    ):
        if not os.path.isdir(target_root):
            continue

        prefix = f"{batch_id}__"
        for filename in os.listdir(target_root):
            full_path = os.path.join(target_root, filename)
            if not os.path.isfile(full_path):
                continue
            if not filename.startswith(prefix):
                continue
            counts[member_kind] += 1

    expected = {
        "trn": int(manifest_row["trn_file_count"] or 0),
        "era": int(manifest_row["era_file_count"] or 0),
        "html": int(manifest_row["html_file_count"] or 0),
    }
    return expected, counts


def _cleanup_edi_processing_dirs(manifest_rows):
    for manifest_row in manifest_rows:
        manifest_id = int(manifest_row["id"])
        era_processing_dir = _edi_manifest_processing_member_dir(manifest_id, "ERA")
        html_processing_dir = _edi_manifest_processing_member_dir(manifest_id, "HTML")
        _safe_remove_tree(era_processing_dir, ZIP_835_ERA_FOLDER)
        _safe_remove_tree(html_processing_dir, ZIP_835_HTML_FOLDER)


def _cleanup_edi_processing_roots_if_idle(conn):
    active_count = conn.execute(
        """
        SELECT COUNT(*)
        FROM EDI_BatchManifest
        WHERE UPPER(TRIM(status)) IN ('UPLOADING', 'PENDING_ACCEPTANCE', 'PARTIAL', 'TRANSFER_PENDING')
        """
    ).fetchone()[0]
    if int(active_count or 0) > 0:
        return

    for processing_root in (ZIP_835_ERA_PROCESSING_FOLDER, ZIP_835_HTML_PROCESSING_FOLDER):
        if not processing_root or not os.path.isdir(processing_root):
            continue
        for child_name in os.listdir(processing_root):
            child_path = os.path.join(processing_root, child_name)
            if os.path.isdir(child_path):
                shutil.rmtree(child_path)
            elif os.path.isfile(child_path):
                os.remove(child_path)


def _safe_remove_tree(path: str, allowed_root: str):
    if not path:
        return

    resolved_path = Path(path).resolve()
    resolved_root = Path(allowed_root).resolve()
    try:
        resolved_path.relative_to(resolved_root)
    except ValueError as exc:
        raise HTTPException(status_code=500, detail="Refusing to remove a path outside the pending root") from exc

    if resolved_path.exists():
        shutil.rmtree(resolved_path)


def _copy_manifest_file(source_path: str, destination_path: str):
    if os.path.exists(destination_path):
        return False
    os.makedirs(os.path.dirname(destination_path), exist_ok=True)
    shutil.copy2(source_path, destination_path)
    return True


def _promote_edi_manifest_files(conn, manifest_rows, destination_batch_id: str | None = None):
    promoted = {"trn": 0, "era": 0, "html": 0}
    manifest_dirs = []

    for manifest_row in manifest_rows:
        manifest_id = int(manifest_row["id"])
        batch_id = str(destination_batch_id or manifest_row["batch_id"] or f"835-{manifest_id}").strip()
        manifest_dirs.append(_edi_manifest_dir(manifest_id))

        for member_kind, target_root, source_dir in (
            ("TRN", ZIP_835_TRN_ARCHIVE_FOLDER, _edi_manifest_member_dir(manifest_id, "TRN")),
            ("ERA", ZIP_835_ERA_FOLDER, _edi_manifest_processing_member_dir(manifest_id, "ERA")),
            ("HTML", ZIP_835_HTML_FOLDER, _edi_manifest_processing_member_dir(manifest_id, "HTML")),
        ):
            if not os.path.isdir(source_dir):
                continue

            for filename in sorted(os.listdir(source_dir)):
                source_path = os.path.join(source_dir, filename)
                if not os.path.isfile(source_path):
                    continue

                promoted_name = f"{batch_id}__{filename}"
                destination_path = os.path.join(target_root, promoted_name)
                if _copy_manifest_file(source_path, destination_path):
                    promoted[member_kind.lower()] += 1

    primary_manifest_dir = manifest_dirs[0] if manifest_dirs else ""
    return promoted, primary_manifest_dir


def _ensure_ediload_table(conn):
    cur = conn.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS EDILoad (
            id INTEGER PRIMARY KEY,
            check_date TEXT,
            check_number TEXT,
            check_amount REAL,
            filename TEXT,
            batchnum TEXT,
            transnum TEXT,
            timestamp TEXT,
            matchstatus TEXT
        )
        """
    )
    conn.commit()


def _load_835_trn_folder_into_ediload(conn):
    if not os.path.exists(ZIP_835_TRN_FOLDER):
        raise HTTPException(status_code=404, detail="TRN folder does not exist")

    os.makedirs(ZIP_835_TRN_ARCHIVE_FOLDER, exist_ok=True)

    _ensure_ediload_table(conn)
    cur = conn.cursor()
    conn.execute("BEGIN IMMEDIATE")
    cur.execute("DELETE FROM EDILoad")

    work_state = cur.execute("SELECT batchnum, transnum FROM work_state WHERE id = 1").fetchone()
    batchnum = str(work_state[0]).strip() if work_state and work_state[0] not in (None, "") else "1"
    try:
        next_trans = int(str(work_state[1]).strip() or "0") + 1 if work_state and work_state[1] not in (None, "") else 1
    except ValueError:
        next_trans = 1

    load_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    total_files = 0
    loaded_files = 0
    blocked_files = 0
    inserted_rows = 0
    blocked_rows = 0
    last_transnum = ""

    for filename in sorted(os.listdir(ZIP_835_TRN_FOLDER)):
        full_path = os.path.join(ZIP_835_TRN_FOLDER, filename)
        if not os.path.isfile(full_path):
            continue
        if filename.startswith("~$"):
            continue
        if not filename.lower().endswith((".trn", ".txt")):
            continue

        total_files += 1
        parsed_rows = _parse_835_trn_file(full_path)
        if not parsed_rows:
            blocked_files += 1
            continue

        new_rows = []
        duplicate_numbers = []
        for row in parsed_rows:
            check_date, check_number, check_amount = row
            exists = cur.execute(
                "SELECT 1 FROM EDILoad WHERE check_number = ?",
                (check_number,),
            ).fetchone()
            if exists:
                duplicate_numbers.append(check_number)
            else:
                new_rows.append(row)

        if not new_rows:
            blocked_files += 1
            blocked_rows += len(duplicate_numbers)
            continue

        file_frame = pd.DataFrame(
            [
                {
                    "check_date": check_date,
                    "check_number": check_number,
                    "check_amount": check_amount,
                    "filename": filename,
                    "batchnum": batchnum,
                    "transnum": str(next_trans + index),
                    "timestamp": load_timestamp,
                    "matchstatus": "UNMATCHED",
                }
                for index, (check_date, check_number, check_amount) in enumerate(new_rows)
            ]
        )

        _append_table_from_dataframe(
            conn,
            "EDILoad",
            file_frame[
                [
                    "check_date",
                    "check_number",
                    "check_amount",
                    "filename",
                    "batchnum",
                    "transnum",
                    "timestamp",
                    "matchstatus",
                ]
            ],
        )

        inserted_rows += int(len(file_frame))
        last_transnum = str(next_trans + len(file_frame) - 1)
        next_trans += len(file_frame)

        archive_path = os.path.join(ZIP_835_TRN_ARCHIVE_FOLDER, filename)
        shutil.move(full_path, archive_path)
        loaded_files += 1

    if inserted_rows > 0:
        cur.execute(
            """
            UPDATE work_state
            SET transnum = ?,
                timestamp = ?,
                matchstatus = ?
            WHERE id = 1
            """,
            (last_transnum, load_timestamp, "LOADED"),
        )

    conn.commit()

    status_tag = "EDILOAD LOADED" if blocked_files == 0 else "EDILOAD PARTIAL"
    status = "loaded" if blocked_files == 0 else "partial"
    if inserted_rows == 0:
        status_tag = "EDILOAD BLOCKED"
        status = "blocked"

    return {
        "status": status,
        "statusTag": status_tag,
        "message": (
            f"Loaded {loaded_files} TRN file(s) into EDILoad."
            if inserted_rows > 0
            else "No TRN rows qualified for EDILoad."
        ),
        "table": "EDILoad",
        "rowsLoaded": inserted_rows,
        "filesLoaded": loaded_files,
        "filesBlocked": blocked_files,
        "blockedRows": blocked_rows,
        "timestamp": load_timestamp,
        "movedTo": ZIP_835_TRN_ARCHIVE_FOLDER,
        "totalFiles": total_files,
    }


def _parse_835_trn_file(path: str):
    with open(path, "r", encoding="utf-8") as handle:
        return _parse_835_trn_text(handle.read())


def _parse_835_trn_text(text: str):
    rows = []
    lines = [line.strip() for line in StringIO(text) if line.strip()]

    if len(lines) < 3:
        return rows

    for line in lines[2:]:
        parts = line.split()
        if len(parts) < 3:
            continue

        check_date = normalize_mmddyyyy(parts[0])
        check_number = str(parts[1]).strip()
        try:
            check_amount = float(str(parts[2]).replace(",", "").strip())
        except Exception:
            continue

        if not check_date or not check_number:
            continue

        rows.append((check_date, check_number, check_amount))

    return rows


def _load_835_trn_rows_into_ediload(conn, trn_file_rows):
    ensure_source_table_columns(conn)
    _ensure_ediload_table(conn)
    cur = conn.cursor()
    conn.execute("BEGIN IMMEDIATE")
    cur.execute("DELETE FROM EDILoad")

    work_state = cur.execute("SELECT batchnum, transnum FROM work_state WHERE id = 1").fetchone()
    batchnum = str(work_state[0]).strip() if work_state and work_state[0] not in (None, "") else "1"
    try:
        next_trans = int(str(work_state[1]).strip() or "0") + 1 if work_state and work_state[1] not in (None, "") else 1
    except ValueError:
        next_trans = 1

    load_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    total_files = 0
    loaded_files = 0
    blocked_files = 0
    inserted_rows = 0
    blocked_rows = 0
    last_transnum = ""
    existing_edi_rows = cur.execute('SELECT "check_number" FROM EDI').fetchall()
    edi_numbers = {
        str(row[0]).strip()
        for row in existing_edi_rows
        if row and str(row[0]).strip()
    }
    seen_numbers = set(edi_numbers)

    for file_row in trn_file_rows:
        filename = str(file_row.get("filename") or "").strip()
        parsed_rows = list(file_row.get("rows") or [])
        if not filename:
            continue

        total_files += 1
        if not parsed_rows:
            blocked_files += 1
            continue

        new_rows = []
        duplicate_numbers = []
        for row in parsed_rows:
            check_date, check_number, check_amount = row
            if check_number in seen_numbers:
                duplicate_numbers.append(check_number)
            else:
                new_rows.append(row)
                seen_numbers.add(check_number)

        if not new_rows:
            blocked_files += 1
            blocked_rows += len(duplicate_numbers)
            continue

        file_frame = pd.DataFrame(
            [
                {
                    "check_date": check_date,
                    "check_number": check_number,
                    "check_amount": check_amount,
                    "filename": filename,
                    "batchnum": batchnum,
                    "transnum": str(next_trans + index),
                    "timestamp": load_timestamp,
                    "matchstatus": "UNMATCHED",
                }
                for index, (check_date, check_number, check_amount) in enumerate(new_rows)
            ]
        )

        _append_table_from_dataframe(
            conn,
            "EDILoad",
            file_frame[
                [
                    "check_date",
                    "check_number",
                    "check_amount",
                    "filename",
                    "batchnum",
                    "transnum",
                    "timestamp",
                    "matchstatus",
                ]
            ],
        )

        inserted_rows += int(len(file_frame))
        last_transnum = str(next_trans + len(file_frame) - 1)
        next_trans += len(file_frame)
        loaded_files += 1

    if inserted_rows > 0:
        cur.execute(
            """
            UPDATE work_state
            SET transnum = ?,
                timestamp = ?,
                matchstatus = ?
            WHERE id = 1
            """,
            (last_transnum, load_timestamp, "LOADED"),
        )

    conn.commit()

    status_tag = "EDILOAD LOADED" if blocked_files == 0 else "EDILOAD PARTIAL"
    status = "loaded" if blocked_files == 0 else "partial"
    if inserted_rows == 0:
        status_tag = "EDILOAD BLOCKED"
        status = "blocked"

    return {
        "status": status,
        "statusTag": status_tag,
        "message": (
            f"Loaded {loaded_files} TRN file(s) into EDILoad."
            if inserted_rows > 0
            else "No TRN rows qualified for EDILoad."
        ),
        "table": "EDILoad",
        "rowsLoaded": inserted_rows,
        "filesLoaded": loaded_files,
        "filesBlocked": blocked_files,
        "blockedRows": blocked_rows,
        "timestamp": load_timestamp,
        "movedTo": None,
        "totalFiles": total_files,
    }


@app.post("/835/upload-stage")
async def post_835_upload_stage(file: UploadFile = File(...), upload_group_id: str | None = Form(None)):
    filename = file.filename or ""
    if not filename.lower().endswith(".zip"):
        raise HTTPException(status_code=400, detail="Please choose a zip file")

    conn = None
    manifest_id = None
    batch_id = None
    upload_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    accepted_count = 0
    blocked_count = 0
    duplicate_count = 0
    manifest_status = "BLOCKED"
    upload_group_id = str(upload_group_id or "").strip() or uuid.uuid4().hex
    batch_id = _short_edi_batch_id(upload_group_id)

    try:
        file_bytes = await file.read()
        if not file_bytes:
            raise HTTPException(status_code=400, detail="Please choose a non-empty zip file")

        extracted_counts = {"trn": 0, "era": 0, "html": 0}
        trn_file_rows = []
        trn_load = None
        zip_hash = _sha256_bytes(file_bytes)

        with zipfile.ZipFile(BytesIO(file_bytes)) as archive:
            conn = get_conn()
            ensure_source_table_columns(conn)
            ensure_edi_manifest_tables(conn)

            existing_edi_rows = conn.execute('SELECT "check_number" FROM EDI').fetchall()
            edi_numbers = {
                str(row[0]).strip()
                for row in existing_edi_rows
                if row and str(row[0]).strip()
            }

            manifest_id, batch_id = _create_edi_batch_manifest(
                conn,
                upload_group_id=upload_group_id,
                batch_id=batch_id,
                zip_filename=filename,
                zip_path="",
                zip_hash=zip_hash,
                created_at=upload_timestamp,
            )
            manifest_dir = _edi_manifest_dir(manifest_id)
            pending_zip_path = os.path.join(manifest_dir, filename)
            _save_bytes_to_path(pending_zip_path, file_bytes)
            conn.execute(
                """
                UPDATE EDI_BatchManifest
                SET zip_path = ?, updated_at = ?
                WHERE id = ?
                """,
                (pending_zip_path, upload_timestamp, manifest_id),
            )
            conn.commit()

            for member in archive.infolist():
                if member.is_dir():
                    continue

                base_name = os.path.basename(member.filename or "").strip()
                if not base_name:
                    continue

                lower_name = base_name.lower()
                if lower_name.endswith((".trn", ".txt")):
                    with archive.open(member, "r") as source:
                        trn_bytes = source.read()
                    trn_destination = os.path.join(_edi_manifest_member_dir(manifest_id, "TRN"), base_name)
                    _save_bytes_to_path(trn_destination, trn_bytes)
                    trn_text = trn_bytes.decode("utf-8", errors="ignore")
                    parsed_rows = _parse_835_trn_text(trn_text)
                    if not parsed_rows:
                        raise HTTPException(status_code=400, detail=f"TRN file {base_name} did not contain any parseable rows.")

                    trn_file_rows.append({
                        "filename": base_name,
                        "rows": parsed_rows,
                    })
                    extracted_counts["trn"] += 1
                elif lower_name.endswith((".html", ".htm")):
                    with archive.open(member, "r") as source:
                        html_bytes = source.read()
                    html_destination = os.path.join(_edi_manifest_processing_member_dir(manifest_id, "HTML"), base_name)
                    _save_bytes_to_path(html_destination, html_bytes)
                    extracted_counts["html"] += 1
                elif lower_name.endswith(".era"):
                    with archive.open(member, "r") as source:
                        era_bytes = source.read()
                    era_destination = os.path.join(_edi_manifest_processing_member_dir(manifest_id, "ERA"), base_name)
                    _save_bytes_to_path(era_destination, era_bytes)
                    extracted_counts["era"] += 1

            accepted_trn_file_rows, manifest_items_df, accepted_count, blocked_count, duplicate_count = _build_edi_manifest_item_rows(
                manifest_id,
                trn_file_rows,
                edi_numbers,
                upload_timestamp,
            )

            if not manifest_items_df.empty:
                _append_table_from_dataframe(
                    conn,
                    "EDI_BatchManifestItem",
                    manifest_items_df[
                        [
                            "manifest_id",
                            "row_index",
                            "trn_filename",
                            "check_date",
                            "check_number",
                            "check_amount",
                            "row_status",
                            "blocked_reason",
                            "edi_id",
                            "ediload_id",
                            "edistage_id",
                            "accepted_at",
                            "created_at",
                            "updated_at",
                        ]
                    ],
                )
                conn.commit()

            if accepted_trn_file_rows:
                trn_conn = get_conn()
                try:
                    trn_load = _load_835_trn_rows_into_ediload(trn_conn, accepted_trn_file_rows)
                finally:
                    trn_conn.close()

            if accepted_count == 0:
                manifest_status = "BLOCKED"
            elif blocked_count > 0:
                manifest_status = "PARTIAL"
            else:
                manifest_status = "PENDING_ACCEPTANCE"

            manifest_notes = (
                "All TRN rows were blocked by duplicates in EDI."
                if accepted_count == 0
                else (
                    f"{blocked_count} TRN row(s) were blocked as duplicates in EDI."
                    if blocked_count > 0
                    else "TRN rows were loaded and ERA/HTML were held pending acceptance."
                )
            )

            _update_edi_batch_manifest(
                conn,
                manifest_id,
                status=manifest_status,
                trn_file_count=extracted_counts["trn"],
                era_file_count=extracted_counts["era"],
                html_file_count=extracted_counts["html"],
                accepted_count=accepted_count,
                blocked_count=blocked_count,
                duplicate_count=duplicate_count,
                notes=manifest_notes,
                updated_at=upload_timestamp,
            )
            conn.commit()

        rows_loaded = extracted_counts["trn"] + extracted_counts["era"] + extracted_counts["html"]
        return {
            "status": "success",
            "statusTag": manifest_status,
            "filename": filename,
            "rowsLoaded": rows_loaded,
            "blockedCount": blocked_count,
            "extractedCounts": extracted_counts,
            "trnLoad": trn_load,
            "manifestId": manifest_id,
            "batchId": batch_id,
            "uploadGroupId": upload_group_id,
            "pendingFolder": _edi_manifest_dir(manifest_id) if manifest_id is not None else None,
            "destinations": {
                "trn": _edi_manifest_member_dir(manifest_id, "TRN") if manifest_id is not None else None,
                "era": _edi_manifest_processing_member_dir(manifest_id, "ERA") if manifest_id is not None else None,
                "html": _edi_manifest_processing_member_dir(manifest_id, "HTML") if manifest_id is not None else None,
            },
        }
    except zipfile.BadZipFile:
        raise HTTPException(status_code=400, detail="Please choose a valid zip file")
    except HTTPException:
        raise
    except Exception as exc:
        if conn is not None and manifest_id is not None:
            try:
                _update_edi_batch_manifest(
                    conn,
                    manifest_id,
                    status="ERROR",
                    trn_file_count=0,
                    era_file_count=0,
                    html_file_count=0,
                    accepted_count=0,
                    blocked_count=0,
                    duplicate_count=0,
                    notes=str(exc),
                    updated_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                )
                conn.commit()
            except Exception:
                conn.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to process 835 zip file: {exc}")
    finally:
        if conn is not None:
            conn.close()


@app.post("/835/load-trn-folder")
def post_835_load_trn_folder():
    conn = get_conn()
    try:
        return _load_835_trn_folder_into_ediload(conn)
    except HTTPException:
        conn.rollback()
        raise
    except Exception as exc:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to load TRN files: {exc}")
    finally:
        conn.close()


@app.post("/835/stage-edi")
def post_835_stage_edi():
    conn = get_conn()
    try:
        ensure_source_table_columns(conn)
        cur = conn.cursor()
        conn.execute("BEGIN IMMEDIATE")

        load_df = pd.read_sql_query("SELECT * FROM EDILoad ORDER BY id ASC", conn)
        if load_df.empty:
            raise HTTPException(status_code=400, detail="EDILoad is empty. Load TRN files first.")

        work_state = cur.execute(
            "SELECT batchnum, transnum FROM work_state WHERE id = 1"
        ).fetchone()
        batchnum = str(work_state[0]).strip() if work_state and work_state[0] not in (None, "") else "1"
        try:
            next_trans = int(str(work_state[1]).strip() or "0") + 1 if work_state and work_state[1] not in (None, "") else 1
        except ValueError:
            next_trans = 1

        stage_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        staged_df = load_df.drop(columns=["id"], errors="ignore").copy()
        staged_df["batchnum"] = batchnum
        staged_df["transnum"] = [str(next_trans + index) for index in range(len(staged_df))]
        staged_df["timestamp"] = stage_timestamp
        if "matchstatus" in staged_df.columns:
            staged_df["matchstatus"] = "STAGED"

        _replace_table_from_dataframe(conn, "EDIStage", staged_df)
        conn.commit()

        end_trans = next_trans + len(staged_df) - 1
        cur.execute(
            """
            UPDATE work_state
            SET transnum = ?,
                timestamp = ?,
                matchstatus = ?
            WHERE id = 1
            """,
            (str(end_trans), stage_timestamp, "STAGED"),
        )
        conn.commit()

        return {
            "status": "staged",
            "statusTag": "EDI STAGED",
            "message": f"Copied {len(staged_df)} row(s) from EDILoad to EDIStage.",
            "table": "EDIStage",
            "rowsStaged": int(len(staged_df)),
            "filesStaged": int(staged_df["filename"].nunique()) if "filename" in staged_df.columns else 0,
            "batchnum": batchnum,
            "startTransnum": str(next_trans),
            "endTransnum": str(end_trans),
            "timestamp": stage_timestamp,
        }
    except HTTPException:
        conn.rollback()
        raise
    except Exception as exc:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to stage EDI data: {exc}")
    finally:
        conn.close()


@app.post("/835/vet-edi")
def post_835_vet_edi():
    conn = get_conn()
    try:
        ensure_source_table_columns(conn)
        cur = conn.cursor()
        conn.execute("BEGIN IMMEDIATE")

        stage_df = pd.read_sql_query("SELECT * FROM EDIStage ORDER BY id ASC", conn)
        if stage_df.empty:
            raise HTTPException(status_code=400, detail="EDIStage is empty. Run staging first.")

        edi_df = pd.read_sql_query('SELECT "check_number" FROM EDI', conn)
        edi_numbers = {
            str(value).strip()
            for value in edi_df.get("check_number", pd.Series(dtype=str)).fillna("").astype(str).tolist()
            if str(value).strip()
        }

        working_df = stage_df.drop(columns=["id"], errors="ignore").copy()
        working_df["check_number"] = working_df.get("check_number", pd.Series(dtype=str)).fillna("").astype(str).str.strip()
        working_df["check_date"] = working_df.get("check_date", pd.Series(dtype=str)).fillna("").astype(str).str.strip()
        working_df["check_amount"] = working_df.get("check_amount", pd.Series(dtype=str)).fillna("").astype(str).str.strip()

        duplicate_mask = [str(value).strip() in edi_numbers for value in working_df["check_number"].tolist()]
        load_df = working_df.loc[[not value for value in duplicate_mask]].copy()
        duplicate_df = working_df.loc[duplicate_mask].copy()

        load_df["matchstatus"] = "VETTED"
        duplicate_rows = [
            {
                "row": int(index) + 1,
                "checkNumber": str(row.get("check_number", "")).strip(),
                "date": str(row.get("check_date", "")).strip(),
                "amount": str(row.get("check_amount", "")).strip(),
                "status": "BLOCKED",
            }
            for index, row in duplicate_df.iterrows()
        ]

        _replace_table_from_dataframe(
            conn,
            "EDIVett",
            load_df[
                [
                    "check_date",
                    "check_number",
                    "check_amount",
                    "filename",
                    "batchnum",
                    "transnum",
                    "timestamp",
                    "matchstatus",
                ]
            ].rename(
                columns={
                    "check_date": "check_date",
                    "check_number": "check_number",
                    "check_amount": "check_amount",
                }
            ),
        )
        conn.commit()

        total_rows = int(len(working_df))
        duplicate_count = int(len(duplicate_df))
        loaded_count = int(len(load_df))
        all_duplicates = total_rows > 0 and loaded_count == 0
        status_tag = "VETTED" if duplicate_count == 0 else "PARTIAL VETTED"
        status = "vetted" if duplicate_count == 0 else "partial"
        if all_duplicates:
            status_tag = "BLOCKED"
            status = "blocked"

        last_transnum = ""
        if loaded_count > 0 and "transnum" in load_df.columns and not load_df["transnum"].empty:
            last_transnum = str(load_df["transnum"].iloc[-1]).strip()

        vet_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cur.execute(
            """
            UPDATE work_state
            SET transnum = COALESCE(NULLIF(?, ''), transnum),
                timestamp = ?,
                matchstatus = ?
            WHERE id = 1
            """,
            (last_transnum, vet_timestamp, status_tag),
        )
        conn.commit()

        message = (
            f"{loaded_count} row(s) loaded into EDIVett."
            if loaded_count > 0
            else "No rows qualified for EDIVett."
        )
        if duplicate_count > 0:
            message = f"{duplicate_count} duplicate row(s) were blocked."
            if loaded_count > 0:
                message += f" {loaded_count} clean row(s) loaded into EDIVett."

        return {
            "status": status,
            "statusTag": status_tag,
            "message": message,
            "table": "EDIVett",
            "rowsLoaded": loaded_count,
            "totalRows": total_rows,
            "filesLoaded": int(working_df["filename"].nunique()) if "filename" in working_df.columns else 0,
            "duplicateCount": duplicate_count,
            "allDuplicates": all_duplicates,
            "duplicateRows": duplicate_rows,
            "timestamp": vet_timestamp,
        }
    except HTTPException:
        conn.rollback()
        raise
    except Exception as exc:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to vet EDI data: {exc}")
    finally:
        conn.close()


@app.post("/835/approval-stage")
async def post_835_approval_stage(request: Request):
    payload = await request.json() if request.headers.get("content-type", "").lower().startswith("application/json") else {}
    decision = str(payload.get("decision", "")).strip().lower()
    if decision not in ("approve", "deny"):
        raise HTTPException(status_code=400, detail="decision must be approve or deny")

    conn = get_conn()
    refresh_result = None
    refresh_warning = ""
    try:
        ensure_source_table_columns(conn)
        ensure_edi_manifest_tables(conn)
        cur = conn.cursor()
        approval_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        manifest_row = _get_latest_edi_manifest(conn)
        manifest_id = int(manifest_row["id"]) if manifest_row else None
        manifest_batch_id = str(manifest_row["batch_id"] or "") if manifest_row else ""
        manifest_group_rows = _collect_edi_manifest_group(conn, manifest_row) if manifest_row else []

        if decision == "deny":
            if manifest_group_rows:
                for group_row in manifest_group_rows:
                    _update_edi_batch_manifest(
                        conn,
                        int(group_row["id"]),
                        status="REJECTED",
                        trn_file_count=int(group_row["trn_file_count"] or 0),
                        era_file_count=int(group_row["era_file_count"] or 0),
                        html_file_count=int(group_row["html_file_count"] or 0),
                        accepted_count=int(group_row["accepted_count"] or 0),
                        blocked_count=int(group_row["blocked_count"] or 0),
                        duplicate_count=int(group_row["duplicate_count"] or 0),
                        notes="Approval denied and pending files were removed.",
                        updated_at=approval_timestamp,
                    )
                conn.commit()
                for group_row in manifest_group_rows:
                    _safe_remove_tree(_edi_manifest_dir(int(group_row["id"])), EDI_PENDING_ROOT)
                _cleanup_edi_processing_dirs(manifest_group_rows)

            conn.execute("BEGIN IMMEDIATE")
            cur.execute("DELETE FROM EDILoad")
            cur.execute("DELETE FROM EDIStage")
            cur.execute("DELETE FROM EDIVett")
            cur.execute(
                """
                UPDATE work_state
                SET timestamp = NULL,
                    matchstatus = NULL
                WHERE id = 1
                """
            )
            conn.commit()

            return {
                "status": "denied",
                "statusTag": "DENIED",
                "message": "835 approval was denied and the working tables were reset.",
                "tablesReset": ["EDILoad", "EDIStage", "EDIVett"],
                "timestamp": approval_timestamp,
            }

        vetted_df = pd.read_sql_query("SELECT * FROM EDIVett ORDER BY id ASC", conn)
        if vetted_df.empty:
            raise HTTPException(status_code=400, detail="EDIVett is empty. Run vetting before approval.")

        approved_df = vetted_df.drop(columns=["id"], errors="ignore").copy()
        approved_df["matchstatus"] = "UNMATCHED"

        approval_key_columns = ["check_date", "check_number", "check_amount", "filename", "batchnum", "transnum", "timestamp"]
        if _all_rows_already_exist(conn, "EDI", approved_df, approval_key_columns):
            raise HTTPException(
                status_code=409,
                detail="These 835 rows were already approved and are already present in EDI.",
            )

        manifest_group_rows = _collect_edi_manifest_group(conn, manifest_row) if manifest_row else []
        manifest_accepted_count = int(manifest_row["accepted_count"] or 0) if manifest_row else 0
        promoted_manifest = None
        promoted_manifest_files = {"trn": 0, "era": 0, "html": 0}
        if manifest_row and manifest_accepted_count <= 0:
            raise HTTPException(
                status_code=409,
                detail={
                    "message": "No accepted TRN rows are available for this manifest, so EDI was not updated.",
                    "manifestId": manifest_id,
                    "batchId": manifest_batch_id,
                    "acceptedCount": manifest_accepted_count,
                    "blockedCount": int(manifest_row["blocked_count"] or 0),
                    "duplicateCount": int(manifest_row["duplicate_count"] or 0),
                },
            )

        if manifest_row and manifest_accepted_count > 0:
            promoted_manifest_files, manifest_dir = _promote_edi_manifest_files(
                conn,
                manifest_group_rows,
                destination_batch_id=manifest_batch_id,
            )
            expected_counts = {
                "trn": sum(int(row["trn_file_count"] or 0) for row in manifest_group_rows),
                "era": sum(int(row["era_file_count"] or 0) for row in manifest_group_rows),
                "html": sum(int(row["html_file_count"] or 0) for row in manifest_group_rows),
            }
            group_manifest_row = manifest_group_rows[0] if manifest_group_rows else manifest_row
            live_counts = _count_manifest_live_files(group_manifest_row, batch_id=manifest_batch_id)[1]
            transfer_ready = all(live_counts[key] >= expected_counts[key] for key in ("trn", "era", "html"))
            if not transfer_ready:
                transfer_pending_notes = (
                    "Waiting for ERA/HTML transfer to complete before EDI approval. "
                    f"Expected TRN/ERA/HTML counts {expected_counts}, found {live_counts}."
                )
                _update_edi_batch_manifest(
                    conn,
                    manifest_id,
                    status="TRANSFER_PENDING",
                    trn_file_count=int(manifest_row["trn_file_count"] or 0),
                    era_file_count=int(manifest_row["era_file_count"] or 0),
                    html_file_count=int(manifest_row["html_file_count"] or 0),
                    accepted_count=int(manifest_row["accepted_count"] or 0),
                    blocked_count=int(manifest_row["blocked_count"] or 0),
                    duplicate_count=int(manifest_row["duplicate_count"] or 0),
                    notes=transfer_pending_notes,
                    updated_at=approval_timestamp,
                )
                conn.commit()
                raise HTTPException(
                    status_code=409,
                    detail={
                        "message": "ERA/HTML transfer verification failed. EDI was not updated.",
                        "expectedCounts": expected_counts,
                        "foundCounts": live_counts,
                        "manifestId": manifest_id,
                        "batchId": manifest_batch_id,
                    },
                )
            promoted_manifest = {
                "manifestId": manifest_id,
                "batchId": manifest_batch_id,
                "manifestDir": manifest_dir,
            }

        conn.execute("BEGIN IMMEDIATE")
        approved_rows = _append_table_from_dataframe(
            conn,
            "EDI",
            approved_df[
                [
                    "check_date",
                    "check_number",
                    "check_amount",
                    "filename",
                    "batchnum",
                    "transnum",
                    "timestamp",
                    "matchstatus",
                ]
            ],
        )

        last_transnum = ""
        if "transnum" in approved_df.columns and not approved_df["transnum"].empty:
            last_transnum = str(approved_df["transnum"].iloc[-1]).strip()

        cur.execute("DELETE FROM EDILoad")
        cur.execute("DELETE FROM EDIStage")
        cur.execute("DELETE FROM EDIVett")
        cur.execute(
            """
            UPDATE work_state
            SET transnum = COALESCE(NULLIF(?, ''), transnum),
                timestamp = ?,
                matchstatus = ?
            WHERE id = 1
            """,
            (last_transnum, approval_timestamp, "APPROVED"),
        )

        if manifest_group_rows:
            manifest_status = "APPROVED_PARTIAL" if int(manifest_row["blocked_count"] or 0) > 0 else "APPROVED"
            for group_row in manifest_group_rows:
                cur.execute(
                    """
                    UPDATE EDI_BatchManifest
                    SET status = ?,
                        notes = ?,
                        updated_at = ?
                    WHERE id = ?
                    """,
                    (
                        manifest_status,
                        "Approved and promoted to live folders." if promoted_manifest else "Approved.",
                        approval_timestamp,
                        int(group_row["id"]),
                    ),
                )
                cur.execute(
                    """
                    UPDATE EDI_BatchManifestItem
                    SET row_status = 'APPROVED',
                        updated_at = ?
                    WHERE manifest_id = ? AND UPPER(TRIM(row_status)) = 'ACCEPTED'
                    """,
                    (approval_timestamp, int(group_row["id"])),
                )
        conn.commit()

        try:
            refresh_result = commit_all_strong_matches()
        except Exception as exc:
            refresh_warning = str(exc)

        if manifest_group_rows:
            _cleanup_edi_processing_dirs(manifest_group_rows)
            for group_row in manifest_group_rows:
                _safe_remove_tree(_edi_manifest_dir(int(group_row["id"])), EDI_PENDING_ROOT)
            _cleanup_edi_processing_roots_if_idle(conn)

        response = {
            "status": "approved",
            "statusTag": "APPROVED",
            "message": f"835 approval completed and {approved_rows} vetted row(s) were appended to EDI.",
            "rowsApproved": approved_rows,
            "table": "EDI",
            "timestamp": approval_timestamp,
            "tablesReset": ["EDILoad", "EDIStage", "EDIVett"],
            "manifestId": manifest_id,
            "manifestBatchId": manifest_batch_id,
            "manifestGroupSize": len(manifest_group_rows) if manifest_group_rows else 0,
            "promotedFiles": promoted_manifest_files,
        }
        if refresh_result is not None:
            response["matchRefresh"] = refresh_result
        if refresh_warning:
            response["matchRefreshWarning"] = refresh_warning
        return response
    except HTTPException:
        conn.rollback()
        raise
    except Exception as exc:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to approve 835 data: {exc}")
    finally:
        conn.close()


# ------------------------------------------------------------
# EFT IMPORT STAGING
# ------------------------------------------------------------
@app.post("/eft/upload-stage")
async def post_eft_upload_stage(file: UploadFile = File(...)):
    filename = file.filename or ""
    lower_name = filename.lower()
    if not filename.startswith("DEP_1101_TRAN"):
        raise HTTPException(status_code=400, detail="Please choose a DEP_1101_TRAN file from Downloads")
    if not lower_name.endswith((".xls", ".xlsx")):
        raise HTTPException(status_code=400, detail="Please upload a DEP_1101_TRAN .xls or .xlsx file")

    try:
        file_bytes = await file.read()
        df = pd.read_excel(BytesIO(file_bytes), dtype=str).fillna("")
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {exc}")

    required_headers = [
        "As-Of Date",
        "Debit Amt",
        "Credit Amt",
        "Descriptive Text 1",
    ]

    optional_headers = [
        "As-Of-Time",
        "Bank ID",
        "Bank Name",
        "State",
        "Acct No",
        "Acct Type",
        "Acct Name",
        "Currency",
        "IBAN",
        "BAI Type Code",
        "Tran Desc",
        "0 Day Flt Amt",
        "1 Day Flt Amt",
        "2+ Day Flt Amt",
        "Customer Ref No",
        "Value Date",
        "Location",
        "Bank Reference",
        "Tran Status",
        "Descriptive Text 2",
        "Descriptive Text 3",
        "Descriptive Text 4",
        "Descriptive Text 5",
        "Descriptive Text 6",
        "Descriptive Text 7",
        "Descriptive Text 8",
        "Descriptive Text 9",
        "Descriptive Text 10",
        "Descriptive Text 11",
        "Descriptive Text 12",
        "Descriptive Text 13",
        "Descriptive Text 14",
        "Descriptive Text 15",
        "Descriptive Text 16",
        "Descriptive Text 17",
        "Descriptive Text 18",
        "Descriptive Text 19",
        "Descriptive Text 20",
        "Descriptive Text 21",
        "Descriptive Text 22",
        "Description",
        "Unique ID",
        "Discretionary Data",
        "UETR",
        "Payment Fee Deduction",
        "Beneficiary Final Wire Received Status",
    ]

    missing_required_headers = [header for header in required_headers if header not in df.columns]
    if missing_required_headers:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required column(s) in DEP_1101_TRAN file: {', '.join(missing_required_headers)}",
        )

    conn = get_conn()
    try:
        ensure_eftload_schema(conn)
        cur = conn.cursor()
        conn.execute("BEGIN IMMEDIATE")
        cur.execute("DELETE FROM EFTLoad")

        working_df = df.copy()
        for header in required_headers + optional_headers:
            if header not in working_df.columns:
                working_df[header] = ""
        working_df = working_df[required_headers + optional_headers].copy()
        working_df.insert(0, "batchnum", None)
        working_df.insert(1, "transnum", None)
        working_df.insert(2, "timestamp", None)
        working_df.insert(3, "matchstatus", None)

        eftload_columns = [
            "batchnum",
            "transnum",
            "timestamp",
            "matchstatus",
            *required_headers,
            *optional_headers,
        ]
        quoted_columns = ", ".join(_quote_identifier(column) for column in eftload_columns)
        placeholders = ", ".join(["?"] * len(eftload_columns))
        insert_sql = f"INSERT INTO {_quote_identifier('EFTLoad')} ({quoted_columns}) VALUES ({placeholders})"
        cur.executemany(insert_sql, working_df.itertuples(index=False, name=None))
        conn.commit()

        row_count = int(conn.execute("SELECT COUNT(*) FROM EFTLoad").fetchone()[0] or 0)
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return {
        "status": "loaded",
        "statusTag": "EFTLOAD REFRESHED",
        "filename": filename,
        "rowsLoaded": row_count,
        "table": "EFTLoad",
        "appendMode": False,
    }


# ------------------------------------------------------------
# EFT STAGING
# ------------------------------------------------------------
@app.post("/eft/transform-stage")
def post_eft_transform_stage():
    conn = get_conn()
    try:
        ensure_source_table_columns(conn)
        ensure_eftload_schema(conn)
        cur = conn.cursor()

        load_count_row = cur.execute("SELECT COUNT(*) FROM EFTLoad").fetchone()
        load_count = int(load_count_row[0] or 0) if load_count_row else 0
        if load_count == 0:
            raise HTTPException(status_code=400, detail="EFTLoad is empty. Load DEP_1101_TRAN first.")

        work_state = cur.execute(
            "SELECT batchnum, transnum FROM work_state WHERE id = 1"
        ).fetchone()
        batchnum = str(work_state[0]).strip() if work_state and work_state[0] not in (None, "") else "1"
        try:
            next_trans = int(str(work_state[1]).strip() or "0") + 1 if work_state and work_state[1] not in (None, "") else 1
        except ValueError:
            next_trans = 1

        source_df = pd.read_sql_query("SELECT * FROM EFTLoad ORDER BY rowid ASC", conn)
        if source_df.empty:
            raise HTTPException(status_code=400, detail="EFTLoad is empty. Load DEP_1101_TRAN first.")

        stage_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        staged_rows = []
        for index, row in source_df.iterrows():
            payer, check_number = _parse_eft_descriptive_text_1(row.get("Descriptive Text 1", ""))
            staged_rows.append(
                {
                    "Date": _normalize_yyyy_mm_dd_to_mmddyyyy(row.get("As-Of Date", "")),
                    "Amount": str(row.get("Credit Amt", "")).strip(),
                    "CheckNumber": check_number,
                    "Payer": payer,
                    "batchnum": batchnum,
                    "transnum": str(next_trans + index),
                    "timestamp": stage_timestamp,
                    "matchstatus": "STAGED",
                    "Descriptive Text 1": str(row.get("Descriptive Text 1", "")).strip(),
                }
            )

        staged_df = pd.DataFrame(staged_rows)

        _replace_table_from_dataframe(conn, "EFTStage", staged_df)
        conn.commit()

        end_trans = next_trans + len(staged_df) - 1
        cur.execute(
            """
            UPDATE work_state
            SET transnum = ?, timestamp = ?, matchstatus = ?
            WHERE id = 1
            """,
            (str(end_trans), stage_timestamp, "STAGED"),
        )
        conn.commit()

        return {
            "status": "staged",
            "statusTag": "EFT STAGED",
            "rowsStaged": int(len(staged_df)),
            "batchnum": batchnum,
            "startTransnum": str(next_trans),
            "endTransnum": str(end_trans),
            "timestamp": stage_timestamp,
            "table": "EFTStage",
            "fieldMap": {
                "Date": "As-Of Date",
                "Amount": "Credit Amt",
                "Descriptive Text 1": "Descriptive Text 1",
            },
        }
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


# ------------------------------------------------------------
# EFT VETTING
# ------------------------------------------------------------
@app.post("/eft/vet-stage")
def post_eft_vet_stage():
    conn = get_conn()
    try:
        ensure_source_table_columns(conn)
        ensure_eftload_schema(conn)
        cur = conn.cursor()

        stage_df = pd.read_sql_query("SELECT * FROM EFTStage ORDER BY rowid ASC", conn)
        if stage_df.empty:
            raise HTTPException(status_code=400, detail="EFTStage is empty. Run staging first.")

        eft_df = pd.read_sql_query('SELECT "Date" FROM EFT', conn)
        eft_dates = {
            _normalize_yyyy_mm_dd_to_mmddyyyy(value)
            for value in eft_df.get("Date", pd.Series(dtype=str)).fillna("").astype(str).tolist()
            if _normalize_yyyy_mm_dd_to_mmddyyyy(value)
        }

        working_df = stage_df.drop(columns=["id"], errors="ignore").copy()
        working_df["Date"] = working_df.get("Date", pd.Series(dtype=str)).fillna("").astype(str).map(_normalize_yyyy_mm_dd_to_mmddyyyy)
        working_df["Payer"] = working_df.get("Payer", pd.Series(dtype=str)).fillna("").astype(str).str.strip()
        working_df["CheckNumber"] = working_df.get("CheckNumber", pd.Series(dtype=str)).fillna("").astype(str).str.strip()
        working_df["Amount"] = working_df.get("Amount", pd.Series(dtype=str)).fillna("").astype(str).str.strip()
        working_df["Descriptive Text 1"] = working_df.get("Descriptive Text 1", pd.Series(dtype=str)).fillna("").astype(str)

        duplicate_mask = []
        blank_payer_mask = []
        blank_date_mask = []
        for date_value, payer_value in zip(working_df["Date"].tolist(), working_df["Payer"].tolist()):
            normalized_date = str(date_value).strip()
            payer = str(payer_value).strip()
            blank_date = normalized_date == ""
            blank_payer = payer == ""
            duplicate = False
            if not blank_date and normalized_date in eft_dates:
                duplicate = True

            duplicate_mask.append(duplicate)
            blank_payer_mask.append(blank_payer)
            blank_date_mask.append(blank_date)

        duplicate_df = working_df.loc[
            [duplicate_mask[index] and not blank_payer_mask[index] for index in range(len(working_df))]
        ].copy()
        blank_payer_df = working_df.loc[blank_payer_mask].copy()
        blank_date_df = working_df.loc[blank_date_mask].copy()

        load_mask = [
            not duplicate_mask[index] and not blank_payer_mask[index] and not blank_date_mask[index]
            for index in range(len(working_df))
        ]
        load_df = working_df.loc[load_mask].copy()
        load_df["matchstatus"] = "VETTED"

        duplicate_rows = []
        for index, row in duplicate_df.iterrows():
            duplicate_rows.append(
                {
                    "row": int(index) + 1,
                    "date": str(row.get("Date", "")).strip(),
                    "payer": str(row.get("Payer", "")).strip(),
                    "checkNumber": str(row.get("CheckNumber", "")).strip(),
                    "amount": str(row.get("Amount", "")).strip(),
                    "status": "BLOCKED",
                }
            )

        blank_payer_rows = []
        for index, row in blank_payer_df.iterrows():
            blank_payer_rows.append(
                {
                    "row": int(index) + 1,
                    "date": str(row.get("Date", "")).strip(),
                    "payer": str(row.get("Payer", "")).strip(),
                    "checkNumber": str(row.get("CheckNumber", "")).strip(),
                    "amount": str(row.get("Amount", "")).strip(),
                    "status": "BLOCKED",
                }
            )

        blank_date_rows = []
        for index, row in blank_date_df.iterrows():
            blank_date_rows.append(
                {
                    "row": int(index) + 1,
                    "date": str(row.get("Date", "")).strip(),
                    "payer": str(row.get("Payer", "")).strip(),
                    "checkNumber": str(row.get("CheckNumber", "")).strip(),
                    "amount": str(row.get("Amount", "")).strip(),
                    "status": "BLOCKED",
                }
            )

        blocked_row_indices = {
            index
            for index in range(len(working_df))
            if duplicate_mask[index] or blank_payer_mask[index] or blank_date_mask[index]
        }
        blocked_count = int(len(blocked_row_indices))
        qualified_count = int(len(load_df))
        total_rows = int(len(working_df))
        all_blocked = total_rows > 0 and qualified_count == 0
        status_tag = "BLOCKED" if blocked_count > 0 else "VETTED"
        status_message = (
            f"{int(len(duplicate_df))} duplicate date row(s) blocked, "
            f"{int(len(blank_payer_df))} blank payer row(s) blocked, "
            f"{int(len(blank_date_df))} blank date row(s) blocked."
        )
        status_message += f" {blocked_count} unique row(s) blocked."
        if qualified_count > 0:
            status_message += f" {qualified_count} row(s) loaded into EFTVett."
        else:
            status_message += " No rows qualified for EFTVett."

        vetted_columns = [
            "Date",
            "Amount",
            "Payer",
            "CheckNumber",
            "batchnum",
            "transnum",
            "timestamp",
            "matchstatus",
        ]
        vetted_df = load_df[vetted_columns].copy() if qualified_count > 0 else load_df.iloc[0:0][vetted_columns].copy()
        _replace_table_from_dataframe(conn, "EFTVett", vetted_df)
        conn.commit()

        last_transnum = ""
        if qualified_count > 0 and "transnum" in load_df.columns and not load_df["transnum"].empty:
            last_transnum = str(load_df["transnum"].iloc[-1]).strip()

        vet_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cur.execute(
            """
            UPDATE work_state
            SET transnum = COALESCE(NULLIF(?, ''), transnum),
                timestamp = ?,
                matchstatus = ?
            WHERE id = 1
            """,
            (last_transnum, vet_timestamp, status_tag),
        )
        conn.commit()

        return {
            "status": "blocked" if blocked_count > 0 else "vetted",
            "statusTag": f"EFT {status_tag}",
            "message": status_message,
            "table": "EFTVett",
            "rowsLoaded": qualified_count,
            "totalRows": total_rows,
            "duplicateCount": int(len(duplicate_df)),
            "blankPayerCount": int(len(blank_payer_df)),
            "blankDateCount": int(len(blank_date_df)),
            "blockedCount": blocked_count,
            "allBlocked": all_blocked,
            "duplicateRows": duplicate_rows,
            "blankPayerRows": blank_payer_rows,
            "blankDateRows": blank_date_rows,
        }
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


# ------------------------------------------------------------
# EFT APPROVAL
# ------------------------------------------------------------
@app.post("/eft/approval-stage")
async def post_eft_approval_stage(request: Request):
    payload = await request.json() if request.headers.get("content-type", "").lower().startswith("application/json") else {}
    decision = str(payload.get("decision", "")).strip().lower()
    if decision not in ("approve", "approve_partial", "deny"):
        raise HTTPException(status_code=400, detail="decision must be approve, approve_partial, or deny")

    conn = get_conn()
    try:
        ensure_eft_tables(conn)
        cur = conn.cursor()
        conn.execute("BEGIN IMMEDIATE")
        approval_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if decision == "deny":
            cur.execute("DELETE FROM EFTLoad")
            cur.execute("DELETE FROM EFTStage")
            cur.execute("DELETE FROM EFTVett")
            cur.execute(
                """
                UPDATE work_state
                SET timestamp = NULL,
                    matchstatus = NULL
                WHERE id = 1
                """
            )
            conn.commit()

            return {
                "status": "denied",
                "statusTag": "DENIED",
                "message": "EFT approval was denied and the working tables were reset.",
                "tablesReset": ["EFTLoad", "EFTStage", "EFTVett"],
            }

        vetted_df = pd.read_sql_query("SELECT * FROM EFTVett ORDER BY rowid ASC", conn)
        if vetted_df.empty:
            raise HTTPException(status_code=400, detail="EFTVett is empty. Run vetting before approval.")

        approved_df = vetted_df[
            [
                "Date",
                "Amount",
                "Payer",
                "CheckNumber",
                "batchnum",
                "transnum",
                "timestamp",
            ]
        ].copy()
        approved_df["matchstatus"] = "UNMATCHED"

        approval_key_columns = ["Date", "Amount", "Payer", "CheckNumber", "batchnum", "transnum", "timestamp"]
        if _all_rows_already_exist(conn, "EFT", approved_df, approval_key_columns):
            raise HTTPException(
                status_code=409,
                detail="These EFT rows were already approved and are already present in EFT.",
            )

        approved_rows = _append_table_from_dataframe(
            conn,
            "EFT",
            approved_df[
                [
                    "Date",
                    "Amount",
                    "Payer",
                    "CheckNumber",
                    "batchnum",
                    "transnum",
                    "timestamp",
                    "matchstatus",
                ]
            ],
        )

        last_transnum = ""
        if "transnum" in approved_df.columns and not approved_df["transnum"].empty:
            last_transnum = str(approved_df["transnum"].iloc[-1]).strip()

        approval_status = "PARTIAL APPROVED" if decision == "approve_partial" else "APPROVED"
        cur.execute("DELETE FROM EFTLoad")
        cur.execute("DELETE FROM EFTStage")
        cur.execute("DELETE FROM EFTVett")
        cur.execute(
            """
            UPDATE work_state
            SET transnum = COALESCE(NULLIF(?, ''), transnum),
                timestamp = ?,
                matchstatus = ?
            WHERE id = 1
            """,
            (last_transnum, approval_timestamp, approval_status),
        )
        conn.commit()

        return {
            "status": "approved",
            "statusTag": approval_status,
            "message": f"EFT approval completed and {approved_rows} vetted row(s) were appended to EFT.",
            "rowsApproved": approved_rows,
            "table": "EFT",
            "timestamp": approval_timestamp,
            "tablesReset": ["EFTLoad", "EFTStage", "EFTVett"],
        }
    finally:
        conn.close()


# ------------------------------------------------------------
# LOCKBOX IMPORT STAGING
# ------------------------------------------------------------
@app.post("/lockbox/upload-stage")
async def post_lockbox_upload_stage(file: UploadFile = File(...)):
    filename = file.filename or ""
    lower_name = filename.lower()
    if not filename.startswith("SearchResults"):
        raise HTTPException(status_code=400, detail="Please choose a SearchResults file from Downloads")
    if not lower_name.endswith((".xls", ".xlsx")):
        raise HTTPException(status_code=400, detail="Please upload a SearchResults .xls or .xlsx file")

    try:
        file_bytes = await file.read()
        df = pd.read_excel(BytesIO(file_bytes), dtype=str).fillna("")
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {exc}")

    conn = get_conn()
    try:
        ensure_source_table_columns(conn)
        cur = conn.cursor()
        cur.execute("DELETE FROM LockboxLoad")
        conn.commit()

        df.to_sql("LockboxLoad", conn, if_exists="append", index=False)
        conn.commit()

        row_count = int(conn.execute("SELECT COUNT(*) FROM LockboxLoad").fetchone()[0] or 0)
    finally:
        conn.close()

    return {
        "status": "loaded",
        "filename": filename,
        "rowsLoaded": row_count,
        "table": "LockboxLoad",
    }


@app.post("/lockbox/transform-stage")
def post_lockbox_transform_stage():
    conn = get_conn()
    try:
        ensure_source_table_columns(conn)
        cur = conn.cursor()

        load_count_row = cur.execute("SELECT COUNT(*) FROM LockboxLoad").fetchone()
        load_count = int(load_count_row[0] or 0) if load_count_row else 0
        if load_count == 0:
            raise HTTPException(status_code=400, detail="LockboxLoad is empty. Upload SearchResults.xls first.")

        work_state = cur.execute(
            "SELECT batchnum, transnum FROM work_state WHERE id = 1"
        ).fetchone()
        batchnum = str(work_state[0]).strip() if work_state and work_state[0] not in (None, "") else "1"
        try:
            next_trans = int(str(work_state[1]).strip() or "0") + 1 if work_state and work_state[1] not in (None, "") else 1
        except ValueError:
            next_trans = 1

        source_df = pd.read_sql_query("SELECT * FROM LockboxLoad ORDER BY id ASC", conn)
        if source_df.empty:
            raise HTTPException(status_code=400, detail="LockboxLoad is empty. Upload SearchResults.xls first.")

        transform_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        source_df = source_df.drop(columns=["id"], errors="ignore")
        source_df["batchnum"] = batchnum
        source_df["transnum"] = [str(next_trans + index) for index in range(len(source_df))]
        source_df["timestamp"] = transform_timestamp
        source_df["matchstatus"] = "TRANSFORMED"

        cur.execute("DELETE FROM LockboxStage")
        conn.commit()
        source_df.to_sql("LockboxStage", conn, if_exists="append", index=False)

        end_trans = next_trans + len(source_df) - 1
        cur.execute(
            """
            UPDATE work_state
            SET transnum = ?, timestamp = ?, matchstatus = ?
            WHERE id = 1
            """,
            (str(end_trans), transform_timestamp, "TRANSFORMED"),
        )
        conn.commit()

        return {
            "status": "transformed",
            "statusTag": "TRANSFORMED",
            "rowsStaged": int(len(source_df)),
            "batchnum": batchnum,
            "startTransnum": str(next_trans),
            "endTransnum": str(end_trans),
            "timestamp": transform_timestamp,
            "table": "LockboxStage",
        }
    finally:
        conn.close()


@app.post("/lockbox/vet-stage")
async def post_lockbox_vet_stage(request: Request):
    payload = await request.json() if request.headers.get("content-type", "").lower().startswith("application/json") else {}
    decision = str(payload.get("decision", "")).strip().lower()
    if decision not in ("", "partial", "reject"):
        raise HTTPException(status_code=400, detail="decision must be partial or reject")

    conn = get_conn()
    try:
        ensure_source_table_columns(conn)
        cur = conn.cursor()

        stage_df = pd.read_sql_query("SELECT * FROM LockboxStage ORDER BY id ASC", conn)
        if stage_df.empty:
          raise HTTPException(status_code=400, detail="LockboxStage is empty. Transform the lockbox file first.")

        lockbox_df = pd.read_sql_query('SELECT "Check Number" FROM Lockbox', conn)
        lockbox_checks = {
            str(value).strip()
            for value in lockbox_df.get("Check Number", pd.Series(dtype=str)).fillna("").astype(str).tolist()
            if str(value).strip()
        }

        working_df = stage_df.drop(columns=["id"], errors="ignore").copy()
        working_df["Check Number"] = working_df["Check Number"].fillna("").astype(str)

        seen_stage_checks = set()
        duplicate_mask = []
        for check_number in working_df["Check Number"].tolist():
            normalized_check = str(check_number).strip()
            is_duplicate = False
            if normalized_check:
                if normalized_check in lockbox_checks or normalized_check in seen_stage_checks:
                    is_duplicate = True
                else:
                    seen_stage_checks.add(normalized_check)
            duplicate_mask.append(is_duplicate)

        duplicate_df = working_df.loc[duplicate_mask].copy()
        qualifying_df = working_df.loc[[not value for value in duplicate_mask]].copy()

        duplicate_rows = []
        for index, row in duplicate_df.iterrows():
            duplicate_rows.append(
                {
                    "row": int(index) + 1,
                    "checkNumber": str(row.get("Check Number", "")).strip(),
                    "transactionNumber": str(row.get("Transaction Number", "")).strip(),
                    "depositDate": str(row.get("Deposit Date", "")).strip(),
                    "payor": str(row.get("Payor", "")).strip(),
                    "checkAmount": str(row.get("Check Amount", "")).strip(),
                    "status": str(row.get("Status", "")).strip(),
                }
            )

        total_rows = int(len(working_df))
        duplicate_count = int(len(duplicate_df))
        qualified_count = int(len(qualifying_df))
        all_duplicates = total_rows > 0 and duplicate_count == total_rows
        partial_available = duplicate_count > 0 and qualified_count > 0

        if decision == "reject":
            cur.execute("DELETE FROM LockboxVett")
            conn.commit()
            return {
                "status": "rejected",
                "statusTag": "REJECTED",
                "message": "Lockbox vetting was rejected and no rows were loaded into LockboxVett.",
                "table": "LockboxVett",
                "totalRows": total_rows,
                "duplicateCount": duplicate_count,
                "qualifiedCount": qualified_count,
                "allDuplicates": all_duplicates,
                "decisionRequired": False,
                "duplicateRows": duplicate_rows,
            }

        if all_duplicates:
            cur.execute("DELETE FROM LockboxVett")
            conn.commit()
            return {
                "status": "rejected",
                "statusTag": "REJECTED",
                "message": "All rows were duplicates, so the file was rejected.",
                "table": "LockboxVett",
                "totalRows": total_rows,
                "duplicateCount": duplicate_count,
                "qualifiedCount": qualified_count,
                "allDuplicates": True,
                "decisionRequired": False,
                "duplicateRows": duplicate_rows,
            }

        if decision not in ("partial",) and partial_available:
            return {
                "status": "review",
                "statusTag": "PARTIAL REVIEW",
                "message": "Some rows are duplicates. Choose partial upload to keep the clean rows, or reject the file for a complete overhaul.",
                "table": "LockboxVett",
                "totalRows": total_rows,
                "duplicateCount": duplicate_count,
                "qualifiedCount": qualified_count,
                "allDuplicates": False,
                "decisionRequired": True,
                "duplicateRows": duplicate_rows,
            }

        if decision == "partial" or (duplicate_count == 0 and qualified_count > 0):
            load_df = qualifying_df.copy() if duplicate_count > 0 else working_df.copy()
            load_df["matchstatus"] = "VETTED"
            cur.execute("DELETE FROM LockboxVett")
            conn.commit()
            load_df.to_sql("LockboxVett", conn, if_exists="append", index=False)
            conn.commit()
            return {
                "status": "vetted",
                "statusTag": "VETTED",
                "message": "Lockbox vetting completed successfully.",
                "table": "LockboxVett",
                "rowsLoaded": int(len(load_df)),
                "totalRows": total_rows,
                "duplicateCount": duplicate_count,
                "qualifiedCount": int(len(load_df)),
                "allDuplicates": False,
                "decisionRequired": False,
                "duplicateRows": duplicate_rows,
            }

        return {
            "status": "review",
            "statusTag": "PARTIAL REVIEW",
            "message": "Review the duplicate rows below before deciding whether to accept the clean rows.",
            "table": "LockboxVett",
            "totalRows": total_rows,
            "duplicateCount": duplicate_count,
            "qualifiedCount": qualified_count,
            "allDuplicates": False,
            "decisionRequired": partial_available,
            "duplicateRows": duplicate_rows,
        }
    finally:
        conn.close()


@app.post("/lockbox/approval-stage")
async def post_lockbox_approval_stage(request: Request):
    payload = await request.json() if request.headers.get("content-type", "").lower().startswith("application/json") else {}
    decision = str(payload.get("decision", "")).strip().lower()
    if decision not in ("approve", "deny"):
        raise HTTPException(status_code=400, detail="decision must be approve or deny")

    conn = get_conn()
    try:
        ensure_source_table_columns(conn)
        cur = conn.cursor()
        conn.execute("BEGIN IMMEDIATE")
        approval_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if decision == "deny":
            cur.execute("DELETE FROM LockboxLoad")
            cur.execute("DELETE FROM LockboxStage")
            cur.execute("DELETE FROM LockboxVett")
            cur.execute(
                """
                UPDATE work_state
                SET timestamp = NULL,
                    matchstatus = NULL
                WHERE id = 1
                """
            )
            conn.commit()

            return {
                "status": "denied",
                "statusTag": "DENIED",
                "message": "Lockbox approval was denied and the working tables were reset.",
                "tablesReset": ["LockboxLoad", "LockboxStage", "LockboxVett"],
            }

        vetted_df = pd.read_sql_query("SELECT * FROM LockboxVett ORDER BY id ASC", conn)
        if vetted_df.empty:
            raise HTTPException(status_code=400, detail="LockboxVett is empty. Run vetting before approval.")

        lockbox_df = vetted_df.drop(columns=["id"], errors="ignore").copy()
        lockbox_df["matchstatus"] = "APPROVED"

        lockbox_key_columns = [column for column in lockbox_df.columns if column != "matchstatus"]
        if _all_rows_already_exist_as_text(conn, "Lockbox", lockbox_df, lockbox_key_columns):
            raise HTTPException(
                status_code=409,
                detail="These Lockbox rows were already approved and are already present in Lockbox.",
            )

        lockbox_df.to_sql("Lockbox", conn, if_exists="append", index=False)

        approved_rows = int(len(lockbox_df))
        last_transnum = ""
        if "transnum" in lockbox_df.columns and not lockbox_df["transnum"].empty:
            last_transnum = str(lockbox_df["transnum"].iloc[-1]).strip()

        cur.execute("DELETE FROM LockboxLoad")
        cur.execute("DELETE FROM LockboxStage")
        cur.execute("DELETE FROM LockboxVett")
        cur.execute(
            """
            UPDATE work_state
            SET transnum = COALESCE(NULLIF(?, ''), transnum),
                timestamp = ?,
                matchstatus = ?
            WHERE id = 1
            """,
            (last_transnum, approval_timestamp, "APPROVED"),
        )
        conn.commit()

        return {
            "status": "approved",
            "statusTag": "APPROVED",
            "message": "Lockbox approval completed and the vetted rows were written to Lockbox.",
            "rowsApproved": approved_rows,
            "table": "Lockbox",
            "timestamp": approval_timestamp,
        }
    finally:
        conn.close()


# ------------------------------------------------------------
# BANKING SPREADSHEET
# ------------------------------------------------------------
@app.get("/banking/spreadsheet")
def get_banking_spreadsheet():
    return build_banking_spreadsheet()


def _era_check_candidates(check_number: str) -> list[str]:
    raw = str(check_number or "").strip()
    normalized = normalize_checknum(raw)
    candidates = [raw, normalized, raw.lstrip("0"), normalized.lstrip("0")]
    return [candidate for candidate in dict.fromkeys(candidates) if candidate]


@app.get("/era/spreadsheet")
def get_era_spreadsheet(work_day: str):
    normalized_work_day = normalize_mmddyyyy(work_day)
    if not normalized_work_day:
        raise HTTPException(status_code=400, detail="work_day is required")

    conn = get_conn()
    init_db()
    row = conn.execute(
        "SELECT bank_day FROM calendar WHERE paperwork_day = ?",
        (normalized_work_day,),
    ).fetchone()
    conn.close()

    bank_day = normalize_mmddyyyy(row[0]) if row and row[0] else None
    if not bank_day:
        return {
            "workDay": normalized_work_day,
            "bankDay": None,
            "rows": [],
            "matchedChecks": 0,
            "matchedFiles": 0,
        }

    spreadsheet = build_banking_spreadsheet()
    if not os.path.exists(ZIP_835_ERA_FOLDER):
        raise HTTPException(status_code=404, detail="ERA folder does not exist")

    era_files = [
        filename
        for filename in sorted(os.listdir(ZIP_835_ERA_FOLDER))
        if filename.lower().endswith(".era")
        and os.path.isfile(os.path.join(ZIP_835_ERA_FOLDER, filename))
    ]

    file_contents: dict[str, str] = {}
    for filename in era_files:
        full_path = os.path.join(ZIP_835_ERA_FOLDER, filename)
        try:
            with open(full_path, "r", errors="ignore") as handle:
                file_contents[filename] = handle.read()
        except Exception:
            file_contents[filename] = ""

    rows = []
    matched_checks = set()

    for group in spreadsheet["groups"]:
        for row_data in group["rows"]:
            if row_data.get("edi") != "Y":
                continue
            if normalize_mmddyyyy(row_data.get("date")) != bank_day:
                continue

            check_number = str(row_data.get("checkNumber") or "").strip()
            if not check_number:
                continue

            candidates = _era_check_candidates(check_number)
            matched_files = [
                filename
                for filename, content in file_contents.items()
                if any(candidate in content for candidate in candidates)
            ]

            if not matched_files:
                continue

            matched_checks.add(check_number)
            for filename in matched_files:
                rows.append(
                    {
                        "source": group["source"],
                        "bankDay": bank_day,
                        "checkNumber": check_number,
                        "payer": str(row_data.get("payer") or "").strip(),
                        "amount": str(row_data.get("amount") or "").strip(),
                        "eraFile": filename,
                    }
                )

    rows.sort(key=lambda item: (item["source"], item["checkNumber"], item["eraFile"]))

    return {
        "workDay": normalized_work_day,
        "bankDay": bank_day,
        "rows": rows,
        "matchedChecks": len(matched_checks),
        "matchedFiles": len(rows),
    }


@app.post("/era/convert")
def post_era_convert(payload: dict):
    work_day = normalize_mmddyyyy(payload.get("work_day"))
    if not work_day:
        raise HTTPException(status_code=400, detail="work_day is required")

    conn = get_conn()
    init_db()
    row = conn.execute(
        "SELECT bank_day FROM calendar WHERE paperwork_day = ?",
        (work_day,),
    ).fetchone()
    conn.close()

    bank_day = normalize_mmddyyyy(row[0]) if row and row[0] else None
    if not bank_day:
        raise HTTPException(status_code=400, detail="No bank day is mapped to the selected posting day")

    if not os.path.exists(ZIP_835_ERA_FOLDER):
        raise HTTPException(status_code=404, detail="ERA folder does not exist")

    os.makedirs(os.path.join(ZIP_835_ERA_FOLDER, "Renamed"), exist_ok=True)

    spreadsheet = build_banking_spreadsheet()
    target_rows = []
    for group in spreadsheet["groups"]:
        for row_data in group["rows"]:
            if row_data.get("edi") != "Y":
                continue
            if normalize_mmddyyyy(row_data.get("date")) != bank_day:
                continue

            check_number = str(row_data.get("checkNumber") or "").strip()
            if not check_number:
                continue

            target_rows.append(
                {
                    "source": group["source"],
                    "checkNumber": check_number,
                    "payer": str(row_data.get("payer") or "").strip(),
                    "amount": str(row_data.get("amount") or "").strip(),
                }
            )

    target_rows.sort(key=lambda item: (item["source"], item["checkNumber"]))
    check_candidates = [row["checkNumber"] for row in target_rows]

    era_files = [
        filename
        for filename in sorted(os.listdir(ZIP_835_ERA_FOLDER))
        if filename.lower().endswith(".era")
        and os.path.isfile(os.path.join(ZIP_835_ERA_FOLDER, filename))
    ]

    renamed = []
    sequence = 1
    date_prefix = datetime.strptime(work_day, "%m/%d/%Y").strftime("%m.%d.%y")
    renamed_folder = os.path.join(ZIP_835_ERA_FOLDER, "Renamed")

    for filename in era_files:
        full_path = os.path.join(ZIP_835_ERA_FOLDER, filename)
        orig_ext = os.path.splitext(filename)[1]
        try:
            with open(full_path, "r", errors="ignore") as handle:
                content = handle.read()
        except Exception:
            continue

        matched_check = ""
        for check_number in check_candidates:
            for candidate in _era_check_candidates(check_number):
                if candidate and candidate in content:
                    matched_check = check_number
                    break
            if matched_check:
                break

        if not matched_check:
            continue

        new_name = f"{date_prefix}-835-{sequence}---{matched_check}{orig_ext}"
        destination = os.path.join(renamed_folder, new_name)

        if os.path.exists(destination):
            raise HTTPException(status_code=409, detail=f"Destination already exists: {new_name}")

        shutil.move(full_path, destination)
        renamed.append(
            {
                "sourceFile": filename,
                "renamedFile": new_name,
                "checkNumber": matched_check,
            }
        )
        sequence += 1

    return {
        "status": "converted" if renamed else "noop",
        "statusTag": "CONVERTED" if renamed else "NO FILES",
        "message": (
            f"Renamed {len(renamed)} ERA file(s) and moved them to 2.ERA/Renamed."
            if renamed
            else "No ERA files matched the selected day."
        ),
        "workDay": work_day,
        "bankDay": bank_day,
        "renamedCount": len(renamed),
        "outputFolder": renamed_folder,
        "renamedFiles": renamed,
    }


@app.get("/html/spreadsheet")
def get_html_spreadsheet(work_day: str):
    normalized_work_day = normalize_mmddyyyy(work_day)
    if not normalized_work_day:
        raise HTTPException(status_code=400, detail="work_day is required")

    conn = get_conn()
    init_db()
    row = conn.execute(
        "SELECT bank_day FROM calendar WHERE paperwork_day = ?",
        (normalized_work_day,),
    ).fetchone()
    conn.close()

    bank_day = normalize_mmddyyyy(row[0]) if row and row[0] else None
    if not bank_day:
        return {
            "workDay": normalized_work_day,
            "bankDay": None,
            "rows": [],
            "matchedChecks": 0,
            "matchedFiles": 0,
        }

    spreadsheet = build_banking_spreadsheet()
    if not os.path.exists(ZIP_835_HTML_FOLDER):
        raise HTTPException(status_code=404, detail="HTML folder does not exist")

    html_files = [
        filename
        for filename in sorted(os.listdir(ZIP_835_HTML_FOLDER))
        if filename.lower().endswith((".html", ".htm"))
        and os.path.isfile(os.path.join(ZIP_835_HTML_FOLDER, filename))
    ]

    file_contents: dict[str, str] = {}
    for filename in html_files:
        full_path = os.path.join(ZIP_835_HTML_FOLDER, filename)
        try:
            with open(full_path, "r", errors="ignore") as handle:
                file_contents[filename] = handle.read()
        except Exception:
            file_contents[filename] = ""

    rows = []
    matched_checks = set()

    for group in spreadsheet["groups"]:
        for row_data in group["rows"]:
            if row_data.get("edi") != "Y":
                continue
            if normalize_mmddyyyy(row_data.get("date")) != bank_day:
                continue

            check_number = str(row_data.get("checkNumber") or "").strip()
            if not check_number:
                continue

            candidates = _era_check_candidates(check_number)
            matched_files = [
                filename
                for filename, content in file_contents.items()
                if any(candidate in content for candidate in candidates)
            ]

            if not matched_files:
                continue

            matched_checks.add(check_number)
            for filename in matched_files:
                rows.append(
                    {
                        "source": group["source"],
                        "bankDay": bank_day,
                        "checkNumber": check_number,
                        "htmlFile": filename,
                    }
                )

    rows.sort(key=lambda item: (item["source"], item["checkNumber"], item["htmlFile"]))

    return {
        "workDay": normalized_work_day,
        "bankDay": bank_day,
        "rows": rows,
        "matchedChecks": len(matched_checks),
        "matchedFiles": len(rows),
    }


@app.post("/html/convert")
def post_html_convert(payload: dict):
    work_day = normalize_mmddyyyy(payload.get("work_day"))
    if not work_day:
        raise HTTPException(status_code=400, detail="work_day is required")

    conn = get_conn()
    init_db()
    row = conn.execute(
        "SELECT bank_day FROM calendar WHERE paperwork_day = ?",
        (work_day,),
    ).fetchone()
    conn.close()

    bank_day = normalize_mmddyyyy(row[0]) if row and row[0] else None
    if not bank_day:
        raise HTTPException(status_code=400, detail="No bank day is mapped to the selected posting day")

    if not os.path.exists(ZIP_835_HTML_FOLDER):
        raise HTTPException(status_code=404, detail="HTML folder does not exist")

    os.makedirs(os.path.join(ZIP_835_HTML_FOLDER, "Renamed"), exist_ok=True)

    spreadsheet = build_banking_spreadsheet()
    target_rows = []
    for group in spreadsheet["groups"]:
        for row_data in group["rows"]:
            if row_data.get("edi") != "Y":
                continue
            if normalize_mmddyyyy(row_data.get("date")) != bank_day:
                continue

            check_number = str(row_data.get("checkNumber") or "").strip()
            if not check_number:
                continue

            target_rows.append(
                {
                    "source": group["source"],
                    "checkNumber": check_number,
                    "payer": str(row_data.get("payer") or "").strip(),
                    "amount": str(row_data.get("amount") or "").strip(),
                }
            )

    target_rows.sort(key=lambda item: (item["source"], item["checkNumber"]))
    check_candidates = [row["checkNumber"] for row in target_rows]

    html_files = [
        filename
        for filename in sorted(os.listdir(ZIP_835_HTML_FOLDER))
        if filename.lower().endswith((".html", ".htm"))
        and os.path.isfile(os.path.join(ZIP_835_HTML_FOLDER, filename))
    ]

    renamed = []
    sequence = 1
    date_prefix = datetime.strptime(work_day, "%m/%d/%Y").strftime("%m.%d.%y")
    renamed_folder = os.path.join(ZIP_835_HTML_FOLDER, "Renamed")

    for filename in html_files:
        full_path = os.path.join(ZIP_835_HTML_FOLDER, filename)
        orig_ext = os.path.splitext(filename)[1]
        try:
            with open(full_path, "r", errors="ignore") as handle:
                content = handle.read()
        except Exception:
            continue

        matched_check = ""
        for check_number in check_candidates:
            for candidate in _era_check_candidates(check_number):
                if candidate and candidate in content:
                    matched_check = check_number
                    break
            if matched_check:
                break

        if not matched_check:
            continue

        new_name = f"{date_prefix}-835-{sequence}-{matched_check}{orig_ext}"
        destination = os.path.join(renamed_folder, new_name)

        if os.path.exists(destination):
            raise HTTPException(status_code=409, detail=f"Destination already exists: {new_name}")

        shutil.move(full_path, destination)
        renamed.append(
            {
                "sourceFile": filename,
                "renamedFile": new_name,
                "checkNumber": matched_check,
            }
        )
        sequence += 1

    return {
        "status": "converted" if renamed else "noop",
        "statusTag": "CONVERTED" if renamed else "NO FILES",
        "message": (
            f"Renamed {len(renamed)} HTML file(s) and moved them to 3.HTML/Renamed."
            if renamed
            else "No HTML files matched the selected day."
        ),
        "workDay": work_day,
        "bankDay": bank_day,
        "renamedCount": len(renamed),
        "outputFolder": renamed_folder,
        "renamedFiles": renamed,
    }


@app.get("/otherday/spreadsheet")
def get_otherday_spreadsheet():
    current_work_day = normalize_mmddyyyy(get_current_work_day() or "")
    if not current_work_day:
        raise HTTPException(status_code=400, detail="No current work day is set")

    conn = get_conn()
    try:
        row = conn.execute(
            "SELECT bank_day FROM calendar WHERE paperwork_day = ?",
            (current_work_day,),
        ).fetchone()
        bank_day = normalize_mmddyyyy(row[0]) if row and row[0] else None
        if not bank_day:
            return {
                "currentWorkDay": current_work_day,
                "bankDay": None,
                "rows": [],
                "missingRows": [],
                "rowCount": 0,
                "missingCount": 0,
                "filenamesWithMissing": 0,
            }

        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            """
            SELECT
                check_date,
                check_number,
                check_amount,
                filename,
                matchstatus
            FROM EDI
            ORDER BY check_date ASC, check_number ASC, filename ASC
            """
        ).fetchall()

        def _clean_filename(value: str | None) -> str:
            return str(value or "").strip()

        def _clean_check_number(value: str | None) -> str:
            return str(value or "").strip()

        def _format_amount(value):
            if value in (None, ""):
                return ""
            try:
                return f"{float(value):,.2f}"
            except Exception:
                return str(value)

        count_all: dict[str, int] = {}
        for row_data in rows:
            filename = _clean_filename(row_data["filename"])
            if filename:
                count_all[filename] = count_all.get(filename, 0) + 1

        today_rows = []
        for row_data in rows:
            if normalize_mmddyyyy(row_data["check_date"]) != bank_day:
                continue

            filename = _clean_filename(row_data["filename"])
            if not filename:
                continue

            today_rows.append(
                {
                    "filename": filename,
                    "checkNumber": _clean_check_number(row_data["check_number"]),
                    "ediAmount": _format_amount(row_data["check_amount"]),
                    "bankDay": normalize_mmddyyyy(row_data["check_date"]) or "",
                    "matchstatus": str(row_data["matchstatus"] or "").strip(),
                    "counts": "",
                }
            )

        count_today: dict[str, int] = {}
        for row_data in today_rows:
            filename = row_data["filename"]
            count_today[filename] = count_today.get(filename, 0) + 1

        for row_data in today_rows:
            filename = row_data["filename"]
            row_data["counts"] = f"{count_all.get(filename, 0)} {count_today.get(filename, 0)}"

        filenames_with_missing = [
            filename
            for filename in dict.fromkeys(row_data["filename"] for row_data in today_rows if row_data["filename"])
            if count_all.get(filename, 0) > count_today.get(filename, 0)
        ]

        missing_rows = []
        for row_data in rows:
            filename = _clean_filename(row_data["filename"])
            if not filename or filename not in filenames_with_missing:
                continue

            missing_rows.append(
                {
                    "filename": filename,
                    "checkNumber": _clean_check_number(row_data["check_number"]),
                    "ediAmount": _format_amount(row_data["check_amount"]),
                    "bankDay": normalize_mmddyyyy(row_data["check_date"]) or "",
                    "matchstatus": str(row_data["matchstatus"] or "").strip(),
                    "counts": f"{count_all.get(filename, 0)} {count_today.get(filename, 0)}",
                }
            )

        today_rows.sort(key=lambda item: (item["filename"], item["checkNumber"]))
        missing_rows.sort(key=lambda item: (item["filename"], item["checkNumber"], item["bankDay"]))

        return {
            "currentWorkDay": current_work_day,
            "bankDay": bank_day,
            "rows": today_rows,
            "missingRows": missing_rows,
            "rowCount": len(today_rows),
            "missingCount": len(missing_rows),
            "filenamesWithMissing": len(filenames_with_missing),
        }
    finally:
        conn.close()


@app.get("/duplicatecheck/spreadsheet")
def get_duplicatecheck_spreadsheet():
    current_work_day = normalize_mmddyyyy(get_current_work_day() or "")
    if not current_work_day:
        raise HTTPException(status_code=400, detail="No current work day is set")

    conn = get_conn()
    try:
        row = conn.execute(
            "SELECT bank_day FROM calendar WHERE paperwork_day = ?",
            (current_work_day,),
        ).fetchone()
        bank_day = normalize_mmddyyyy(row[0]) if row and row[0] else None
        if not bank_day:
            return {
                "currentWorkDay": current_work_day,
                "bankDay": None,
                "rows": [],
                "duplicateCount": 0,
                "duplicateFilenames": 0,
                "duplicateFilenameList": [],
            }

        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            """
            SELECT
                e.filename AS filename,
                e.check_number AS edi_check,
                e.check_amount AS edi_amount,
                lb.[Transaction Total] AS lockbox_amount,
                eft.Amount AS eft_amount,
                COALESCE(eft.Date, lb.[Deposit Date]) AS match_date
            FROM EDI e
            LEFT JOIN Lockbox lb
                ON TRIM(e.check_number) = TRIM(lb.[Check Number])
            LEFT JOIN EFT eft
                ON TRIM(e.check_number) = TRIM(eft.CheckNumber)
            WHERE COALESCE(eft.Date, lb.[Deposit Date]) IS NOT NULL
            ORDER BY e.filename, e.check_number
            """
        ).fetchall()

        filtered = [
            row_data
            for row_data in rows
            if normalize_mmddyyyy(row_data["match_date"]) == bank_day
        ]

        filename_counts: dict[str, int] = {}
        for row_data in filtered:
            filename = str(row_data["filename"] or "").strip()
            if filename:
                filename_counts[filename] = filename_counts.get(filename, 0) + 1

        duplicate_filenames = {filename for filename, count in filename_counts.items() if count > 1}

        table_rows = []
        for row_data in filtered:
            filename = str(row_data["filename"] or "").strip()
            if not filename or filename not in duplicate_filenames:
                continue

            table_rows.append(
                {
                    "filename": filename,
                    "ediCheck": str(row_data["edi_check"] or "").strip(),
                    "lockboxAmount": str(row_data["lockbox_amount"] or "").strip(),
                    "eftAmount": str(row_data["eft_amount"] or "").strip(),
                    "date": normalize_mmddyyyy(row_data["match_date"]) or "",
                    "count": filename_counts.get(filename, 0),
                }
            )

        table_rows.sort(key=lambda item: (item["filename"], item["ediCheck"], item["date"]))

        return {
            "currentWorkDay": current_work_day,
            "bankDay": bank_day,
            "rows": table_rows,
            "duplicateCount": len(table_rows),
            "duplicateFilenames": len(duplicate_filenames),
            "duplicateFilenameList": sorted(duplicate_filenames),
        }
    finally:
        conn.close()


@app.get("/balsheet/workday")
def get_balsheet_workday():
    init_db()
    conn = get_conn()
    try:
        row = conn.execute(
            "SELECT current_bank_day, current_work_day, message FROM work_state WHERE id = 1"
        ).fetchone()
    finally:
        conn.close()

    current_bank_day = normalize_mmddyyyy(row[0] if row else None) if row else None
    current_work_day = normalize_mmddyyyy(row[1] if row else None) if row else None
    message = str(row[2] or "") if row else ""
    posting_date = current_work_day or current_bank_day
    if not posting_date:
        posting_date = datetime.today().strftime("%m/%d/%Y")

    return {
        "posting_date": posting_date,
        "current_bank_day": current_bank_day,
        "current_work_day": current_work_day,
        "message": message,
    }


@app.put("/balsheet/workday/message")
def set_balsheet_workday_message(payload: dict | None = None):
    init_db()
    conn = get_conn()
    try:
        payload = payload or {}
        message = str(payload.get("message") or "").strip()
        conn.execute(
            "UPDATE work_state SET message = ? WHERE id = 1",
            (message,),
        )
        conn.commit()
        return {
            "status": "ok",
            "message": message,
        }
    finally:
        conn.close()


@app.get("/balsheet")
def get_balsheet(posting_date: str | None = None):
    init_db()
    conn = get_conn()
    ensure_balsheet_table(conn)

    try:
        conn.row_factory = sqlite3.Row
        normalized_posting_date = normalize_mmddyyyy(posting_date) if posting_date else None
        if normalized_posting_date:
            rows = conn.execute(
                f'SELECT * FROM {_quote_identifier("Balsheet")} WHERE {_quote_identifier("PostingDate")} = ? {_balsheet_order_clause()}',
                (normalized_posting_date,),
            ).fetchall()
        else:
            rows = conn.execute(
                f'SELECT * FROM {_quote_identifier("Balsheet")} {_balsheet_order_clause()}'
            ).fetchall()

        return [_balsheet_row_to_payload(row) for row in rows]
    finally:
        conn.close()


@app.get("/imaging/balsheet-associations")
def get_imaging_balsheet_associations(posting_date: str):
    normalized_posting_date = normalize_mmddyyyy(posting_date)
    if not normalized_posting_date:
        raise HTTPException(status_code=400, detail="posting_date is required")

    init_db()
    conn = get_conn()
    try:
        ensure_balsheet_table(conn)
        ensure_imaging_tables(conn)
        return _build_imaging_association_payload(conn, normalized_posting_date)
    finally:
        conn.close()


@app.post("/imaging/balsheet-associations/refresh")
def refresh_imaging_balsheet_associations(payload: dict | None = None):
    payload = payload or {}
    normalized_posting_date = normalize_mmddyyyy(payload.get("posting_date"))
    if not normalized_posting_date:
        raise HTTPException(status_code=400, detail="posting_date is required")

    init_db()
    conn = get_conn()
    try:
        ensure_balsheet_table(conn)
        ensure_imaging_tables(conn)
        rebuild_imaging_document_index(conn)
        return _build_imaging_association_payload(conn, normalized_posting_date)
    finally:
        conn.close()


@app.get("/imaging/lockbox-associations")
def get_imaging_lockbox_associations(posting_date: str, query: str = ""):
    normalized_posting_date = normalize_mmddyyyy(posting_date)
    if not normalized_posting_date:
        raise HTTPException(status_code=400, detail="posting_date is required")

    init_db()
    conn = get_conn()
    try:
        ensure_imaging_tables(conn)
        root = _imaging_root_folder()
        if not root.exists():
            return {
                "postingDate": normalized_posting_date,
                "query": str(query or "").strip(),
                "results": [],
            }

        query_value = str(query or "").strip()
        search_results = _search_lockbox_documents(normalized_posting_date, query_value)
        return {
            "postingDate": normalized_posting_date,
            "query": query_value,
            "results": search_results,
        }
    finally:
        conn.close()


@app.post("/imaging/lockbox-associations/find-matches")
def find_imaging_lockbox_matches(payload: dict | None = None):
    payload = payload or {}
    normalized_posting_date = normalize_mmddyyyy(payload.get("posting_date"))
    if not normalized_posting_date:
        raise HTTPException(status_code=400, detail="posting_date is required")

    init_db()
    conn = get_conn()
    try:
        ensure_balsheet_table(conn)
        ensure_imaging_tables(conn)
        rebuild_imaging_document_index(conn)
        return _build_imaging_lockbox_recommendations_payload(conn, normalized_posting_date)
    finally:
        conn.close()


@app.post("/imaging/balsheet-links/confirm")
def confirm_imaging_balsheet_link(payload: dict | None = None):
    payload = payload or {}
    entry_id = str(payload.get("entry_id") or "").strip()
    file_path = str(payload.get("file_path") or "").strip()
    if not entry_id or not file_path:
        raise HTTPException(status_code=400, detail="entry_id and file_path are required")

    init_db()
    conn = get_conn()
    try:
        ensure_balsheet_table(conn)
        ensure_imaging_tables(conn)
        conn.row_factory = sqlite3.Row

        result = _upsert_imaging_balsheet_link(
            conn=conn,
            entry_id=entry_id,
            file_path=file_path,
            link_id=str(payload.get("link_id") or "").strip() or f"IML-{uuid.uuid4().hex[:12].upper()}",
            check_number=str(payload.get("check_number") or ""),
            match_method=str(payload.get("match_method") or "manual").strip() or "manual",
            confidence=float(payload.get("confidence") or 1),
            bookmark_page=int(payload.get("bookmark_page") or 0),
            bookmark_title=str(payload.get("bookmark_title") or ""),
            source_query=str(payload.get("source_query") or ""),
            lockbox_image_date=normalize_mmddyyyy(payload.get("lockbox_image_date"))
            or normalize_mmddyyyy(payload.get("posting_date"))
            or "",
        )
        conn.commit()
        return {
            "status": "ok",
            "linkId": result["linkId"],
            "entryId": entry_id,
            "filePath": file_path,
            "bookmarkPage": result.get("bookmarkPage", 0),
            "bookmarkTitle": result.get("bookmarkTitle", ""),
            "lockboxImageDate": result.get("lockboxImageDate", ""),
        }
    finally:
        conn.close()


@app.delete("/imaging/balsheet-links/{link_id}")
def delete_imaging_balsheet_link(link_id: str):
    link_id = str(link_id or "").strip()
    if not link_id:
        raise HTTPException(status_code=400, detail="link_id is required")

    init_db()
    conn = get_conn()
    try:
        ensure_imaging_tables(conn)
        cur = conn.execute(
            f'DELETE FROM {_quote_identifier("Imaging_BalsheetDocumentLinks")} WHERE {_quote_identifier("link_id")} = ?',
            (link_id,),
        )
        if cur.rowcount <= 0:
            raise HTTPException(status_code=404, detail="Link not found")
        conn.commit()
        return {"status": "ok", "linkId": link_id}
    finally:
        conn.close()


def _upsert_imaging_balsheet_link(
    *,
    conn,
    entry_id: str,
    file_path: str,
    link_id: str,
    check_number: str,
    match_method: str,
    confidence: float,
    bookmark_page: int = 0,
    bookmark_title: str = "",
    source_query: str = "",
    lockbox_image_date: str = "",
):
    balsheet_row = conn.execute(
        f'SELECT * FROM {_quote_identifier("Balsheet")} WHERE {_quote_identifier("EntryID")} = ?',
        (entry_id,),
    ).fetchone()
    if not balsheet_row:
        raise HTTPException(status_code=404, detail="Balsheet row not found")

    document_row = conn.execute(
        f'SELECT * FROM {_quote_identifier("Imaging_DocumentFileIndex")} WHERE {_quote_identifier("file_path")} = ?',
        (file_path,),
    ).fetchone()
    if not document_row:
        rebuild_imaging_document_index(conn)
        document_row = conn.execute(
            f'SELECT * FROM {_quote_identifier("Imaging_DocumentFileIndex")} WHERE {_quote_identifier("file_path")} = ?',
            (file_path,),
        ).fetchone()
    if not document_row:
        raise HTTPException(status_code=404, detail="Document file not found")

    now = datetime.now().isoformat(timespec="seconds")
    normalized_posting_date = normalize_mmddyyyy(balsheet_row["PostingDate"]) or str(balsheet_row["PostingDate"] or "")
    conn.execute(
        f"""
        INSERT INTO {_quote_identifier("Imaging_BalsheetDocumentLinks")} (
            {_quote_identifier("link_id")},
            {_quote_identifier("entry_id")},
            {_quote_identifier("posting_date")},
            {_quote_identifier("lockbox_image_date")},
            {_quote_identifier("amount")},
            {_quote_identifier("payer")},
            {_quote_identifier("check_number")},
            {_quote_identifier("file_path")},
            {_quote_identifier("file_name")},
            {_quote_identifier("match_method")},
            {_quote_identifier("confidence")},
            {_quote_identifier("bookmark_page")},
            {_quote_identifier("bookmark_title")},
            {_quote_identifier("source_query")},
            {_quote_identifier("confirmed")},
            {_quote_identifier("created_at")},
            {_quote_identifier("updated_at")}
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?)
        ON CONFLICT({_quote_identifier("entry_id")}, {_quote_identifier("file_path")}) DO UPDATE SET
            {_quote_identifier("posting_date")} = excluded.{_quote_identifier("posting_date")},
            {_quote_identifier("lockbox_image_date")} = excluded.{_quote_identifier("lockbox_image_date")},
            {_quote_identifier("amount")} = excluded.{_quote_identifier("amount")},
            {_quote_identifier("payer")} = excluded.{_quote_identifier("payer")},
            {_quote_identifier("check_number")} = excluded.{_quote_identifier("check_number")},
            {_quote_identifier("file_name")} = excluded.{_quote_identifier("file_name")},
            {_quote_identifier("match_method")} = excluded.{_quote_identifier("match_method")},
            {_quote_identifier("confidence")} = excluded.{_quote_identifier("confidence")},
            {_quote_identifier("bookmark_page")} = excluded.{_quote_identifier("bookmark_page")},
            {_quote_identifier("bookmark_title")} = excluded.{_quote_identifier("bookmark_title")},
            {_quote_identifier("source_query")} = excluded.{_quote_identifier("source_query")},
            {_quote_identifier("confirmed")} = excluded.{_quote_identifier("confirmed")},
            {_quote_identifier("updated_at")} = excluded.{_quote_identifier("updated_at")}
        """,
        (
            link_id,
            entry_id,
            normalized_posting_date,
            str(lockbox_image_date or normalized_posting_date or ""),
            float(balsheet_row["Amount"] or 0),
            str(balsheet_row["Payer"] or ""),
            _normalize_imaging_check_number(check_number or balsheet_row["Check Number"] or ""),
            file_path,
            str(document_row["file_name"] or ""),
            match_method,
            confidence,
            int(bookmark_page or 0),
            str(bookmark_title or ""),
            str(source_query or ""),
            now,
            now,
        ),
    )

    return {
        "linkId": link_id,
        "entryId": entry_id,
        "filePath": file_path,
        "documentFileName": str(document_row["file_name"] or ""),
        "bookmarkPage": int(bookmark_page or 0),
        "bookmarkTitle": str(bookmark_title or ""),
        "lockboxImageDate": str(lockbox_image_date or normalized_posting_date or ""),
    }


@app.post("/imaging/balsheet-associations/confirm-exact")
def confirm_imaging_balsheet_exact_matches(payload: dict | None = None):
    payload = payload or {}
    posting_date = normalize_mmddyyyy(payload.get("posting_date"))
    if not posting_date:
        raise HTTPException(status_code=400, detail="posting_date is required")

    init_db()
    conn = get_conn()
    try:
        ensure_balsheet_table(conn)
        ensure_imaging_tables(conn)
        conn.row_factory = sqlite3.Row

        association_payload = _build_imaging_association_payload(conn, posting_date)
        committed = []
        skipped = []

        for row in association_payload["rows"]:
            if row["linkedFiles"]:
                skipped.append({"entryId": row["entryId"], "reason": "already_linked"})
                continue

            top_match = row["matches"][0] if row["matches"] else None
            if not top_match or float(top_match.get("confidence") or 0) < 1.0:
                skipped.append({"entryId": row["entryId"], "reason": "no_exact_match"})
                continue

            result = _upsert_imaging_balsheet_link(
                conn=conn,
                entry_id=row["entryId"],
                file_path=str(top_match["filePath"]),
                link_id=f"IML-{uuid.uuid4().hex[:12].upper()}",
                check_number=row["checkNumber"],
                match_method=str(top_match.get("matchMethod") or "exact"),
                confidence=float(top_match.get("confidence") or 1),
            )
            committed.append(result)

        conn.commit()
        refreshed = _build_imaging_association_payload(conn, posting_date)
        return {
            "status": "ok",
            "postingDate": posting_date,
            "committedCount": len(committed),
            "skippedCount": len(skipped),
            "committed": committed,
            "skipped": skipped,
            "data": refreshed,
        }
    finally:
        conn.close()


@app.get("/imaging/balsheet-links/{link_id}/open")
def open_imaging_balsheet_link(link_id: str):
    conn = get_conn()
    try:
        ensure_imaging_tables(conn)
        row = conn.execute(
            f'SELECT {_quote_identifier("file_path")} FROM {_quote_identifier("Imaging_BalsheetDocumentLinks")} WHERE {_quote_identifier("link_id")} = ?',
            (link_id,),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Link not found")
        return _safe_imaging_file_response(str(row[0]))
    finally:
        conn.close()


@app.get("/imaging/files/open")
def open_imaging_file(path: str):
    if not path:
        raise HTTPException(status_code=400, detail="path is required")
    return _safe_imaging_file_response(path)


@app.post("/balsheet/import-banking")
def import_balsheet_from_banking(payload: dict | None = None):
    init_db()
    conn = get_conn()
    ensure_balsheet_table(conn)

    payload = payload or {}
    posting_date = normalize_mmddyyyy(payload.get("posting_date")) or normalize_mmddyyyy(get_current_work_day() or "") or normalize_mmddyyyy(get_current_bank_day() or "") or datetime.today().strftime("%m/%d/%Y")
    bank_day = normalize_mmddyyyy(payload.get("bank_day")) or normalize_mmddyyyy(get_current_bank_day() or "")
    if not bank_day:
        raise HTTPException(status_code=400, detail="bank_day is required")

    imported_rows = 0
    removed_rows = 0

    try:
        spreadsheet = build_banking_spreadsheet()

        cur = conn.cursor()
        cur.execute(
            f'DELETE FROM {_quote_identifier("Balsheet")} WHERE {_quote_identifier("PostingDate")} = ? AND {_quote_identifier("EntryID")} LIKE ?',
            (posting_date, "BANK-%"),
        )
        removed_rows = cur.rowcount if cur.rowcount is not None else 0

        for group in spreadsheet.get("groups", []):
            source = str(group.get("source") or "")
            type_value = "EFT" if source == "EFT" else "Lockbox"
            rows = group.get("rows", [])
            if not isinstance(rows, list):
                continue

            for row in rows:
                if not isinstance(row, dict):
                    continue

                row_date = normalize_mmddyyyy(row.get("date")) or ""
                if row_date != bank_day:
                    continue

                edi_value = str(row.get("edi", "") or "").strip().upper()
                poster_value = "Raul" if edi_value == "Y" else "Nick"

                entry = {
                    "entry_id": f'BANK-{source}-{row.get("id")}',
                    "posting_date": posting_date,
                    "type": type_value,
                    "amount": row.get("amount", 0),
                    "payer": row.get("payer", ""),
                    "check_number": row.get("checkNumber", ""),
                    "edi": row.get("edi", ""),
                    "poster": poster_value,
                    "eob": "",
                    "unposted": 0,
                    "misc": 0,
                    "misc_type": "",
                    "notes": "",
                    "nick": 0,
                    "raul": 0,
                    "needs": "",
                    "from_date": "",
                    "to_date": "",
                }
                _balsheet_insert_or_replace(conn, entry)
                imported_rows += 1

        conn.commit()
        return {
            "status": "ok",
            "postingDate": posting_date,
            "rowsImported": imported_rows,
            "rowsRemoved": removed_rows,
        }
    finally:
        conn.close()


@app.delete("/balsheet")
def clear_balsheet(posting_date: str | None = None):
    init_db()
    normalized_posting_date = normalize_mmddyyyy(posting_date or "")
    if not normalized_posting_date:
        raise HTTPException(status_code=400, detail="posting_date is required")

    conn = get_conn()
    ensure_balsheet_table(conn)

    try:
        cur = conn.cursor()
        cur.execute(
            f'DELETE FROM {_quote_identifier("Balsheet")} WHERE {_quote_identifier("PostingDate")} = ?',
            (normalized_posting_date,),
        )
        deleted_rows = cur.rowcount if cur.rowcount is not None else 0
        conn.commit()
        return {
            "status": "ok",
            "postingDate": normalized_posting_date,
            "rowsDeleted": deleted_rows,
        }
    finally:
        conn.close()


@app.get("/balsheet/notes")
def get_balsheet_notes(post_date: str | None = None):
    init_db()
    conn = get_conn()
    ensure_balsheet_notes_table(conn)

    try:
        conn.row_factory = sqlite3.Row
        normalized_post_date = normalize_mmddyyyy(post_date) if post_date else None
        if normalized_post_date:
            rows = conn.execute(
                f'SELECT rowid, {_quote_identifier("post_date")}, {_quote_identifier("notes")}, {_quote_identifier("message")} FROM {_quote_identifier("Balsheet_notes")} WHERE {_quote_identifier("post_date")} = ? ORDER BY rowid ASC',
                (normalized_post_date,),
            ).fetchall()
        else:
            rows = conn.execute(
                f'SELECT rowid, {_quote_identifier("post_date")}, {_quote_identifier("notes")}, {_quote_identifier("message")} FROM {_quote_identifier("Balsheet_notes")} ORDER BY {_quote_identifier("post_date")} ASC, rowid ASC'
            ).fetchall()

        return [_balsheet_note_row_to_payload(row) for row in rows]
    finally:
        conn.close()


@app.post("/balsheet/notes")
def post_balsheet_note(note: dict):
    init_db()
    conn = get_conn()
    ensure_balsheet_notes_table(conn)

    try:
        normalized = _normalize_balsheet_note_payload(note)
        cur = conn.cursor()
        cur.execute(
            f'INSERT INTO {_quote_identifier("Balsheet_notes")} ({_quote_identifier("post_date")}, {_quote_identifier("notes")}, {_quote_identifier("message")}) VALUES (?, ?, ?)',
            (normalized["post_date"], normalized["notes"], normalized["message"]),
        )
        conn.commit()
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            f'SELECT rowid, {_quote_identifier("post_date")}, {_quote_identifier("notes")}, {_quote_identifier("message")} FROM {_quote_identifier("Balsheet_notes")} WHERE rowid = ?',
            (cur.lastrowid,),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=500, detail="Failed to save Balsheet note")
        return _balsheet_note_row_to_payload(row)
    finally:
        conn.close()


@app.put("/balsheet/notes/{rowid}")
def put_balsheet_note(rowid: int, note: dict):
    init_db()
    conn = get_conn()
    ensure_balsheet_notes_table(conn)

    try:
        conn.row_factory = sqlite3.Row
        existing = conn.execute(
            f'SELECT rowid, {_quote_identifier("post_date")}, {_quote_identifier("notes")}, {_quote_identifier("message")} FROM {_quote_identifier("Balsheet_notes")} WHERE rowid = ?',
            (rowid,),
        ).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Balsheet note not found")

        normalized = _normalize_balsheet_note_payload(note, rowid=rowid)
        conn.execute(
            f'UPDATE {_quote_identifier("Balsheet_notes")} SET {_quote_identifier("post_date")} = ?, {_quote_identifier("notes")} = ?, {_quote_identifier("message")} = ? WHERE rowid = ?',
            (normalized["post_date"], normalized["notes"], normalized["message"], rowid),
        )
        conn.commit()

        row = conn.execute(
            f'SELECT rowid, {_quote_identifier("post_date")}, {_quote_identifier("notes")}, {_quote_identifier("message")} FROM {_quote_identifier("Balsheet_notes")} WHERE rowid = ?',
            (rowid,),
        ).fetchone()
        return _balsheet_note_row_to_payload(row)
    finally:
        conn.close()


@app.delete("/balsheet/notes/{rowid}")
def delete_balsheet_note(rowid: int):
    init_db()
    conn = get_conn()
    ensure_balsheet_notes_table(conn)

    try:
        cur = conn.cursor()
        cur.execute(
            f'DELETE FROM {_quote_identifier("Balsheet_notes")} WHERE rowid = ?',
            (rowid,),
        )
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="Balsheet note not found")
        conn.commit()
        return {"status": "ok", "rowid": rowid}
    finally:
        conn.close()


def _generate_misc_id() -> str:
    return f"MISC-{datetime.now().strftime('%m%d%Y-%H%M%S%f')}"


def _misc_row_to_payload(row):
    return {
        "misc_id": str(row["misc_id"] or ""),
        "posting_date": normalize_mmddyyyy(row["posting_date"]) or str(row["posting_date"] or ""),
        "amount": row["amount"],
        "misc_type": str(row["misc_type"] or ""),
        "details": str(row["details"] or ""),
        "created_at": str(row["created_at"] or ""),
    }


def _normalize_misc_payload(payload: dict, misc_id: str | None = None):
    posting_date = normalize_mmddyyyy(payload.get("posting_date")) or ""
    if not posting_date:
        raise HTTPException(status_code=400, detail="posting_date is required")

    return {
        "misc_id": misc_id or str(payload.get("misc_id") or "").strip() or _generate_misc_id(),
        "posting_date": posting_date,
        "amount": _normalize_balsheet_amount(payload.get("amount")),
        "misc_type": str(payload.get("misc_type") or "").strip(),
        "details": str(payload.get("details") or "").strip(),
        "created_at": str(payload.get("created_at") or "").strip() or datetime.now().isoformat(timespec="seconds"),
    }


@app.get("/misc")
def get_misc(posting_date: str | None = None):
    init_db()
    conn = get_conn()
    ensure_misc_table(conn)

    try:
        conn.row_factory = sqlite3.Row
        normalized_posting_date = normalize_mmddyyyy(posting_date) if posting_date else None
        if normalized_posting_date:
            rows = conn.execute(
                f'SELECT * FROM {_quote_identifier("Misc")} WHERE {_quote_identifier("posting_date")} = ? ORDER BY {_quote_identifier("created_at")} ASC, {_quote_identifier("misc_id")} ASC',
                (normalized_posting_date,),
            ).fetchall()
        else:
            rows = conn.execute(
                f'SELECT * FROM {_quote_identifier("Misc")} ORDER BY {_quote_identifier("posting_date")} ASC, {_quote_identifier("created_at")} ASC, {_quote_identifier("misc_id")} ASC'
            ).fetchall()
        return [_misc_row_to_payload(row) for row in rows]
    finally:
        conn.close()


@app.post("/misc")
def post_misc(payload: dict):
    init_db()
    conn = get_conn()
    ensure_misc_table(conn)

    if not isinstance(payload, dict):
        raise HTTPException(status_code=400, detail="Misc payload must be an object")

    try:
        normalized = _normalize_misc_payload(payload)
        columns = [name for name, _ in MISC_TABLE_COLUMNS]
        conn.execute(
            f'INSERT INTO {_quote_identifier("Misc")} ({", ".join(_quote_identifier(name) for name in columns)}) VALUES ({", ".join(["?"] * len(columns))})',
            tuple(normalized[name] for name in columns),
        )
        conn.commit()
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            f'SELECT * FROM {_quote_identifier("Misc")} WHERE {_quote_identifier("misc_id")} = ?',
            (normalized["misc_id"],),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=500, detail="Failed to save Misc entry")
        return _misc_row_to_payload(row)
    finally:
        conn.close()


@app.put("/misc/{misc_id}")
def put_misc(misc_id: str, payload: dict):
    init_db()
    conn = get_conn()
    ensure_misc_table(conn)

    if not isinstance(payload, dict):
        raise HTTPException(status_code=400, detail="Misc payload must be an object")

    try:
        conn.row_factory = sqlite3.Row
        existing = conn.execute(
            f'SELECT * FROM {_quote_identifier("Misc")} WHERE {_quote_identifier("misc_id")} = ?',
            (misc_id,),
        ).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Misc entry not found")

        normalized = _normalize_misc_payload(payload, misc_id=misc_id)
        set_clause = ", ".join(f'{_quote_identifier(name)} = ?' for name, _ in MISC_TABLE_COLUMNS[1:])
        conn.execute(
            f'UPDATE {_quote_identifier("Misc")} SET {set_clause} WHERE {_quote_identifier("misc_id")} = ?',
            tuple(normalized[name] for name, _ in MISC_TABLE_COLUMNS[1:]) + (misc_id,),
        )
        conn.commit()

        row = conn.execute(
            f'SELECT * FROM {_quote_identifier("Misc")} WHERE {_quote_identifier("misc_id")} = ?',
            (misc_id,),
        ).fetchone()
        return _misc_row_to_payload(row)
    finally:
        conn.close()


@app.delete("/misc/{misc_id}")
def delete_misc(misc_id: str):
    init_db()
    conn = get_conn()
    ensure_misc_table(conn)

    try:
        cur = conn.cursor()
        cur.execute(
            f'DELETE FROM {_quote_identifier("Misc")} WHERE {_quote_identifier("misc_id")} = ?',
            (misc_id,),
        )
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="Misc entry not found")
        conn.commit()
        return {"status": "ok", "misc_id": misc_id}
    finally:
        conn.close()


@app.post("/balsheet")
def post_balsheet(entry: dict):
    init_db()
    conn = get_conn()
    ensure_balsheet_table(conn)

    try:
        entry_id = _balsheet_insert_or_replace(conn, entry)
        conn.commit()
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            f'SELECT * FROM {_quote_identifier("Balsheet")} WHERE {_quote_identifier("EntryID")} = ?',
            (entry_id,),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=500, detail="Failed to save Balsheet entry")
        return _balsheet_row_to_payload(row)
    finally:
        conn.close()


@app.post("/balsheet/bulk")
def post_balsheet_bulk(payload: dict):
    init_db()
    conn = get_conn()
    ensure_balsheet_table(conn)

    entries = payload.get("entries", [])
    source_attachment_id = payload.get("source_attachment_id")
    posting_date = normalize_mmddyyyy(payload.get("posting_date")) if payload.get("posting_date") else None
    if not isinstance(entries, list):
        raise HTTPException(status_code=400, detail="entries must be a list")

    inserted = 0
    try:
        for entry in entries:
            if isinstance(entry, dict):
                next_entry = dict(entry)
                if posting_date:
                    next_entry["posting_date"] = posting_date
                _balsheet_insert_or_replace(conn, next_entry)
            else:
                _balsheet_insert_or_replace(conn, {"posting_date": posting_date} if posting_date else {})
            inserted += 1
        conn.commit()
        return {
            "status": "ok",
            "rowsImported": inserted,
            "sourceAttachmentId": source_attachment_id,
        }
    finally:
        conn.close()


@app.put("/balsheet/{entry_id}")
def put_balsheet_entry(entry_id: str, entry: dict):
    init_db()
    conn = get_conn()
    ensure_balsheet_table(conn)

    try:
        conn.row_factory = sqlite3.Row
        existing = conn.execute(
            f'SELECT * FROM {_quote_identifier("Balsheet")} WHERE {_quote_identifier("EntryID")} = ?',
            (entry_id,),
        ).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Balsheet entry not found")

        normalized = _normalize_balsheet_payload(entry, entry_id=entry_id)
        set_clause = ", ".join(f'{_quote_identifier(name)} = ?' for name, _ in BALSHEET_TABLE_COLUMNS[1:])
        conn.execute(
            f'UPDATE {_quote_identifier("Balsheet")} SET {set_clause} WHERE {_quote_identifier("EntryID")} = ?',
            tuple(normalized[name] for name, _ in BALSHEET_TABLE_COLUMNS[1:]) + (entry_id,),
        )
        conn.commit()

        row = conn.execute(
            f'SELECT * FROM {_quote_identifier("Balsheet")} WHERE {_quote_identifier("EntryID")} = ?',
            (entry_id,),
        ).fetchone()
        return _balsheet_row_to_payload(row)
    finally:
        conn.close()


@app.delete("/balsheet/{entry_id}")
def delete_balsheet_entry(entry_id: str):
    init_db()
    conn = get_conn()
    ensure_balsheet_table(conn)

    try:
        cur = conn.cursor()
        cur.execute(
            f'DELETE FROM {_quote_identifier("Balsheet")} WHERE {_quote_identifier("EntryID")} = ?',
            (entry_id,),
        )
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="Balsheet entry not found")
        conn.commit()
        return {"status": "ok", "entry_id": entry_id}
    finally:
        conn.close()


# ------------------------------------------------------------
# NEXT PENDING FILE
# ------------------------------------------------------------
@app.get("/attachments/{attachment_id}/next")
def get_next(attachment_id: int, day: str | None = None):
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, filename, site, snapshot_path, review_status, batch_date, batch_id, processed_at
        FROM imported_files
        WHERE review_status = 'Pending' AND id > ?
        ORDER BY id ASC
    """, (attachment_id,))

    rows = cur.fetchall()
    conn.close()

    desired_day = _normalize_pending_day(day) if day else None
    row = None
    for candidate in rows:
        if desired_day and _row_pending_day(candidate, 5, 6, 7) != desired_day:
            continue
        row = candidate
        break

    if not row:
        return {"done": True}

    return {
        "id": row[0],
        "filename": row[1],
        "site": row[2],
        "snapshot": row[3],
        "status": row[4],
        "done": False
    }


@app.get("/attachments/{attachment_id}/previous")
def get_previous(attachment_id: int, day: str | None = None):
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, filename, site, snapshot_path, review_status, batch_date, batch_id, processed_at
        FROM imported_files
        WHERE review_status = 'Pending' AND id < ?
        ORDER BY id DESC
    """, (attachment_id,))

    rows = cur.fetchall()
    conn.close()

    desired_day = _normalize_pending_day(day) if day else None
    row = None
    for candidate in rows:
        if desired_day and _row_pending_day(candidate, 5, 6, 7) != desired_day:
            continue
        row = candidate
        break

    if not row:
        return {"done": True}

    return {
        "id": row[0],
        "filename": row[1],
        "site": row[2],
        "snapshot": row[3],
        "status": row[4],
        "done": False
    }


@app.put("/attachments/{attachment_id}/site")
def update_attachment_site(attachment_id: int, payload: dict):
    site = (payload.get("site") or "").strip()
    if not site:
        raise HTTPException(status_code=400, detail="site is required")

    conn = get_conn()
    cur = conn.cursor()

    cur.execute(
        """
        UPDATE imported_files
        SET site = ?
        WHERE id = ?
        """,
        (site, attachment_id),
    )
    if cur.rowcount == 0:
        conn.close()
        raise HTTPException(status_code=404, detail="Attachment not found")

    conn.commit()
    conn.close()
    return {"status": "ok", "id": attachment_id, "site": site}


# ------------------------------------------------------------
# SNAPSHOT IMAGE
# ------------------------------------------------------------
@app.get("/attachments/{attachment_id}/snapshot")
def get_snapshot(attachment_id: int):
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("SELECT snapshot_path FROM imported_files WHERE id = ?", (attachment_id,))
    row = cur.fetchone()

    if not row or not row[0]:
        fallback_path = os.path.join(SNAPSHOTS_FOLDER, f"{attachment_id}.png")
        if not os.path.exists(fallback_path):
            conn.close()
            raise HTTPException(status_code=404, detail="Snapshot not found")

        cur.execute(
            """
            UPDATE imported_files
            SET snapshot_path = ?
            WHERE id = ?
            """,
            (fallback_path, attachment_id),
        )
        conn.commit()
        conn.close()
        return FileResponse(fallback_path)

    snapshot_path = row[0]

    if not os.path.exists(snapshot_path):
        fallback_path = os.path.join(SNAPSHOTS_FOLDER, f"{attachment_id}.png")
        if not os.path.exists(fallback_path):
            conn.close()
            raise HTTPException(status_code=404, detail="Snapshot file missing")

        cur.execute(
            """
            UPDATE imported_files
            SET snapshot_path = ?
            WHERE id = ?
            """,
            (fallback_path, attachment_id),
        )
        conn.commit()
        conn.close()
        return FileResponse(fallback_path)

    conn.close()
    return FileResponse(snapshot_path)


@app.post("/attachments/{attachment_id}/repair-snapshot")
def repair_snapshot(attachment_id: int):
    conn = get_conn()
    cur = conn.cursor()

    cur.execute(
        """
        SELECT id, filename, original_filename, moved_to, snapshot_path
        FROM imported_files
        WHERE id = ?
        """,
        (attachment_id,),
    )
    row = cur.fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Attachment not found")

    _, filename, original_filename, moved_to, snapshot_path = row
    target_path = os.path.join(SNAPSHOTS_FOLDER, f"{attachment_id}.png")

    if snapshot_path and os.path.exists(snapshot_path):
        conn.close()
        return {
            "status": "ok",
            "mode": "existing",
            "id": attachment_id,
            "snapshot_path": snapshot_path,
            "source_path": snapshot_path,
        }

    candidate_paths = []
    for candidate in (
        moved_to,
        os.path.join(EMAILS_FOLDER, original_filename) if original_filename else None,
        os.path.join(EMAILS_FOLDER, filename) if filename else None,
    ):
        if candidate and os.path.exists(candidate):
            candidate_paths.append(candidate)

    if candidate_paths:
        try:
            from site_snapshotgenerator import _snapshot_for_file

            _snapshot_for_file(candidate_paths[0], target_path, original_filename or filename or os.path.basename(candidate_paths[0]))
        except Exception as exc:
            conn.close()
            raise HTTPException(status_code=500, detail=f"Failed to repair snapshot: {exc}") from exc

        cur.execute(
            """
            UPDATE imported_files
            SET snapshot_path = ?
            WHERE id = ?
            """,
            (target_path, attachment_id),
        )
        conn.commit()
        conn.close()
        return {
            "status": "ok",
            "mode": "generated",
            "id": attachment_id,
            "snapshot_path": target_path,
            "source_path": candidate_paths[0],
        }

    cur.execute(
        """
        SELECT snapshot_path
        FROM imported_files
        WHERE id != ?
          AND snapshot_path IS NOT NULL
          AND snapshot_path != ''
          AND (
            filename = ?
            OR original_filename = ?
          )
        ORDER BY CASE WHEN original_filename = ? THEN 0 ELSE 1 END, id ASC
        LIMIT 1
        """,
        (attachment_id, filename, original_filename, original_filename),
    )
    sibling = cur.fetchone()

    if sibling and sibling[0] and os.path.exists(sibling[0]):
        try:
            shutil.copy2(sibling[0], target_path)
        except Exception as exc:
            conn.close()
            raise HTTPException(status_code=500, detail=f"Failed to copy snapshot: {exc}") from exc

        cur.execute(
            """
            UPDATE imported_files
            SET snapshot_path = ?
            WHERE id = ?
            """,
            (target_path, attachment_id),
        )
        conn.commit()
        conn.close()
        return {
            "status": "ok",
            "mode": "copied",
            "id": attachment_id,
            "snapshot_path": target_path,
            "source_path": sibling[0],
        }

    conn.close()
    raise HTTPException(status_code=404, detail="Snapshot source not found")


@app.get("/attachments/{attachment_id}/original")
def get_original_attachment(attachment_id: int):
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("SELECT moved_to, original_filename, filename FROM imported_files WHERE id = ?", (attachment_id,))
    row = cur.fetchone()

    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Attachment not found")

    moved_to, original_filename, stored_filename = row
    candidate_paths = [moved_to]

    if original_filename:
        candidate_paths.append(os.path.join(EMAILS_FOLDER, original_filename))
    if stored_filename:
        candidate_paths.append(os.path.join(EMAILS_FOLDER, stored_filename))

    for candidate_path in candidate_paths:
        if candidate_path and os.path.exists(candidate_path):
            conn.close()
            return FileResponse(candidate_path)

    conn.close()
    raise HTTPException(status_code=404, detail="Original file not found")


@app.get("/keyproof/{attachment_id}")
def get_keyproof_saved_state(attachment_id: int):
    conn = get_conn()
    conn.row_factory = sqlite3.Row
    try:
        _ensure_attachment_exists(conn, attachment_id)
        return _load_saved_payload(conn, "keyproof", attachment_id)
    finally:
        conn.close()


@app.put("/keyproof/{attachment_id}")
def save_keyproof_saved_state(attachment_id: int, payload: dict):
    if not isinstance(payload, dict):
        raise HTTPException(status_code=400, detail="Keyproof payload must be an object")

    conn = get_conn()
    conn.row_factory = sqlite3.Row
    try:
        _ensure_attachment_exists(conn, attachment_id)
        saved = _save_saved_payload(conn, "keyproof", attachment_id, payload)
        conn.execute(
            f'''
            UPDATE {_quote_identifier("imported_files")}
            SET {_quote_identifier("amount")} = ?
            WHERE {_quote_identifier("id")} = ?
            ''',
            (_keyproof_total_from_payload(payload), attachment_id),
        )
        conn.commit()
        return saved
    finally:
        conn.close()


@app.delete("/keyproof/{attachment_id}")
def delete_keyproof_saved_state(attachment_id: int):
    conn = get_conn()
    try:
        _ensure_attachment_exists(conn, attachment_id)
        return _delete_saved_payload(conn, "keyproof", attachment_id)
    finally:
        conn.close()


@app.get("/itemization/{attachment_id}")
def get_itemization_saved_state(attachment_id: int):
    conn = get_conn()
    conn.row_factory = sqlite3.Row
    try:
        _ensure_attachment_exists(conn, attachment_id)
        return _load_saved_payload(conn, "itemization", attachment_id)
    finally:
        conn.close()


@app.put("/itemization/{attachment_id}")
def save_itemization_saved_state(attachment_id: int, payload: dict):
    if not isinstance(payload, dict):
        raise HTTPException(status_code=400, detail="Itemization payload must be an object")

    conn = get_conn()
    conn.row_factory = sqlite3.Row
    try:
        _ensure_attachment_exists(conn, attachment_id)
        return _save_saved_payload(conn, "itemization", attachment_id, payload)
    finally:
        conn.close()


@app.delete("/itemization/{attachment_id}")
def delete_itemization_saved_state(attachment_id: int):
    conn = get_conn()
    try:
        _ensure_attachment_exists(conn, attachment_id)
        return _delete_saved_payload(conn, "itemization", attachment_id)
    finally:
        conn.close()


@app.get("/keyproof/flywire/{attachment_id}")
def get_keyproof_flywire(attachment_id: int):
    conn = get_conn()
    conn.row_factory = sqlite3.Row
    try:
        document = conn.execute(
            f'''
            SELECT *
            FROM {_quote_identifier("Import_FlywireDocuments")}
            WHERE {_quote_identifier("attachment_id")} = ?
            ORDER BY {_quote_identifier("id")} DESC
            LIMIT 1
            ''',
            (attachment_id,),
        ).fetchone()

        if not document:
            return {"document": None, "summary": None, "rows": []}

        rows = conn.execute(
            f'''
            SELECT *
            FROM {_quote_identifier("Import_FlywireRows")}
            WHERE {_quote_identifier("document_id")} = ?
            ORDER BY {_quote_identifier("position")} ASC, {_quote_identifier("id")} ASC
            ''',
            (document["id"],),
        ).fetchall()
        return _flywire_document_payload(document, rows)
    finally:
        conn.close()


@app.post("/keyproof/flywire/{attachment_id}")
async def upload_keyproof_flywire(attachment_id: int, file: UploadFile = File(...)):
    file_name = os.path.basename(file.filename or "").strip()
    if not file_name:
        raise HTTPException(status_code=400, detail="Fly Wire filename is required")

    extension = os.path.splitext(file_name)[1].lower()
    if extension not in {".xlsx", ".xlsm", ".xls"}:
        raise HTTPException(status_code=400, detail="Please upload a Fly Wire .xls or .xlsx workbook")

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Fly Wire workbook is empty")

    conn = get_conn()
    conn.row_factory = sqlite3.Row
    try:
        attachment_row = conn.execute(
            f'''
            SELECT id, filename, batch_id, batch_date
            FROM {_quote_identifier("imported_files")}
            WHERE {_quote_identifier("id")} = ?
            ''',
            (attachment_id,),
        ).fetchone()
        if not attachment_row:
            raise HTTPException(status_code=404, detail="Attachment not found")
        return _import_flywire_document(conn, attachment_row, file_name, file_bytes)
    finally:
        conn.close()


@app.post("/keyproof/flywire/{attachment_id}/autofind")
def autofind_keyproof_flywire(attachment_id: int):
    conn = get_conn()
    conn.row_factory = sqlite3.Row
    try:
        attachment_row = conn.execute(
            f'''
            SELECT id, filename, batch_id, batch_date
            FROM {_quote_identifier("imported_files")}
            WHERE {_quote_identifier("id")} = ?
            ''',
            (attachment_id,),
        ).fetchone()
        if not attachment_row:
            raise HTTPException(status_code=404, detail="Attachment not found")

        batch_id, batch_date = _flywire_attachment_context(attachment_row)
        candidate_path = _find_flywire_candidate_path(batch_id, batch_date)
        if not candidate_path or not os.path.exists(candidate_path):
            raise HTTPException(status_code=404, detail="No matching Fly Wire file was found in the configured email folder")

        with open(candidate_path, "rb") as input_file:
            file_bytes = input_file.read()
        file_name = os.path.basename(candidate_path)
        return _import_flywire_document(conn, attachment_row, file_name, file_bytes)
    finally:
        conn.close()


@app.delete("/keyproof/flywire/{attachment_id}")
def delete_keyproof_flywire(attachment_id: int):
    conn = get_conn()
    try:
        _remove_existing_flywire_document(conn, attachment_id)
        conn.commit()
        return {"ok": True}
    finally:
        conn.close()


# ------------------------------------------------------------
# APPROVE FILE
# ------------------------------------------------------------
@app.post("/attachments/{attachment_id}/approve")
def approve_attachment(attachment_id: int):
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        UPDATE imported_files
        SET review_status = 'Approved'
        WHERE id = ?
    """, (attachment_id,))

    conn.commit()
    conn.close()

    return {"status": "approved", "id": attachment_id}


# ------------------------------------------------------------
# REJECT FILE
# ------------------------------------------------------------
@app.post("/attachments/{attachment_id}/reject")
def reject_attachment(attachment_id: int):
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        UPDATE imported_files
        SET review_status = 'Rejected'
        WHERE id = ?
    """, (attachment_id,))

    conn.commit()
    conn.close()

    return {"status": "rejected", "id": attachment_id}


@app.post("/attachments/{attachment_id}/restore-pending")
def restore_attachment_pending(attachment_id: int):
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        UPDATE imported_files
        SET review_status = 'Pending'
        WHERE id = ?
    """, (attachment_id,))

    conn.commit()
    conn.close()

    return {"status": "pending", "id": attachment_id}


@app.get("/auth/roles")
def list_roles():
    conn = get_conn()
    try:
        ensure_auth_tables(conn)
        rows = conn.execute(
            f'SELECT * FROM {_quote_identifier("roles")} ORDER BY {_quote_identifier("name")} ASC'
        ).fetchall()
        return [_role_row_to_payload(row) for row in rows]
    finally:
        conn.close()


@app.post("/auth/roles")
def create_role(role: dict):
    name = str(role.get("name") or "").strip()
    description = str(role.get("description") or "").strip()
    permissions_json = _json_permissions(role.get("permissions", []))

    if not name:
        raise HTTPException(status_code=400, detail="name is required")

    conn = get_conn()
    try:
        ensure_auth_tables(conn)
        now = _utc_now_iso()
        cur = conn.cursor()
        cur.execute(
            f'INSERT INTO {_quote_identifier("roles")} ({_quote_identifier("name")}, {_quote_identifier("description")}, {_quote_identifier("permissions_json")}, {_quote_identifier("is_system")}, {_quote_identifier("active")}, {_quote_identifier("created_at")}, {_quote_identifier("updated_at")}) VALUES (?, ?, ?, 0, ?, ?, ?)',
            (name, description, permissions_json, 1 if role.get("active", True) else 0, now, now),
        )
        conn.commit()
        created = cur.execute(
            f'SELECT * FROM {_quote_identifier("roles")} WHERE id = ?',
            (cur.lastrowid,),
        ).fetchone()
        return _role_row_to_payload(created)
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=400, detail="Role already exists")
    finally:
        conn.close()


@app.put("/auth/roles/{role_id}")
def update_role(role_id: int, role: dict):
    conn = get_conn()
    try:
        ensure_auth_tables(conn)
        existing = _fetch_role_or_404(conn, role_id)
        name = str(role.get("name") or existing["name"]).strip()
        description = str(role.get("description") or existing["description"]).strip()
        permissions_json = _json_permissions(role.get("permissions", existing["permissions_json"]))
        active = 1 if role.get("active", bool(existing["active"])) else 0

        conn.execute(
            f'UPDATE {_quote_identifier("roles")} SET {_quote_identifier("name")} = ?, {_quote_identifier("description")} = ?, {_quote_identifier("permissions_json")} = ?, {_quote_identifier("active")} = ?, {_quote_identifier("updated_at")} = ? WHERE id = ?',
            (name, description, permissions_json, active, _utc_now_iso(), role_id),
        )
        conn.commit()
        updated = conn.execute(
            f'SELECT * FROM {_quote_identifier("roles")} WHERE id = ?',
            (role_id,),
        ).fetchone()
        return _role_row_to_payload(updated)
    finally:
        conn.close()


@app.get("/auth/users")
def list_users():
    conn = get_conn()
    try:
        ensure_auth_tables(conn)
        rows = conn.execute(
            f"""
            SELECT
                u.id,
                u.signin,
                u.display_name,
                u.role_id,
                u.active,
                u.last_login_at,
                u.created_at,
                u.updated_at,
                r.name AS role_name
            FROM {_quote_identifier("users")} u
            LEFT JOIN {_quote_identifier("roles")} r ON r.id = u.role_id
            ORDER BY u.signin COLLATE NOCASE ASC
            """
        ).fetchall()
        return [_user_row_to_payload(row) for row in rows]
    finally:
        conn.close()


@app.post("/auth/users")
def create_user(user: dict):
    conn = get_conn()
    try:
        ensure_auth_tables(conn)
        user_id = _create_user_record(conn, user)
        conn.commit()
        created = conn.execute(
            f"""
            SELECT
                u.id,
                u.signin,
                u.display_name,
                u.role_id,
                u.active,
                u.last_login_at,
                u.created_at,
                u.updated_at,
                r.name AS role_name
            FROM {_quote_identifier("users")} u
            LEFT JOIN {_quote_identifier("roles")} r ON r.id = u.role_id
            WHERE u.id = ?
            """,
            (user_id,),
        ).fetchone()
        return _user_row_to_payload(created)
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=400, detail="Signin already exists")
    finally:
        conn.close()


@app.put("/auth/users/{user_id}")
def update_user(user_id: int, user: dict):
    conn = get_conn()
    try:
        ensure_auth_tables(conn)
        existing = conn.execute(
            f'SELECT * FROM {_quote_identifier("users")} WHERE id = ?',
            (user_id,),
        ).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="User not found")

        signin = str(user.get("signin") or existing["signin"]).strip()
        display_name = str(user.get("display_name") or existing["display_name"]).strip()
        role_id = int(user.get("role_id") or existing["role_id"])
        active = 1 if user.get("active", bool(existing["active"])) else 0

        _fetch_role_or_404(conn, role_id)

        params = [signin, display_name, role_id, active, _utc_now_iso(), user_id]
        update_clause = [
            f'{_quote_identifier("signin")} = ?',
            f'{_quote_identifier("display_name")} = ?',
            f'{_quote_identifier("role_id")} = ?',
            f'{_quote_identifier("active")} = ?',
            f'{_quote_identifier("updated_at")} = ?',
        ]

        if user.get("password"):
            update_clause.insert(2, f'{_quote_identifier("password_hash")} = ?')
            params.insert(2, _hash_password(user.get("password")))

        conn.execute(
            f'UPDATE {_quote_identifier("users")} SET {", ".join(update_clause)} WHERE id = ?',
            tuple(params),
        )
        conn.commit()
        updated = conn.execute(
            f"""
            SELECT
                u.id,
                u.signin,
                u.display_name,
                u.role_id,
                u.active,
                u.last_login_at,
                u.created_at,
                u.updated_at,
                r.name AS role_name
            FROM {_quote_identifier("users")} u
            LEFT JOIN {_quote_identifier("roles")} r ON r.id = u.role_id
            WHERE u.id = ?
            """,
            (user_id,),
        ).fetchone()
        return _user_row_to_payload(updated)
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=400, detail="Signin already exists")
    finally:
        conn.close()


@app.post("/auth/login")
def login(payload: dict):
    signin = str(payload.get("signin") or "").strip()
    password = str(payload.get("password") or "")
    if not signin or not password:
        raise HTTPException(status_code=400, detail="signin and password are required")

    conn = get_conn()
    try:
        ensure_auth_tables(conn)
        user = _get_user_by_signin(conn, signin)
        if not user or not user["active"]:
            raise HTTPException(status_code=401, detail="Invalid signin or password")
        if not _verify_password(password, user["password_hash"]):
            raise HTTPException(status_code=401, detail="Invalid signin or password")

        now = _utc_now_iso()
        conn.execute(
            f'UPDATE {_quote_identifier("users")} SET {_quote_identifier("last_login_at")} = ?, {_quote_identifier("updated_at")} = ? WHERE id = ?',
            (now, now, user["id"]),
        )
        conn.commit()

        permissions = []
        try:
            permissions = json.loads(user["role_permissions_json"] or "[]")
        except Exception:
            permissions = []

        return {
            "id": user["id"],
            "signin": user["signin"],
            "display_name": user["display_name"],
            "role": {
                "id": user["role_id"],
                "name": user["role_name"],
                "description": user["role_description"],
                "permissions": permissions,
            },
            "permissions": permissions,
        }
    finally:
        conn.close()


@app.post("/auth/bootstrap-admin")
def bootstrap_admin(payload: dict):
    signin = str(payload.get("signin") or "").strip()
    password = str(payload.get("password") or "")
    display_name = str(payload.get("display_name") or "Administrator").strip()

    if not signin or not password:
        raise HTTPException(status_code=400, detail="signin and password are required")

    conn = get_conn()
    try:
        ensure_auth_tables(conn)
        existing_user_count = conn.execute(
            f'SELECT COUNT(*) FROM {_quote_identifier("users")}'
        ).fetchone()[0]
        if existing_user_count:
            raise HTTPException(status_code=409, detail="Bootstrap is only allowed when no users exist")

        admin_role = conn.execute(
            f'SELECT id FROM {_quote_identifier("roles")} WHERE LOWER({_quote_identifier("name")}) = LOWER(?)',
            ("Admin",),
        ).fetchone()
        if not admin_role:
            raise HTTPException(status_code=500, detail="Admin role is missing")

        user_id = _create_user_record(
            conn,
            {
                "signin": signin,
                "password": password,
                "role_id": admin_role[0],
                "display_name": display_name,
                "active": True,
            },
        )
        conn.commit()
        created = conn.execute(
            f"""
            SELECT
                u.id,
                u.signin,
                u.display_name,
                u.role_id,
                u.active,
                u.last_login_at,
                u.created_at,
                u.updated_at,
                r.name AS role_name
            FROM {_quote_identifier("users")} u
            LEFT JOIN {_quote_identifier("roles")} r ON r.id = u.role_id
            WHERE u.id = ?
            """,
            (user_id,),
        ).fetchone()
        return _user_row_to_payload(created)
    finally:
        conn.close()


@app.get("/sites")
def get_sites():
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("SELECT id, name, description, active FROM sites ORDER BY name;")
    rows = cur.fetchall()

    conn.close()

    return [
        {
            "id": row[0],
            "name": row[1],
            "description": row[2],
            "active": row[3],
        }
        for row in rows
    ]


@app.post("/sites")
def add_site(site: dict):
    name = site.get("name")
    description = site.get("description", "")

    if not name:
        raise HTTPException(status_code=400, detail="Site name is required")

    conn = get_conn()
    cur = conn.cursor()

    try:
        cur.execute(
            "INSERT INTO sites (name, description, active) VALUES (?, ?, 1);",
            (name, description),
        )
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        raise HTTPException(status_code=400, detail="Site already exists")

    conn.close()
    return {"status": "ok", "message": "Site added"}


@app.put("/sites/{site_id}")
def update_site(site_id: int, site: dict):
    name = site.get("name")
    description = site.get("description")
    active = site.get("active")

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("SELECT id FROM sites WHERE id = ?;", (site_id,))
    if not cur.fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="Site not found")

    cur.execute(
        "UPDATE sites SET name = ?, description = ?, active = ? WHERE id = ?;",
        (name, description, active, site_id),
    )

    conn.commit()
    conn.close()

    return {"status": "ok", "message": "Site updated"}


@app.delete("/sites/{site_id}")
def delete_site(site_id: int):
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("SELECT id FROM sites WHERE id = ?;", (site_id,))
    if not cur.fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="Site not found")

    cur.execute("DELETE FROM sites WHERE id = ?;", (site_id,))
    conn.commit()
    conn.close()

    return {"status": "ok", "message": "Site deleted"}
